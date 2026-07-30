# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
"""Puente entre los engines y la app de escritorio.

Los engines de `engines/` son UI-agnósticos, así que la app de escritorio los
reusa TAL CUAL (misma lógica que los Colab: mismos selectores, misma zona, mismo
Excel). Este módulo sólo aporta:

  1. Una API uniforme para las 3 herramientas: `run(params, emit) -> ruta_excel`.
  2. La traducción de los eventos de cada engine a un stream único (`emit`) que
     el servidor manda al frontend por SSE.
  3. El loop tienda×subcategoría de Maestra Sección, que en Colab vive dentro
     del launcher de ipywidgets y por eso no era reusable.

DIFERENCIAS A PROPÓSITO vs Colab (son la razón de ser de la app):
  - **Sin checkpoints**: no hay VM efímera que se caiga, así que no se monta
    Drive ni se persisten parciales. Un corte se reintenta y ya.
  - **Paralelismo agresivo**: corriendo desde el PC del usuario la IP es
    residencial (Cloudflare es mucho menos hostil que con las IPs de Google) y
    no hay límite de sesión, así que Fast usa más workers que en Colab.
"""
from __future__ import annotations

import asyncio
import re
from datetime import datetime
from pathlib import Path

# Paralelismo local de Fast. Se bajó de 12→8: con 12 (y peor si corría junto a
# otros jobs Sodimac) Cloudflare empezaba a devolver challenges y se acumulaban
# tandas de "error de red persistente" (categorías incompletas). 8 es el número
# que el Colab probó estable. Ver también el backoff de _fetch_page en el engine.
FAST_WORKERS = 8
# Presupuesto TOTAL de workers de Fast repartido entre los jobs Fast activos
# (el server calcula params["_workers"] al lanzar). Evita que N× Fast disparen
# demasiadas requests concurrentes a Sodimac → 429/bloqueo de Cloudflare.
FAST_WORKER_BUDGET = 8

# UA propio: `maestra_sodimac` no exporta USER_AGENT (sí lo hace sodimac_engine).
USER_AGENT = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe(name: str) -> str:
    return re.sub(r"[^\w\s-]", "", str(name)).strip().replace(" ", "_") or "salida"


def _outname(base: str, tag: str) -> str:
    """Nombre de Excel único por job: base_TIMESTAMP[_tag].xlsx.
    El `tag` (id corto del job) evita colisiones cuando 2 jobs de la misma
    herramienta terminan en el mismo segundo."""
    return f"{base}_{_ts()}" + (f"_{tag}" if tag else "") + ".xlsx"


def _shots_dir(outdir: Path, tag: str):
    """Carpeta de screenshots aislada por job (no se pisan entre paralelos)."""
    d = outdir / "_shots" / (tag or "run")
    d.mkdir(parents=True, exist_ok=True)
    return d


def _cleanup_shots(shot_dir):
    """Borra la carpeta temporal de screenshots del job tras escribir el Excel: las
    imágenes YA quedaron embebidas en el .xlsx, no se necesitan más. Sin esto,
    Documents/Cruzer/_shots crecería sin control con cada scrape que use fotos."""
    if not shot_dir:
        return
    try:
        import shutil
        shutil.rmtree(shot_dir, ignore_errors=True)
    except Exception:  # noqa: BLE001
        pass


_REPO_ROOT = Path(__file__).resolve().parent.parent  # desktop/ → raíz del repo


def _colab_mk7_template():
    """Devuelve los BYTES EXACTOS del formato de carga del Colab (MK7).

    Fuente única de verdad: el blob base64 embebido en launchers/mk7.py
    (`_FORMATO_CARGA_B64`). Así el desktop entrega el MISMO archivo que el
    Colab (5 columnas: SKU Easy · Desc. Producto · SKU Sodimac · SKU Falabella
    · SKU Construmart) y nunca diverge. Lanza si no lo encuentra (para no
    servir en silencio un formato distinto).
    """
    import re
    src = (_REPO_ROOT / "launchers" / "mk7.py").read_text(encoding="utf-8")
    m = re.search(r'_FORMATO_CARGA_B64\s*=\s*"([^"]+)"', src)
    if not m:
        raise RuntimeError("No encontré _FORMATO_CARGA_B64 en launchers/mk7.py")
    import base64
    return base64.b64decode(m.group(1))


def build_template_bytes(tool: str):
    """Devuelve (nombre_archivo, bytes .xlsx) del 'formato de carga'.

    - mk7: bytes EXACTOS del template del Colab (5 columnas), no una copia.
    - ferni_sku: mismo diseño que el _download_formato del Colab (3 columnas,
      ejemplos de puertas): cabecera #334E68, freeze A2, SKU como texto.
    """
    if tool != "ferni_sku":  # mk7
        return "formato_carga.xlsx", _colab_mk7_template()

    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    rows = [["SKU Easy", "Desc. Producto", "SKU Sodimac"],
            ["E001", "Puerta Madera Terciada Carpintera 90x200 (ejemplo)", "139566229"],
            ["E002", "Puerta MDF Milano 60x200 (ejemplo, oferta)", "120822458"],
            ["E003", "Puerta Madera Terciada Carpintera 75x200 (ejemplo)", "139566225"]]
    wb = Workbook(); ws = wb.active; ws.title = "SKUs"
    for r in rows:
        ws.append(r)
    fill = PatternFill("solid", fgColor="334E68")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for c in range(1, 4):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill; cell.font = font; cell.alignment = align
    for i, w in enumerate((12, 55, 16), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for ci in (1, 3):
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row:
                cell.number_format = "@"
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf)
    return "formato_carga_puertas.xlsx", buf.getvalue()


# ── 1. MK7 — Buscador por SKU ───────────────────────────────────────────────
def _post_maestra_async(rows, fuente):
    """Sube las filas a la Maestra Sodimac (Google Sheet auto-actualizada) en un
    hilo aparte: no bloquea ni puede romper el job. No-op si la URL no está seteada."""
    try:
        import threading
        from engines import _maestra_post as _mp
        threading.Thread(target=_mp.post_maestra, args=(list(rows), fuente), daemon=True).start()
    except Exception:  # noqa: BLE001
        pass


def _page_progress_cb(emit, store_name, sc_name):
    """Tercera barra 'Páginas': solo emite cuando el paginador YA reveló el total
    (page_progress_cb da (page_num, total_pages); total_pages es None hasta que se
    detecta). Así la barra aparece "cuando corresponde" con x/n real."""
    def cb(page_num, total_pages):
        if total_pages:
            emit({"type": "progress", "phase": "pagina", "done": page_num, "total": total_pages,
                  "msg": f"{store_name} · {sc_name} · pág {page_num}/{total_pages}"})
    return cb


def _mk7_positional_progress(emit, total_stores):
    """Adapta el progress_cb POSICIONAL de Falabella/Construmart
    (i, total, store, count, rows_so_far) al stream de eventos del desktop."""
    def cb(i, total, store, count, rows_so_far):
        tot = total or total_stores
        emit({"type": "progress", "phase": "zona", "done": max(0, i - 1), "total": tot,
              "msg": f"Tienda {i}/{tot}: {(store or {}).get('name', '')}"})
        emit({"type": "count", "rows": rows_so_far or 0})
    return cb


async def _run_mk7_generic(params, emit, outdir, tag, mod, out_base, label):
    """MK7 para Falabella/Construmart: mismo contrato de engine
    (read_input / search_skus(metas, stores) / write_output(rows, path) / ALL_STORES)."""
    stores = [s for s in mod.ALL_STORES if s["id"] in set(params["store_ids"])]
    if not stores:
        raise ValueError("No seleccionaste ninguna tienda.")
    df, desc_col, sku_col, easy_col = mod.read_input(params["input_path"])
    has_easy, has_desc = easy_col in df.columns, desc_col in df.columns
    metas, seen = [], set()
    for _, r in df.iterrows():
        sku = str(r[sku_col]).strip()
        if not sku or sku in seen:
            continue
        seen.add(sku)
        metas.append({"sku": sku,
                      "easy": str(r[easy_col]).strip() if has_easy else "",
                      "desc": str(r[desc_col]).strip() if has_desc else ""})
    if not metas:
        raise ValueError("El archivo no tiene SKUs válidos.")

    run_id = params.get("resume_run_id")
    prior_rows, prior_done = [], set()
    if run_id:
        emit({"type": "info", "msg": "Reanudando proceso anterior…"})
        prior_rows, prior_done = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
    else:
        run_id = _ckpts.create_run_id()
        _ckpts.write_meta(CHECKPOINT_DIR, run_id, {
            "tool": "mk7", "section_name": out_base, "store_ids": params["store_ids"]
        })
        _ckpts.ensure_jsonl(CHECKPOINT_DIR, run_id)

    all_rows = list(prior_rows)
    def _on_row(r):
        all_rows.append(r)
        _ckpts.append_row(CHECKPOINT_DIR, run_id, r)

    # El engine de falabella/construmart soporta on_row y skip_pairs
    skip_pairs = {(r.get("Zona") or r.get("Tienda"), r.get("SKU")) for r in all_rows}

    emit({"type": "info", "msg": f"{len(metas)} SKUs · {len(stores)} tienda(s) · {label}"})
    try:
        await mod.search_skus(
            metas, stores, screenshot=bool(params.get("screenshots")), headless=True,
            progress_cb=_mk7_positional_progress(emit, len(stores)),
            on_row=_on_row, skip_pairs=skip_pairs)
    except Exception as e:
        emit({"type": "error", "msg": str(e)})
        raise

    if not all_rows:
        raise RuntimeError(f"No se encontró ningún SKU en {label} para las tiendas elegidas.")
    out = outdir / _outname(out_base, tag)
    emit({"type": "count", "rows": len(all_rows)})
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    mod.write_output(all_rows, str(out))
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    return out


async def run_mk7(params, emit, outdir: Path, tag: str = ""):
    """params: {retailer, input_path, store_ids, screenshots}. Devuelve la ruta del Excel."""
    retailer = params.get("retailer", "sodimac")
    if retailer == "falabella":
        from engines import falabella_engine as _fe
        return await _run_mk7_generic(params, emit, outdir, tag, _fe, "MK7_Falabella", "Falabella")
    if retailer == "construmart":
        from engines import construmart_engine as _ce
        return await _run_mk7_generic(params, emit, outdir, tag, _ce, "MK7_Construmart", "Construmart")
    from engines import sodimac_engine as se

    stores = [s for s in se.ALL_STORES if s["id"] in set(params["store_ids"])]
    if not stores:
        raise ValueError("No seleccionaste ninguna tienda.")

    df, desc_col, sku_col, easy_col = se.read_input(params["input_path"])
    skus = [str(v).strip() for v in df[sku_col].tolist() if str(v).strip()]
    skus = [s for s in dict.fromkeys(skus) if s.isdigit()]  # dedup + sólo numéricos
    if not skus:
        raise ValueError("El archivo no tiene SKUs numéricos válidos.")

    emit({"type": "info", "msg": f"{len(skus)} SKUs · {len(stores)} tienda(s)"})
    shots = _shots_dir(outdir, tag) if params.get("screenshots") else None
    
    run_id = params.get("resume_run_id")
    prior_rows, prior_done = [], set()
    if run_id:
        emit({"type": "info", "msg": "Reanudando proceso anterior…"})
        prior_rows, prior_done = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
    else:
        run_id = _ckpts.create_run_id()
        _ckpts.write_meta(CHECKPOINT_DIR, run_id, {
            "tool": "mk7", "section_name": "MK7_Sodimac", "store_ids": params["store_ids"]
        })
        _ckpts.ensure_jsonl(CHECKPOINT_DIR, run_id)

    all_rows = list(prior_rows)
    def _on_match(m):
        all_rows.append(m)
        _ckpts.append_row(CHECKPOINT_DIR, run_id, m)
    
    # Sodimac engine usa skip_store_ids para evitar tiendas, y procesa zonas completas
    # Wait, skip_store_ids is available in sodimac_engine.search_skus_mk6
    # Y on_match is available
    skip_store_ids = {s["id"] for s in stores if s["id"] in prior_done}

    def _sodimac_wrap(ev):
        _zone_progress(emit, len(stores), {"zone": 0})(ev)
        if ev.get("event") == "zone_end":
            st = ev.get("store") or {}
            sid = st.get("id")
            if sid and not ev.get("zone_failed"):
                _ckpts.append_done(CHECKPOINT_DIR, run_id, sid, "ALL")

    try:
        await se.search_skus_mk6(
            skus, stores, headless=True, screenshot_dir=(str(shots) if shots else None),
            progress_cb=_sodimac_wrap, on_match=_on_match, skip_store_ids=skip_store_ids)
    except Exception as e:
        emit({"type": "error", "msg": str(e)})
        raise

    if not all_rows:
        raise RuntimeError("No se encontró ningún SKU en las tiendas seleccionadas.")

    out = outdir / _outname("MK7_Sodimac", tag)
    emit({"type": "count", "rows": len(all_rows)})
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    se.write_output(df, desc_col, sku_col, easy_col, all_rows, str(out), stores=stores)
    _post_maestra_async(all_rows, "MK7")  # consolidar en la Maestra Sodimac
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    _cleanup_shots(shots)
    return out


# ── 2. Maestra Sección ──────────────────────────────────────────────────────
class _NullProg:
    """Los engines esperan un objeto de progreso estilo rich; acá no aplica."""
    def update(self, *a, **k): pass
    def advance(self, *a, **k): pass


async def discover_sections_desktop(emit, include_landing=True, retailer="sodimac"):
    """Descubre el árbol de secciones (abre el navegador una vez) para el retailer.

    `include_landing=False` para Ferni: su motor deduplica sólo DENTRO de cada
    subcategoría, así que las entradas "Todo X" (roll-up) le duplicarían filas.
    Falabella/Construmart no aceptan include_landing (descubridor propio).
    """
    from playwright.async_api import async_playwright
    from playwright_stealth import Stealth

    emit({"type": "info", "msg": "Abriendo navegador y descubriendo secciones…"})
    async with Stealth().use_async(async_playwright()) as pw:
        b = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
        try:
            ctx = await b.new_context(user_agent=USER_AGENT,
                                      viewport={"width": 1280, "height": 900})
            pg = await ctx.new_page()
            if retailer == "falabella":
                from engines import maestra_falabella as m
                tree = await m.discover_sections(pg)
            elif retailer == "construmart":
                from engines import maestra_construmart as m
                tree = await m.discover_sections(pg)
            else:
                from engines import maestra_sodimac as m
                tree = await m.discover_sections(pg, include_landing=include_landing)
        finally:
            await b.close()
    # subcats pueden ser 2- o 3-tuplas (Falabella) → tomamos nombre y URL.
    return [{"section": sec,
             "subcats": [{"name": s[0], "url": s[1]} for s in subs]}
            for sec, subs in tree]


def stores_for_retailer(retailer):
    """Lista estática de tiendas (Sodimac/Falabella = 42 Easy). Construmart es
    dinámico → discover_stores_desktop (async)."""
    if retailer == "falabella":
        from engines import maestra_falabella as m
    else:
        from engines import maestra_sodimac as m
    return [{"id": s["id"], "name": s["name"], "region": s["region"],
             "comuna": s.get("comuna", "")} for s in m.ALL_STORES]


async def discover_stores_desktop(retailer="construmart"):
    """Descubre las tiendas de Construmart (popup del sitio). Abre el navegador."""
    from playwright.async_api import async_playwright
    from playwright_stealth import Stealth
    from engines import maestra_construmart as mc
    async with Stealth().use_async(async_playwright()) as pw:
        b = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
        try:
            ctx = await b.new_context(user_agent=USER_AGENT, viewport={"width": 1280, "height": 900})
            pg = await ctx.new_page()
            stores = await mc.discover_stores(pg)
        finally:
            await b.close()
    return [{"id": s["id"], "name": s["name"], "region": s.get("region", ""),
             "comuna": s.get("comuna", "")} for s in stores]


async def run_seccion(params, emit, outdir: Path, tag: str = ""):
    """params: {retailer, section, subcats:[{name,url}], store_ids, ...}."""
    retailer = params.get("retailer", "sodimac")
    if retailer == "falabella":
        return await _run_seccion_falabella(params, emit, outdir, tag)
    if retailer == "construmart":
        return await _run_seccion_construmart(params, emit, outdir, tag)
    from playwright.async_api import async_playwright
    from playwright_stealth import Stealth
    from engines import maestra_sodimac as ms

    stores = [s for s in ms.ALL_STORES if s["id"] in set(params.get("store_ids", []))]
    subcats = [(s["name"], s["url"]) for s in params.get("subcats", [])]
    section_name = params.get("section") or "Sección"
    only_sod = not params.get("include_non_sodimac")
    shots = params.get("screenshots", False)

    from engines import _checkpoints as _ckpts
    CHECKPOINT_DIR = outdir / "_checkpoints"
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
    
    run_id = params.get("resume_run_id")
    prior_rows = []
    prior_done = set()
    
    if run_id:
        prior_rows = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
        prior_done = _ckpts.read_done(CHECKPOINT_DIR, run_id)
        _ckpts.touch_run(CHECKPOINT_DIR, run_id)
        meta = _ckpts.read_meta(CHECKPOINT_DIR, run_id) or {}
        if meta:
            stores = [s for s in ms.ALL_STORES if s["id"] in set(meta.get("store_ids", []))]
            subcats = [(s["name"], s["url"]) for s in meta.get("subcats", [])]
            section_name = meta.get("section_name", section_name)
            only_sod = meta.get("only_sod", only_sod)
            shots = meta.get("shots", shots)
    else:
        run_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{_safe(section_name)[:20]}"
        _ckpts.start_run(CHECKPOINT_DIR, run_id, {
            "tool": "seccion",
            "retailer": retailer,
            "section_name": section_name,
            "store_ids": [s["id"] for s in stores],
            "subcats": [{"name": n, "url": u} for n, u in subcats],
            "only_sod": only_sod,
            "shots": shots,
        })

    if not stores or not subcats:
        raise ValueError("Falta seleccionar tienda(s) o subcategoría(s).")

    # PARALLEL-SAFE: dir por-job pasado como parámetro a scrape_subcat, NO por la
    # global ms.SCREENSHOT_DIR (que 2 jobs de Sección se pisarían).
    shot_dir = _shots_dir(outdir, tag) if shots else None

    all_rows = list(prior_rows)
    n_st, n_sub = len(stores), len(subcats)
    prog = _NullProg()

    async with Stealth().use_async(async_playwright()) as pw:
        b = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
        try:
            ctx = await b.new_context(user_agent=USER_AGENT,
                                      viewport={"width": 1280, "height": 900},
                                      color_scheme="light")
            page = await ctx.new_page()
            for si, store in enumerate(stores, 1):
                emit({"type": "progress", "phase": "zona", "done": si, "total": n_st, "msg": store["name"]})
                emit({"type": "info", "msg": f"Fijando zona: {store['name']}…"})
                ok = await ms.set_zone_with_retry(page, store["region"], store["comuna"])
                if not ok:
                    emit({"type": "warn", "msg": f"No se pudo fijar zona en {store['name']}, se salta."})
                    continue
                # dedup por SKU dentro de cada tienda (igual que el Colab)
                seen = {r["SKU"] for r in all_rows
                        if r.get("Tienda") == store["id"] and r.get("SKU")}
                for sj, (sc_name, sc_url) in enumerate(subcats, 1):
                    if (store["id"], sc_name) in prior_done:
                        emit({"type": "info", "msg": f"Saltando {store['name']} · {sc_name} (ya completado)"})
                        continue

                    emit({"type": "progress", "phase": "subcat", "done": sj, "total": n_sub,
                          "msg": f"{store['name']} · {section_name} · {sc_name}",
                          "frac": ((si - 1) * n_sub + sj) / (n_st * n_sub)})
                    res = await ms.scrape_subcat(
                        page, section_name, sc_name, sc_url, prog, None,
                        capture_screenshots=shots, only_sodimac=only_sod,
                        auto_breadcrumb=(section_name == "Custom"),
                        screenshot_dir=shot_dir,
                        page_progress_cb=_page_progress_cb(emit, store["name"], sc_name))
                    for r in res.get("rows", []):
                        if only_sod and "SODIMAC" not in (r.get("Vendedor") or "").upper():
                            continue
                        sku = r.get("SKU")
                        if not sku or sku in seen:
                            continue
                        seen.add(sku)
                        r2 = {"Tienda": store["id"],
                                         "Nombre Tienda": store["name"], **r}
                        all_rows.append(r2)
                        _ckpts.append_row(CHECKPOINT_DIR, run_id, r2)
                    _ckpts.append_done(CHECKPOINT_DIR, run_id, store["id"], sc_name)
                    emit({"type": "count", "rows": len(all_rows)})
        finally:
            await b.close()

    if not all_rows:
        raise RuntimeError("No se obtuvo ninguna fila.")
    out = outdir / _outname(f"Seccion_{_safe(section_name)}", tag)
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    ms.write_excel(all_rows, str(out), with_images=shots)
    _post_maestra_async(all_rows, "Maestra")  # consolidar en la Maestra Sodimac
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    _cleanup_shots(shot_dir)
    return out


# ── 2b. Maestra Falabella / Construmart ─────────────────────────────────────
async def _run_seccion_falabella(params, emit, outdir, tag=""):
    """Igual a Sodimac pero por ZONA (region/comuna) con maestra_falabella."""
    from playwright.async_api import async_playwright
    from playwright_stealth import Stealth
    from engines import maestra_falabella as mf

    stores = [s for s in mf.ALL_STORES if s["id"] in set(params["store_ids"])]
    subcats = [(s["name"], s["url"]) for s in params["subcats"]]
    if not stores or not subcats:
        raise ValueError("Falta seleccionar tienda(s) o subcategoría(s).")
    section_name = params.get("section") or "Sección"
    only_fa = not params.get("include_non_sodimac")  # "solo Falabella" (mismo toggle)
    shots = params.get("screenshots", False)
    
    run_id = params.get("resume_run_id")
    prior_rows, prior_done = [], set()
    if run_id:
        emit({"type": "info", "msg": "Reanudando proceso anterior…"})
        prior_rows, prior_done = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
    else:
        run_id = _ckpts.create_run_id()
        _ckpts.write_meta(CHECKPOINT_DIR, run_id, {
            "tool": "seccion", "section_name": section_name,
            "subcats": [s[0] for s in subcats], "store_ids": params["store_ids"]
        })
        _ckpts.ensure_jsonl(CHECKPOINT_DIR, run_id)

    all_rows = list(prior_rows)
    prog = _NullProg()
    n_st, n_sub = len(stores), len(subcats)

    async with Stealth().use_async(async_playwright()) as pw:
        b = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
        try:
            ctx = await b.new_context(user_agent=USER_AGENT,
                                      viewport={"width": 1280, "height": 900}, color_scheme="light")
            page = await ctx.new_page()
            for si, store in enumerate(stores, 1):
                emit({"type": "progress", "phase": "zona", "done": si, "total": n_st, "msg": store["name"]})
                emit({"type": "info", "msg": f"Fijando zona: {store['name']}…"})
                ok = await mf.set_zone_with_retry(page, store["region"], store["comuna"])
                if not ok:
                    emit({"type": "warn", "msg": f"No se pudo fijar zona en {store['name']}, se salta."})
                    continue
                seen = {r["SKU"] for r in all_rows if r.get("Tienda") == store["id"] and r.get("SKU")}
                for sj, (sc_name, sc_url) in enumerate(subcats, 1):
                    if (store["id"], sc_name) in prior_done:
                        emit({"type": "info", "msg": f"Saltando {store['name']} · {sc_name} (ya completado)"})
                        continue

                    emit({"type": "progress", "phase": "subcat", "done": sj, "total": n_sub,
                          "msg": f"{store['name']} · {section_name} · {sc_name}",
                          "frac": ((si - 1) * n_sub + sj) / (n_st * n_sub)})
                    res = await mf.scrape_subcat(page, section_name, sc_name, sc_url, prog, None,
                                                 download_images=shots, only_falabella=only_fa,
                                                 page_progress_cb=_page_progress_cb(emit, store["name"], sc_name))
                    for r in res.get("rows", []):
                        sku = r.get("SKU")
                        if not sku or sku in seen:
                            continue
                        seen.add(sku)
                        r2 = {"Tienda": store["id"], "Nombre Tienda": store["name"], **r}
                        all_rows.append(r2)
                        _ckpts.append_row(CHECKPOINT_DIR, run_id, r2)
                    _ckpts.append_done(CHECKPOINT_DIR, run_id, store["id"], sc_name)
                    emit({"type": "count", "rows": len(all_rows)})
        finally:
            await b.close()

    if not all_rows:
        raise RuntimeError("No se obtuvo ninguna fila.")
    out = outdir / _outname(f"Seccion_Falabella_{_safe(section_name)}", tag)
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    mf.write_excel(all_rows, str(out), with_images=shots)
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    return out


async def _run_seccion_construmart(params, emit, outdir, tag=""):
    """Construmart: por TIENDA (discover_stores dinámico + set_store). Las filas ya
    traen Tienda/Nombre Tienda desde el engine."""
    from playwright.async_api import async_playwright
    from playwright_stealth import Stealth
    from engines import maestra_construmart as mc

    subcats = [(s["name"], s["url"]) for s in params["subcats"]]
    if not subcats:
        raise ValueError("Falta seleccionar subcategoría(s).")
    section_name = params.get("section") or "Sección"
    shots = params.get("screenshots", False)
    wanted = set(params.get("store_ids") or [])
    
    run_id = params.get("resume_run_id")
    prior_rows, prior_done = [], set()
    if run_id:
        emit({"type": "info", "msg": "Reanudando proceso anterior…"})
        prior_rows, prior_done = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
    else:
        run_id = _ckpts.create_run_id()
        _ckpts.write_meta(CHECKPOINT_DIR, run_id, {
            "tool": "seccion", "section_name": section_name,
            "subcats": [s[0] for s in subcats], "store_ids": params["store_ids"]
        })
        _ckpts.ensure_jsonl(CHECKPOINT_DIR, run_id)

    all_rows = list(prior_rows)
    prog = _NullProg()

    async with Stealth().use_async(async_playwright()) as pw:
        b = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
        try:
            ctx = await b.new_context(user_agent=USER_AGENT,
                                      viewport={"width": 1280, "height": 900}, color_scheme="light")
            page = await ctx.new_page()
            all_st = await mc.discover_stores(page)
            stores = [s for s in all_st if s["id"] in wanted] or all_st[:1]
            n_st, n_sub = len(stores), len(subcats)
            for si, store in enumerate(stores, 1):
                emit({"type": "progress", "phase": "zona", "done": si, "total": n_st, "msg": store["name"]})
                emit({"type": "info", "msg": f"Fijando tienda: {store['name']}…"})
                ok = await mc.set_store_with_retry(page, store["id"], store.get("region"))
                if not ok:
                    emit({"type": "warn", "msg": f"No se pudo fijar {store['name']}, se salta."})
                    continue
                for sj, (sc_name, sc_url) in enumerate(subcats, 1):
                    if (store["id"], sc_name) in prior_done:
                        emit({"type": "info", "msg": f"Saltando {store['name']} · {sc_name} (ya completado)"})
                        continue

                    emit({"type": "progress", "phase": "subcat", "done": sj, "total": n_sub,
                          "msg": f"{store['name']} · {section_name} · {sc_name}",
                          "frac": ((si - 1) * n_sub + sj) / (n_st * n_sub)})
                    res = await mc.scrape_subcat(page, section_name, sc_name, sc_url, store, prog, None,
                                                 download_images=shots)
                    for r in res.get("rows", []):
                        all_rows.append(r)
                        _ckpts.append_row(CHECKPOINT_DIR, run_id, r)
                    _ckpts.append_done(CHECKPOINT_DIR, run_id, store["id"], sc_name)
                    emit({"type": "count", "rows": len(all_rows)})
        finally:
            await b.close()

    if not all_rows:
        raise RuntimeError("No se obtuvo ninguna fila.")
    out = outdir / _outname(f"Seccion_Construmart_{_safe(section_name)}", tag)
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    mc.write_excel(all_rows, str(out), with_images=shots)
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    return out


# ── 3. Fast — Precios por mayor (browserless) ───────────────────────────────
async def run_fast(params, emit, outdir: Path, tag: str = ""):
    """params: {store_ids, sections:[str]|None, url:str|None, wholesale_only, _workers}."""
    from engines import mayoristas_fast as mf
    from engines import maestra_sodimac as ms

    stores = [s for s in ms.ALL_STORES if s["id"] in set(params["store_ids"])]
    if not stores:
        raise ValueError("No seleccionaste ninguna tienda.")
    # _workers lo fija el server: presupuesto total repartido entre los Fast activos.
    workers = int(params.get("_workers") or FAST_WORKERS)
    wholesale_only = params.get("wholesale_only", True)
    url_scope = (params.get("url") or "").strip()
    sections = params.get("sections") or None

    run_id = params.get("resume_run_id")
    prior_rows, prior_done = [], set()
    if run_id:
        emit({"type": "info", "msg": "Reanudando proceso anterior…"})
        prior_rows, prior_done = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
    else:
        run_id = _ckpts.create_run_id()
        _ckpts.write_meta(CHECKPOINT_DIR, run_id, {
            "tool": "fast", "store_ids": params["store_ids"],
            "url": url_scope, "sections": sections
        })
        _ckpts.ensure_jsonl(CHECKPOINT_DIR, run_id)

    all_rows = list(prior_rows)
    report = []
    prog = {"rows": len(all_rows)}  # contador acumulado en vivo (Fast barre 1 tienda ~9 min)
    
    def _on_row_fast(r):
        _ckpts.append_row(CHECKPOINT_DIR, run_id, r)

    for i, store in enumerate(stores, 1):
        emit({"type": "info", "msg": f"[{i}/{len(stores)}] {store['name']}: fijando zona…"})
        if url_scope:
            cookie = await mf.fetch_zone_cookie(store, headless=True)
            tree = [("URL personalizada", [("URL personalizada", url_scope)])]
        else:
            cookie, tree = await mf.open_session(store, headless=True)
        if not cookie or not tree:
            emit({"type": "warn", "msg": f"No se pudo preparar {store['name']}, se salta."})
            continue
        if sections:
            wanted = set(sections)
            tree = [(sec, subs) for sec, subs in tree if sec in wanted]
        
        if prior_done:
            filtered_tree = []
            for sec, subs in tree:
                new_subs = [(nm, u) for nm, u in subs if (store["id"], nm) not in prior_done]
                if new_subs:
                    filtered_tree.append((sec, new_subs))
            tree = filtered_tree

        n_sub = sum(len(s) for _, s in tree)
        nst = len(stores)
        emit({"type": "progress", "phase": "zona", "done": i, "total": nst, "msg": store["name"]})
        if n_sub == 0:
            emit({"type": "info", "msg": f"{store['name']}: completada, saltando…"})
            continue
        emit({"type": "info", "msg": f"{store['name']}: barriendo {n_sub} subcategorías…"})

        def subcat_cb(done, total, sec, name, kept, scanned, status=None, _i=i, _nst=nst, _sid=store["id"]):
            emit({"type": "progress", "phase": "subcat", "done": done, "total": total,
                  "msg": f"{store['name']} · {sec[:20]} · {name[:20]}",
                  "frac": (_i - 1 + done / max(total, 1)) / _nst})
            # filas EN VIVO: acumula lo conservado por subcat (antes el count solo
            # salía al terminar la tienda entera → ~9 min en 0).
            prog["rows"] += kept
            emit({"type": "count", "rows": prog["rows"]})
            _ckpts.append_done(CHECKPOINT_DIR, run_id, _sid, name)

        rows = await asyncio.to_thread(
            mf.scrape_all_wholesale, cookie, tree, store,
            wholesale_only=wholesale_only, only_sodimac=True,
            subcat_cb=subcat_cb, on_row=_on_row_fast, workers=workers, report=report)
        all_rows.extend(rows)
        prog["rows"] = len(all_rows)  # reconcilia con el total real de la tienda
        emit({"type": "count", "rows": len(all_rows)})

    if report:
        emit({"type": "warn",
              "msg": f"{len(report)} categoría(s) no se pudieron leer completas "
                     f"(error de red persistente)."})
    if not all_rows:
        raise RuntimeError("No se obtuvo ninguna fila.")
    kind = "mayoristas" if wholesale_only else "catalogo"
    out = outdir / _outname(f"Fast_{kind}", tag)
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    ms.write_excel(all_rows, str(out), columns=mf.OUTPUT_COLS, with_images=False)
    _post_maestra_async(all_rows, "Fast")  # consolidar en la Maestra Sodimac
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    return out


# ── 4. Ferni — puertas por SKU ──────────────────────────────────────────────
def _zone_progress(emit, total_zones, state):
    """Traductor de eventos compartido: MK7 y Ferni SKU emiten los mismos."""
    def cb(ev):
        e = ev.get("event")
        if e == "zone_start":
            state["zone"] += 1
            emit({"type": "progress", "phase": "zona",
                  "done": state["zone"] - 1, "total": total_zones,
                  "msg": f"Tienda {state['zone']}/{total_zones}: {ev['store']['name']}"})
        elif e == "batch_done":
            emit({"type": "progress", "phase": "lote",
                  "done": ev.get("batches_done_in_zone", 0),
                  "total": ev.get("total_batches_in_zone", 1),
                  "msg": f"{ev['store']['name']} · lote "
                         f"{ev.get('batches_done_in_zone')}/{ev.get('total_batches_in_zone')}"})
            # contador de filas EN VIVO: antes MK7/Ferni SKU mostraban 0 todo el
            # rato (solo emitían progreso, nunca 'count') → parecía que no capturaba.
            state["rows"] = state.get("rows", 0) + ev.get("found_in_batch", 0)
            emit({"type": "count", "rows": state["rows"]})
        elif e == "fallback_start":
            emit({"type": "progress", "phase": "lote",
                  "done": 0, "total": 1,
                  "msg": f"{ev['store']['name']} · buscando {ev.get('n_skus', 0)} "
                         f"SKU(s) en el catálogo…"})
        elif e == "fallback_done" and ev.get("recovered"):
            state["rows"] = state.get("rows", 0) + ev.get("recovered", 0)
            emit({"type": "count", "rows": state["rows"]})
        elif e == "zone_end" and ev.get("zone_failed"):
            emit({"type": "warn", "msg": f"No se pudo fijar zona en {ev['store']['name']}"})
    return cb


async def run_ferni_sku(params, emit, outdir: Path, tag: str = ""):
    """params: {input_path, store_ids, screenshots}. Puertas: 1 fila por medida."""
    from engines import ferni_sodimac as fs

    stores = fs.stores_by_ids(params["store_ids"])
    if not stores:
        raise ValueError("No seleccionaste ninguna tienda.")
    df, desc_col, sku_col, easy_col = fs.read_input(params["input_path"])
    skus = [str(v).strip() for v in df[sku_col].tolist() if str(v).strip()]
    skus = [s for s in dict.fromkeys(skus) if s.isdigit()]
    if not skus:
        raise ValueError("El archivo no tiene SKUs numéricos válidos.")

    emit({"type": "info", "msg": f"{len(skus)} SKUs de puertas · {len(stores)} tienda(s)"})
    shots = params.get("screenshots", False)
    shot_dir = _shots_dir(outdir, tag) if shots else None

    run_id = params.get("resume_run_id")
    prior_rows, prior_done = [], set()
    if run_id:
        emit({"type": "info", "msg": "Reanudando proceso anterior…"})
        prior_rows, prior_done = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
    else:
        run_id = _ckpts.create_run_id()
        _ckpts.write_meta(CHECKPOINT_DIR, run_id, {
            "tool": "ferni_sku", "store_ids": params["store_ids"]
        })
        _ckpts.ensure_jsonl(CHECKPOINT_DIR, run_id)

    all_rows = list(prior_rows)
    def _on_match_ferni_sku(m):
        all_rows.append(m)
        _ckpts.append_row(CHECKPOINT_DIR, run_id, m)

    skip_store_ids = {s["id"] for s in stores if s["id"] in prior_done}
    def _ferni_sku_wrap(ev):
        _zone_progress(emit, len(stores), {"zone": 0})(ev)
        if ev.get("event") == "zone_end":
            st = ev.get("store") or {}
            sid = st.get("id")
            if sid and not ev.get("zone_failed"):
                _ckpts.append_done(CHECKPOINT_DIR, run_id, sid, "ALL")

    try:
        await fs.search_doors(
            skus, stores, headless=True,
            screenshot_dir=(str(shot_dir) if shot_dir else None),
            progress_cb=_ferni_sku_wrap, on_match=_on_match_ferni_sku,
            skip_store_ids=skip_store_ids)
    except Exception as e:
        emit({"type": "error", "msg": str(e)})
        raise
    
    if not all_rows:
        raise RuntimeError("No se encontró ninguna puerta en las tiendas seleccionadas.")

    # Reporte transparente: qué SKUs no aparecieron en ninguna tienda (ni en la
    # búsqueda ni en el listado de categoría) → no están en el catálogo actual.
    found_skus = {str(m.get("sku_input", "")) for m in all_rows}
    not_found = [s for s in skus if s not in found_skus]
    if not_found:
        emit({"type": "warn",
              "msg": (f"{len(found_skus)} de {len(skus)} SKUs encontrados. "
                      f"{len(not_found)} no están en el catálogo actual de Sodimac "
                      f"(descontinuados o sin stock).")})

    out = outdir / _outname("Ferni_SKU", tag)
    emit({"type": "count", "rows": len(all_rows)})
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    fs.write_output(df, desc_col, sku_col, easy_col, all_rows, str(out),
                    stores=stores, embed_images=shots)
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    _cleanup_shots(shot_dir)
    return out


# ── 5. Ferni Sección — puertas y más, por categoría ─────────────────────────
async def run_ferni_seccion(params, emit, outdir: Path, tag: str = ""):
    """params: {section, subcats:[{name,url}], store_ids, screenshots}."""
    from engines import ferni_maestra_sodimac as fm

    from engines import maestra_sodimac as ms
    stores = [s for s in ms.ALL_STORES if s["id"] in set(params["store_ids"])]
    section_name = params.get("section") or "Sección"
    subcats = [(section_name, s["name"], s["url"]) for s in params["subcats"]]
    if not stores or not subcats:
        raise ValueError("Falta seleccionar tienda(s) o subcategoría(s).")

    shots = params.get("screenshots", False)
    shot_dir = _shots_dir(outdir, tag) if shots else None

    run_id = params.get("resume_run_id")
    prior_rows, prior_done = [], set()
    if run_id:
        emit({"type": "info", "msg": "Reanudando proceso anterior…"})
        prior_rows, prior_done = _ckpts.load_rows(CHECKPOINT_DIR, run_id)
    else:
        run_id = _ckpts.create_run_id()
        _ckpts.write_meta(CHECKPOINT_DIR, run_id, {
            "tool": "ferni_seccion", "section_name": section_name, "store_ids": params["store_ids"]
        })
        _ckpts.ensure_jsonl(CHECKPOINT_DIR, run_id)

    all_rows = list(prior_rows)
    def _on_row_ferni_seccion(r):
        all_rows.append(r)
        _ckpts.append_row(CHECKPOINT_DIR, run_id, r)

    done_keys = set()
    for _sid, _sname in prior_done:
        for _sec, _nm, _url in subcats:
            if _nm == _sname:
                done_keys.add((_sid, _url))

    total = len(stores) * len(subcats)
    state = {"n": 0, "rows": len(all_rows)}

    def progress_cb(ev):
        e = ev.get("event")
        st_info = ev.get("store") or {}
        store = st_info.get("name", "")
        sid = st_info.get("id")
        if e == "subcat_start":
            emit({"type": "progress", "phase": "subcat", "done": state["n"], "total": total,
                  "msg": f"{store} · {ev.get('subcat', '')}"})
        elif e == "subcat_done":
            state["n"] += 1
            emit({"type": "progress", "phase": "subcat", "done": state["n"], "total": total,
                  "msg": f"{store} · {ev.get('subcat', '')}"})
            if sid and not ev.get("skipped"):
                _ckpts.append_done(CHECKPOINT_DIR, run_id, sid, ev.get("subcat"))
        elif e == "zone_end" and ev.get("zone_failed"):
            emit({"type": "warn", "msg": f"No se pudo fijar zona en {store}"})
        elif e == "browser_error":
            emit({"type": "warn", "msg": f"Navegador: {ev.get('msg', '')}"})

    emit({"type": "info", "msg": f"{len(subcats)} subcat(s) × {len(stores)} tienda(s)"})
    try:
        await fm.scrape_maestra(
            subcats, stores, headless=True,
            screenshot_dir=(str(shot_dir) if shot_dir else None),
            progress_cb=progress_cb, on_row=_on_row_ferni_seccion, done_keys=done_keys)
    except Exception as e:
        emit({"type": "error", "msg": str(e)})
        raise
    
    if not all_rows:
        raise RuntimeError("No se obtuvo ninguna fila.")

    out = outdir / _outname(f"Ferni_Seccion_{_safe(section_name)}", tag)
    emit({"type": "count", "rows": len(all_rows)})
    emit({"type": "info", "msg": f"Escribiendo Excel ({len(all_rows)} filas)…"})
    fm.write_excel(all_rows, str(out), with_images=shots)
    _ckpts.cleanup_run(CHECKPOINT_DIR, run_id)
    _cleanup_shots(shot_dir)
    return out


TOOLS = {
    "mk7":          {"label": "MK7 · Buscador por SKU",   "run": run_mk7},
    "seccion":      {"label": "Maestra Sección",          "run": run_seccion},
    "fast":         {"label": "Fast · Precios por mayor", "run": run_fast},
    "ferni_sku":    {"label": "Ferni · Puertas por SKU",  "run": run_ferni_sku},
    "ferni_seccion": {"label": "Ferni · Sección",         "run": run_ferni_seccion},
}

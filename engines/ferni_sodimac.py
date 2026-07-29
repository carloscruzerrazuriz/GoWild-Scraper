# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
# === Scrapper Ferni — Buscador de PUERTAS Sodimac por SKU ===

"""Scrapper Ferni — buscador batch de puertas (productos con selector de medidas).

Por qué existe (vs MK7):
  El MK7 extrae el precio del DOM de la card de búsqueda. Para una puerta, esa
  card solo muestra un RANGO ("desde $30.990") porque la puerta es un producto
  CONFIGURABLE: una sola card agrupa N medidas (60x200, 75x200, ...). Resultado:
  las cartas del MK7 salían mal para puertas (precio genérico, sin la medida).

Cómo lo resuelve Ferni:
  Sodimac sirve la grilla con SSR. En <script id="__NEXT_DATA__"> bajo
  props.pageProps.results cada producto trae un array `variants[]`. Para puertas
  hay una entrada `{type:"SIZES", options:[...]}` donde CADA medida es un objeto:
      - variant / offeringId / extraInfo  → el SKU específico de esa medida
      - size / value                       → la medida (ej. "90X200CM")
      - prices[]                           → {crossed:false}=internet, {crossed:true}=normal
      - url                                → link directo a esa variante
      - mediaUrls[]                        → imágenes de esa variante
      - available / isPurchaseable
  Así, buscando el SKU de la medida específica, lo matcheamos contra ese array y
  obtenemos precio EXACTO + medida EXACTA. Es extracción JSON pura: sin selectores
  CSS frágiles ni la carrera precio-vs-DOM que sufría el MK7.

Bonus de eficiencia: varias medidas de la MISMA puerta colapsan en 1 sola card
del buscador, pero TODAS se resuelven desde su array `variants` → más SKUs por
batch que en el MK7.

Reutiliza verbatim la infraestructura probada del MK7: set_zone, warmup,
guards anti-página-en-blanco, orquestador multi-zona, lanzamiento Chromium.
"""

import re
from pathlib import Path

import pandas as pd
from playwright.async_api import async_playwright, Page
from playwright_stealth import Stealth


BASE_URL = "https://www.sodimac.cl/sodimac-cl"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Guards: SKUs Sodimac buscables en (casi) toda zona. Se anteponen a cada query
# para forzar el render de la grilla (una query de 1 SKU puede volver en blanco
# en headless). Idénticos al MK7.
#   110038221 — UBERMANN martillo carpintero 20oz
#   130607328 — BAUKER set destornilladores 31 pzs
#   110229015 — ENERGIZER pilas AA x4
#   110096085 — AQUACHEM Cloro granulado 2kg
DEFAULT_GUARD_SKUS = ["110038221", "130607328", "110229015", "110096085"]

# Puertas colapsan medidas en 1 card → caben más SKUs por página. 16 reales +
# 4 guards = 20 entra cómodo en una PLP (page size ~24).
DEFAULT_BATCH_SIZE = 16

# Fallback de catálogo: el buscador /buscar de Sodimac tiene un índice INCOMPLETO
# (no devuelve todos los SKUs del catálogo), mientras que el listado de categoría
# sí los trae. Cuando un SKU no aparece en la búsqueda, recorremos el listado
# maestro de Puertas y lo matcheamos ahí (mismo extractor, misma zona → precio
# correcto). Recupera SKUs válidos que el buscador se salta.
PUERTAS_CATALOG_URL = f"{BASE_URL}/lista/CATG10743/puertas"
MAX_FALLBACK_PAGES = 12

ALL_STORES = [
    {"id": "E534", "name": "Antofagasta",   "region": "Antofagasta",   "comuna": "Antofagasta"},
    {"id": "E619", "name": "Arica",         "region": "Arica",         "comuna": "Arica"},
    {"id": "E633", "name": "Bio Bio",       "region": "Biobío",        "comuna": "Hualpén"},
    {"id": "E614", "name": "Calama",        "region": "Antofagasta",   "comuna": "Calama"},
    {"id": "E522", "name": "Cerrillos",     "region": "Metropolitana", "comuna": "Cerrillos"},
    {"id": "E988", "name": "Chicureo",      "region": "Metropolitana", "comuna": "Colina"},
    {"id": "E990", "name": "Chiguayante",   "region": "Biobío",        "comuna": "Chiguayante"},
    {"id": "E525", "name": "Chillán",       "region": "Ñuble",         "comuna": "Chillán"},
    {"id": "E760", "name": "Copiapó",       "region": "Atacama",       "comuna": "Copiapó"},
    {"id": "E983", "name": "Coronel",       "region": "Biobío",        "comuna": "Coronel"},
    {"id": "E511", "name": "Costanera",     "region": "Metropolitana", "comuna": "Providencia"},
    {"id": "E592", "name": "Curicó",        "region": "Maule",         "comuna": "Curicó"},
    {"id": "E781", "name": "El Belloto",    "region": "Valparaíso",    "comuna": "Quilpué"},
    {"id": "E513", "name": "El Llano",      "region": "Metropolitana", "comuna": "San Miguel"},
    {"id": "E510", "name": "Florida",       "region": "Metropolitana", "comuna": "La Florida"},
    {"id": "E502", "name": "Kennedy",       "region": "Metropolitana", "comuna": "Las Condes"},
    {"id": "E514", "name": "La Dehesa",     "region": "Metropolitana", "comuna": "Lo Barnechea"},
    {"id": "E512", "name": "La Reina",      "region": "Metropolitana", "comuna": "La Reina"},
    {"id": "E521", "name": "La Serena",     "region": "Coquimbo",      "comuna": "La Serena"},
    {"id": "E744", "name": "La Unión",      "region": "Ríos",          "comuna": "La Unión"},
    {"id": "E524", "name": "Linares",       "region": "Maule",         "comuna": "Linares"},
    {"id": "E900", "name": "Los Andes",     "region": "Valparaíso",    "comuna": "Los Andes"},
    {"id": "E529", "name": "Los Ángeles",   "region": "Biobío",        "comuna": "Los Ángeles"},
    {"id": "E503", "name": "Maipú",         "region": "Metropolitana", "comuna": "Maipú"},
    {"id": "E643", "name": "Ochagavía",     "region": "Metropolitana", "comuna": "Pedro Aguirre Cerda"},
    {"id": "E585", "name": "Osorno",        "region": "Lagos",         "comuna": "Osorno"},
    {"id": "E775", "name": "Portal Ñuñoa",  "region": "Metropolitana", "comuna": "Ñuñoa"},
    {"id": "E748", "name": "Portal Osorno", "region": "Lagos",         "comuna": "Osorno"},
    {"id": "E506", "name": "Portal Temuco", "region": "Araucanía",     "comuna": "Temuco"},
    {"id": "E659", "name": "Puente Alto",   "region": "Metropolitana", "comuna": "Puente Alto"},
    {"id": "E507", "name": "Puerto Montt",  "region": "Lagos",         "comuna": "Puerto Montt"},
    {"id": "E655", "name": "Quilicura",     "region": "Metropolitana", "comuna": "Quilicura"},
    {"id": "E518", "name": "Quilín",        "region": "Metropolitana", "comuna": "Peñalolén"},
    {"id": "E646", "name": "Quillota",      "region": "Valparaíso",    "comuna": "Quillota"},
    {"id": "E504", "name": "Rancagua",      "region": "O'Higgins",     "comuna": "Rancagua"},
    {"id": "E843", "name": "San Bernardo",  "region": "Metropolitana", "comuna": "San Bernardo"},
    {"id": "E874", "name": "Santa Amalia",  "region": "Metropolitana", "comuna": "La Florida"},
    {"id": "E591", "name": "Talca",         "region": "Maule",         "comuna": "Talca"},
    {"id": "E517", "name": "Temuco",        "region": "Araucanía",     "comuna": "Temuco"},
    {"id": "E520", "name": "Valparaíso",    "region": "Valparaíso",    "comuna": "Valparaíso"},
    {"id": "E830", "name": "Villarrica",    "region": "Araucanía",     "comuna": "Villarrica"},
    {"id": "E508", "name": "Viña del Mar",  "region": "Valparaíso",    "comuna": "Viña del Mar"},
]

# Presets de zona (mismos que MK7). Default puertas: Cerrillos (E522).
DEFAULT_STORE_ID = "E522"
RM_STORE_IDS = [s["id"] for s in ALL_STORES if s["region"] == "Metropolitana"]


def stores_by_ids(ids):
    idset = list(ids)
    by_id = {s["id"]: s for s in ALL_STORES}
    return [by_id[i] for i in idset if i in by_id]


# ─────────────────────────────────────────  Zone (autocomplete)  ───────────
# (idéntico al MK7 sodimac_engine — infra probada, no tocar sin testear)

# ── Zona: delegado al sistema ÚNICO compartido (engines/_zone_sodimac.py) ──
# Antes Ferni tenía su copia propia; ahora comparte la misma set_zone robusta
# (verificación cookie+label, backspace-retry, warmup) que MK7 y Sección.
from engines import _zone_sodimac as _zone

warmup_session     = _zone.warmup_session
_type_autocomplete = _zone._type_autocomplete
set_zone           = _zone.set_zone


# ─────────────────────────────────────────  Batch search (puertas)  ────────

async def _fetch_results_json(page: Page, url: str) -> list | None:
    """Navega y devuelve props.pageProps.results (o None). Polling in-page del
    array results del __NEXT_DATA__. Idéntico al MK7."""
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
    except Exception:
        return None
    extract_js = """() => {
        const s = document.getElementById('__NEXT_DATA__');
        if (!s) return {state: 'no-script'};
        try {
            const d = JSON.parse(s.textContent);
            const r = d?.props?.pageProps?.results;
            if (!Array.isArray(r)) return {state: 'no-results-key'};
            return {state: 'ok', results: r};
        } catch (e) { return {state: 'parse-error', err: String(e)}; }
    }"""
    last_results = None
    for _ in range(20):  # ~10s
        try:
            payload = await page.evaluate(extract_js)
        except Exception:
            payload = None
        if payload and payload.get("state") == "ok":
            r = payload.get("results") or []
            last_results = r
            if len(r) > 0:
                return r
        await page.wait_for_timeout(500)
    return last_results


# Extracción de puertas: parsea results y, para cada SKU buscado, resuelve la
# variante (medida) exacta con su precio. A diferencia del MK7, NO mira el DOM:
# todo sale del JSON SSR, que ya trae las variantes desglosadas.
_EXTRACT_DOORS_JS = r"""(searchedSkus) => {
    const s = document.getElementById('__NEXT_DATA__');
    if (!s) return {state: 'no-script', matches: {}};
    let d; try { d = JSON.parse(s.textContent); } catch (e) { return {state: 'parse-error', matches: {}}; }
    const r = d?.props?.pageProps?.results;
    if (!Array.isArray(r)) return {state: 'no-results', matches: {}};
    const want = new Set(searchedSkus.map(String));
    const firstP = (arr) => (Array.isArray(arr) ? arr[0] : arr) || '';
    // De un array de precios saca {internet, normal}. internet = no tachado;
    // normal = tachado (solo presente cuando hay descuento).
    const pickPrice = (prices) => {
        let internet = '', normal = '';
        for (const p of (prices || [])) {
            const val = firstP(p.price);
            if (p.crossed) { if (!normal) normal = val; }
            else { if (!internet) internet = val; }
        }
        return {internet, normal};
    };
    const out = {};
    for (const p of r) {
        const sz = (p.variants || []).find(v => v.type === 'SIZES');
        const options = sz ? (sz.options || []) : [];
        // Resumen de TODAS las medidas del producto (para la columna comparativa).
        const allSizes = options
            .map(o => `${o.size || o.value}: $${firstP((o.prices || []).find(x => !x.crossed)?.price)}`)
            .join(' | ');
        // 1) Productos con variantes de medida (puertas): matchear por o.variant.
        for (const o of options) {
            const sku = String(o.variant);
            if (!want.has(sku) || out[sku]) continue;
            const pr = pickPrice(o.prices);
            out[sku] = {
                sku_input: sku,
                product_id: String(p.productId || ''),
                marca: p.brand || '',
                descripcion: p.displayName || '',
                vendedor: p.sellerName || '',
                medida: o.size || o.value || '',
                precio_internet: pr.internet,
                precio_normal: pr.normal,
                disponible: o.available ? 'Sí' : 'No',
                n_medidas: options.length,
                todas_las_medidas: allSizes,
                url: (o.url || '').split('?')[0],
                img_url: firstP(o.mediaUrls),
            };
        }
        // 2) Productos SIN variantes de medida (si el SKU buscado no era puerta):
        //    matchear por skuId top-level. Mantiene el robot útil para SKUs mixtos.
        const topSku = String(p.skuId || '');
        if (want.has(topSku) && !out[topSku] && options.length === 0) {
            const pr = pickPrice(p.prices);
            out[topSku] = {
                sku_input: topSku,
                product_id: String(p.productId || ''),
                marca: p.brand || '',
                descripcion: p.displayName || '',
                vendedor: p.sellerName || '',
                medida: '(sin variantes)',
                precio_internet: pr.internet,
                precio_normal: pr.normal,
                disponible: '',
                n_medidas: 0,
                todas_las_medidas: '',
                url: (p.url || '').split('?')[0],
                img_url: firstP(p.mediaUrls),
            };
        }
    }
    return {state: 'ok', n_results: r.length, matches: out};
}"""


# Mapea cada card visible (div grid-pod) a su productId, leyendo el primer
# segmento numérico del href /articulo/{productId}/... Sirve para ubicar la
# card de cada puerta y sacarle la screenshot (como el MK7).
_CARD_PID_JS = r"""(card) => {
    const a = card.querySelector('a[href*="/articulo/"]');
    if (!a) return '';
    const m = (a.href || '').match(/\/articulo\/(\d+)/);
    return m ? m[1] : '';
}"""


async def scan_catalog_doors(
    page: Page,
    wanted: list[str],
    *,
    catalog_url: str = PUERTAS_CATALOG_URL,
    max_pages: int = MAX_FALLBACK_PAGES,
) -> dict:
    """Fallback: recorre el listado de categoría Puertas (paginado) y matchea los
    SKUs que el buscador NO devolvió. Devuelve {sku: data} con la misma forma que
    search_batch_doors, para que write_output los procese sin cambios.

    La zona debe estar fijada en `page` antes de llamar (los precios son de esa
    zona). Se detiene apenas encuentra todos los SKUs pedidos o agota las páginas.

    Devuelve (found, reached): `reached` indica si el listado se pudo cargar
    (para no marcar SKUs como ausentes por una falla de red).
    """
    remaining = list(dict.fromkeys(str(s) for s in wanted))
    if not remaining:
        return {}, True
    found: dict = {}
    reached = False
    for page_i in range(1, max_pages + 1):
        sep = "&" if "?" in catalog_url else "?"
        url = f"{catalog_url}{sep}page={page_i}"
        results = await _fetch_results_json(page, url)
        if not results:
            break
        reached = True
        try:
            payload = await page.evaluate(_EXTRACT_DOORS_JS, remaining)
        except Exception:
            payload = None
        for sku, data in ((payload or {}).get("matches") or {}).items():
            if sku not in found:
                found[sku] = data
        remaining = [s for s in remaining if s not in found]
        if not remaining:
            break
    return found, reached


async def search_batch_doors(
    page: Page,
    guards: list[str],
    chunk: list[str],
    *,
    screenshot_dir: Path = None,
    max_attempts: int = 3,
    expect_results: bool = True,
) -> dict:
    """Corre /buscar?Ntt={guards+chunk} y devuelve {sku: data} de las puertas
    (medidas) matcheadas. La extracción se hace sobre el chunk real (sin guards).

    Si screenshot_dir está dado, además captura la screenshot de la card de cada
    puerta (igual que el MK7) — una por productId, reutilizada para todas sus
    medidas (varias medidas colapsan en la misma card)."""
    query_skus = [g for g in guards if g not in chunk] + chunk
    if not chunk:
        return {}
    q = "+".join(query_skus)
    url = f"{BASE_URL}/buscar?Ntt={q}"

    results = None
    for attempt in range(1, max_attempts + 1):
        results = await _fetch_results_json(page, url)
        if results:
            break
        if not expect_results:
            break
        await page.wait_for_timeout(int(1500 * (2 ** (attempt - 1))))
    if not results:
        return {}

    try:
        payload = await page.evaluate(_EXTRACT_DOORS_JS, chunk)
    except Exception:
        return {}
    matches = (payload or {}).get("matches") or {}
    # Defensa: solo SKUs realmente pedidos en este chunk.
    chunkset = set(chunk)
    matches = {k: v for k, v in matches.items() if k in chunkset}
    if not matches:
        return {}

    # ── Screenshots de las cards (estilo MK7) ──────────────────────────
    if screenshot_dir:
        screenshot_dir = Path(screenshot_dir)
        screenshot_dir.mkdir(parents=True, exist_ok=True)
        # productIds que necesitamos fotografiar (uno por puerta matcheada).
        want_pids = {str(v.get("product_id", "")) for v in matches.values() if v.get("product_id")}
        if want_pids:
            # Quitar overlays + lazy-load de las cards antes de capturar.
            try:
                await page.evaluate("""() => {
                    document.querySelectorAll('[data-testid="overlay"], [class*="overlay"]')
                        .forEach(o => { try { o.remove(); } catch (_) {} });
                }""")
                for _ in range(5):
                    await page.evaluate("window.scrollBy(0, window.innerHeight)")
                    await page.wait_for_timeout(250)
                await page.evaluate("window.scrollTo(0, 0)")
                await page.wait_for_timeout(400)
            except Exception:
                pass
            # Mapear card → productId y capturar.
            pid_to_path: dict[str, str] = {}
            try:
                cards = await page.query_selector_all('div[class*="grid-pod"]')
            except Exception:
                cards = []
            for card in cards:
                try:
                    pid = (await card.evaluate(_CARD_PID_JS) or "").strip()
                except Exception:
                    pid = ""
                if not pid or pid not in want_pids or pid in pid_to_path:
                    continue
                img_path = screenshot_dir / f"{re.sub(r'[^0-9]', '_', pid)}.jpg"
                try:
                    await card.scroll_into_view_if_needed()
                    await page.wait_for_timeout(120)
                    await card.screenshot(path=str(img_path), type="jpeg", quality=78)
                    if img_path.exists() and img_path.stat().st_size > 0:
                        pid_to_path[pid] = str(img_path)
                except Exception:
                    pass
            # Adjuntar el path a cada match por su productId.
            for v in matches.values():
                v["screenshot_path"] = pid_to_path.get(str(v.get("product_id", "")), "")

    return matches


# ─────────────────────────────────────────  Multi-zone orchestrator  ──────
# (misma estructura que MK7 search_skus_mk6, adaptada a la extracción de puertas)

async def search_doors(
    skus: list[str],
    stores: list[dict],
    *,
    guards: list[str] = None,
    batch_size: int = DEFAULT_BATCH_SIZE,
    headless: bool = True,
    screenshot_dir=None,
    progress_cb=None,
    skip_store_ids=None,
    on_match=None,
):
    """Busca cada SKU de puerta en cada zona. Devuelve list[dict]; cada dict es
    un match (SKU × zona) con la medida y precio exactos + store_id/store_found.

    Eventos de progreso (mismos nombres que MK7 para reusar UIs):
      browser_launching / browser_ready / browser_error
      warmup_start / warmup_done {store}
      zone_start {store, n_skus}
      batch_done {store, batch_skus, found_in_batch, batches_done_in_zone, total_batches_in_zone}
      zone_end {store, found_in_zone, zone_failed, retried}
      complete {matches, stats}
    """
    guards = list(guards) if guards else list(DEFAULT_GUARD_SKUS)
    matches_all: list[dict] = []
    skip_store_ids = set(skip_store_ids or ())
    n_skus = len(skus)
    n_batches_per_zone = max(1, (n_skus + batch_size - 1) // batch_size)
    # SKUs que ni el buscador ni el listado de categoría devolvieron en la 1ª zona
    # ⇒ están fuera del catálogo: no reintentamos el fallback en zonas siguientes.
    catalog_absent: set[str] = set()

    async def _run_zone(browser, store, *, retried: bool) -> bool:
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900}, color_scheme="light",
            user_agent=USER_AGENT,
        )
        page = await ctx.new_page()
        try:
            if progress_cb: progress_cb({"event": "warmup_start", "store": store})
            await warmup_session(page)
            if progress_cb: progress_cb({"event": "warmup_done", "store": store})

            if progress_cb:
                progress_cb({"event": "zone_start", "store": store, "n_skus": n_skus})

            ok = await set_zone(page, store["region"], store["comuna"], warmup=False)
            if not ok:
                ok = await set_zone(page, store["region"], store["comuna"], warmup=False)
            if not ok:
                if progress_cb:
                    progress_cb({"event": "zone_end", "store": store,
                                 "found_in_zone": 0, "zone_failed": True,
                                 "retried": retried})
                return False

            zone_shots = (Path(screenshot_dir) / store["id"]) if screenshot_dir else None
            found_in_zone = 0
            matched_here: set[str] = set()

            def _record(sku, data):
                nonlocal found_in_zone
                _row = dict(data, store_found=store["name"], store_id=store["id"])
                matches_all.append(_row)
                matched_here.add(sku)
                if on_match: on_match(_row)
                found_in_zone += 1

            for i in range(0, n_skus, batch_size):
                chunk = skus[i:i + batch_size]
                try:
                    batch_matches = await search_batch_doors(
                        page, guards, chunk, screenshot_dir=zone_shots)
                except Exception:
                    batch_matches = {}
                found_here = 0
                for sku in chunk:
                    if sku in batch_matches:
                        _record(sku, batch_matches[sku])
                        found_here += 1
                if progress_cb:
                    progress_cb({"event": "batch_done", "store": store,
                                 "batch_skus": chunk,
                                 "found_in_batch": found_here,
                                 "batches_done_in_zone": (i // batch_size) + 1,
                                 "total_batches_in_zone": n_batches_per_zone})

            # Fallback: SKUs que el buscador no devolvió pero pueden estar en el
            # listado de categoría (índice de /buscar incompleto). Se omiten los
            # que ya probamos ausentes del catálogo en una zona previa.
            unmatched = [s for s in skus
                         if s not in matched_here and s not in catalog_absent]
            if unmatched:
                if progress_cb:
                    progress_cb({"event": "fallback_start", "store": store,
                                 "n_skus": len(unmatched)})
                try:
                    recovered, reached = await scan_catalog_doors(page, unmatched)
                except Exception:
                    recovered, reached = {}, False
                for sku in unmatched:
                    if sku in recovered:
                        _record(sku, recovered[sku])
                    elif reached:
                        catalog_absent.add(sku)
                if progress_cb:
                    progress_cb({"event": "fallback_done", "store": store,
                                 "recovered": len(recovered)})

            if progress_cb:
                progress_cb({"event": "zone_end", "store": store,
                             "found_in_zone": found_in_zone,
                             "zone_failed": False, "retried": retried})
            return True
        finally:
            try:
                await ctx.close()
            except Exception:
                pass

    if progress_cb: progress_cb({"event": "browser_launching"})
    launch_args = ["--no-sandbox", "--disable-dev-shm-usage"]
    async with Stealth().use_async(async_playwright()) as pw:
        try:
            browser = await pw.chromium.launch(
                headless=headless, args=launch_args, timeout=60_000,
            )
        except Exception as e:
            if progress_cb: progress_cb({"event": "browser_error",
                                          "stage": "launch", "msg": str(e)})
            return []
        if progress_cb: progress_cb({"event": "browser_ready"})
        failed_stores: list[dict] = []
        for store in stores:
            if str(store["id"]) in skip_store_ids:
                continue
            try:
                ok = await _run_zone(browser, store, retried=False)
            except Exception:
                ok = False
            if not ok:
                failed_stores.append(store)
        for store in failed_stores:
            try:
                await _run_zone(browser, store, retried=True)
            except Exception:
                pass
        await browser.close()

    distinct_skus_matched = len({m.get("sku_input", "") for m in matches_all})
    stats = {
        "total_input_skus":      n_skus,
        "distinct_skus_matched": distinct_skus_matched,
        "skus_not_found":        n_skus - distinct_skus_matched,
        "total_rows":            len(matches_all),
        "n_zones":               len(stores),
    }
    if progress_cb:
        progress_cb({"event": "complete", "matches": matches_all, "stats": stats})
    return matches_all


# ─────────────────────────────────────────  Helpers de precio  ─────────────

def _price_to_int(s):
    """'42.990' / '$ 42.990' → 42990. '' → None."""
    if not s:
        return None
    digits = re.sub(r"[^\d]", "", str(s))
    return int(digits) if digits else None


def _pct_descuento(normal, internet):
    n, i = _price_to_int(normal), _price_to_int(internet)
    if n and i and n > 0 and i < n:
        return f"-{round((1 - i / n) * 100)}%"
    return ""


# ─────────────────────────────────────────  IO helpers  ────────────────────

DESC_COL_ALIASES = ["Desc. Producto", "Desc Producto", "Desc", "Descripción", "Descripcion", "Descripción Producto", "Descripcion Easy"]
SKU_COL_ALIASES  = ["SKU Sodimac", "SKU Producto", "SKU", "Cód. Producto", "Código Producto", "Cod Producto"]
EASY_COL_ALIASES = ["SKU Easy", "Cód. Easy", "Codigo Easy", "Cod Easy", "Easy"]


def _pick_col(df, aliases):
    cols = {c.strip().lower(): c for c in df.columns}
    for a in aliases:
        if a.strip().lower() in cols:
            return cols[a.strip().lower()]
    return None


def read_input(path):
    """Lee el Excel/CSV de entrada. Devuelve (df, desc_col, sku_col, easy_col).
    Mismo formato unificado que el MK7: SKU Sodimac requerido; SKU Easy y
    Desc. Producto opcionales (se sintetizan si faltan)."""
    p = Path(path)
    if p.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(p)
    else:
        df = pd.read_csv(p, dtype=str)
    desc_col = _pick_col(df, DESC_COL_ALIASES)
    sku_col  = _pick_col(df, SKU_COL_ALIASES)
    easy_col = _pick_col(df, EASY_COL_ALIASES)
    if not sku_col:
        raise ValueError(
            f"No encontré columna de SKU. Columnas presentes: {list(df.columns)}. "
            f"Esperaba alguna de: {SKU_COL_ALIASES}."
        )
    def _norm(v):
        if pd.isna(v):
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    df[sku_col] = df[sku_col].apply(_norm).str.strip()
    if easy_col:
        df[easy_col] = df[easy_col].apply(_norm).str.strip()
    else:
        df["SKU Easy"] = ""
        easy_col = "SKU Easy"
    if not desc_col:
        df["Desc. Producto"] = ""
        desc_col = "Desc. Producto"
    return df, desc_col, sku_col, easy_col


def write_output(df, desc_col, sku_col, easy_col, matches, output_path, *,
                 stores=None, embed_images=True):
    """Combina los matches en un Excel long-format (una fila por SKU × zona).

    Columnas (la 'Medida' es la razón de ser de este robot):
      1. Tienda                 (store ID, ej. E522)
      2. Nombre Tienda          (ej. Cerrillos)
      3. SKU Easy               (passthrough)
      4. Desc. Producto         (passthrough)
      5. SKU Sodimac            (SKU de la medida buscada)
      6. Marca
      7. Descripción Producto   (nombre del producto puerta)
      8. Medida                 ← la medida exacta (ej. 90X200CM)
      9. Vendedor
     10. Precio Normal          (tachado, si hay oferta)
     11. Precio Internet        (precio exacto de ESA medida)
     12. % Descuento            (calculado)
     13. Todas las Medidas      (resumen medida: precio de todas)
     14. URL                    (link directo a la variante)
     15. Imagen                 (screenshot de la card, hoja 'Con fotos')
    """
    if stores is None:
        stores = ALL_STORES

    EMPTY = {
        "marca": "", "descripcion": "", "vendedor": "", "medida": "",
        "precio_internet": "", "precio_normal": "",
        "todas_las_medidas": "", "url": "", "screenshot_path": "",
    }

    by_sku_store: dict[tuple[str, str], dict] = {}
    for m in matches:
        by_sku_store[(str(m.get("sku_input", "")), str(m.get("store_id", "")))] = m

    cols_present = set(df.columns)
    has_easy = easy_col in cols_present
    has_desc = desc_col in cols_present

    rows: list[dict] = []
    for _, src in df.iterrows():
        sku  = str(src[sku_col])
        easy = src[easy_col] if has_easy else ""
        desc = src[desc_col] if has_desc else ""
        for store in stores:
            m = by_sku_store.get((sku, store["id"]))
            base = {"SKU Easy": easy, "Desc. Producto": desc, "SKU Sodimac": sku,
                    "store_id": store["id"], "store_found": store["name"]}
            base.update(EMPTY)
            if m:
                base.update({k: m.get(k, "") for k in EMPTY})
            else:
                # Sin match en esta tienda: marca explícita para el usuario.
                base["descripcion"] = "No encontrado"
            rows.append(base)

    from . import _locales_easy as _loc
    _rz = [_loc.region_zona(r["store_id"]) for r in rows]

    cols = {}
    cols["Tienda"]               = pd.Series([r["store_id"] for r in rows], dtype=object)
    cols["Nombre Tienda"]        = [r["store_found"] for r in rows]
    cols["Región"]               = [rz[0] for rz in _rz]
    cols["Zona"]                 = [rz[1] for rz in _rz]
    cols["SKU Easy"]             = pd.Series([r["SKU Easy"] for r in rows], dtype=object)
    cols["Desc. Producto"]       = pd.Series([r["Desc. Producto"] for r in rows], dtype=object)
    cols["SKU Sodimac"]          = pd.Series([r["SKU Sodimac"] for r in rows], dtype=object)
    cols["Marca"]                = [r["marca"] for r in rows]
    cols["Descripción Producto"] = [r["descripcion"] for r in rows]
    cols["Medida"]               = [r["medida"] for r in rows]
    cols["Vendedor"]             = [r["vendedor"] for r in rows]
    cols["Precio Normal"]        = [r["precio_normal"] for r in rows]
    cols["Precio Internet"]      = [r["precio_internet"] for r in rows]
    cols["% Descuento"]          = [_pct_descuento(r["precio_normal"], r["precio_internet"]) for r in rows]
    cols["Todas las Medidas"]    = [r["todas_las_medidas"] for r in rows]
    cols["URL"]                  = [r["url"] for r in rows]

    out = pd.DataFrame(cols)
    screenshot_paths = [r["screenshot_path"] for r in rows]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Datos", index=False)
        out_with_imgs = out.copy()
        out_with_imgs["Imagen"] = ""
        out_with_imgs.to_excel(writer, sheet_name="Con fotos", index=False)

    # Post-proceso: SKUs como texto + descarga/embed imágenes + URL truncation.
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(output_path)

        def _force_text_skus(ws):
            headers = {cell.value: cell.column for cell in ws[1]}
            for col_name in ("SKU Easy", "SKU Sodimac"):
                col_idx = headers.get(col_name)
                if not col_idx:
                    continue
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is None or cell.value == "":
                            continue
                        s = str(cell.value)
                        if s.endswith(".0") and s[:-2].isdigit():
                            s = s[:-2]
                        cell.value = s
                        cell.number_format = "@"

        for sheet_name in ("Datos", "Con fotos"):
            if sheet_name in wb.sheetnames:
                _force_text_skus(wb[sheet_name])

        # Embed screenshots de la card (estilo MK7) solo en 'Con fotos'.
        if embed_images and "Con fotos" in wb.sheetnames:
            ws = wb["Con fotos"]
            headers = {cell.value: cell.column for cell in ws[1]}
            img_col_idx = headers.get("Imagen")
            if img_col_idx:
                ws.column_dimensions[get_column_letter(img_col_idx)].width = 26
                for i, path in enumerate(screenshot_paths):
                    if not path or not Path(path).exists():
                        continue
                    ri = i + 2
                    ws.row_dimensions[ri].height = 160
                    try:
                        img = OpenpyxlImage(path); img.width = 170; img.height = 200
                        img.anchor = TwoCellAnchor(
                            editAs="oneCell",
                            _from=AnchorMarker(col=img_col_idx - 1, colOff=0, row=ri - 1, rowOff=0),
                            to=AnchorMarker(col=img_col_idx, colOff=0, row=ri, rowOff=0),
                        )
                        ws.add_image(img)
                    except Exception:
                        pass

        # Estética unificada (cabecera celeste, auto-filtro, anchos auto, URL
        # truncada con spacer, SIN freeze) — helper compartido en _excel_utils.
        from ._excel_utils import apply_clean_style
        for sheet_name in ("Datos", "Con fotos"):
            if sheet_name in wb.sheetnames:
                apply_clean_style(wb[sheet_name])

        wb.save(output_path)
    except Exception:
        pass
    return output_path

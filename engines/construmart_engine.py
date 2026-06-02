# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
"""Scraper interactivo de productos de Construmart, por sección.

Construmart es Magento (Improntus). Estrategia:
  1. Setear tienda (popup #store-popup → select Región → select Tienda → botón).
  2. Descubrir categorías leyendo #mainMenu (3 niveles).
  3. Usuario elige sección + subcategorías.
  4. Para cada subcategoría: paginar con ?p=N&product_list_limit=40 y
     extraer cards del DOM (.products.list.items.product-items > li.product-item).
  5. Descargar la imagen de cada card y embeberla en el Excel.

A diferencia de Falabella, los precios y stock de Construmart dependen de la
tienda seleccionada — por eso se mantiene la opción multi-tienda.
"""

import os
import re
import json
import asyncio
from pathlib import Path
from datetime import datetime

import pandas as pd
from playwright.async_api import async_playwright
from playwright_stealth import Stealth
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

import questionary
from questionary import Style as QStyle
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import (
    Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn, MofNCompleteColumn,
)
from rich import box

console = Console()

HOST = "https://www.construmart.cl"
BASE_URL = "https://www.construmart.cl"

try:
    PROJECT_DIR = Path(__file__).resolve().parent
except NameError:
    PROJECT_DIR = Path.cwd()
SCREENSHOT_DIR = PROJECT_DIR / "cards_screenshots"
PARTIAL_DIR = PROJECT_DIR / "partial_runs"
PROJECT_DIR.mkdir(parents=True, exist_ok=True)
SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
PARTIAL_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_REGION = "REGIÓN METROPOLITANA DE SANTIAGO"
DEFAULT_STORE_ID = "14"  # LAS CONDES
MAX_PAGES_PER_SUBCAT = 350
PER_PAGE = 40

# Tiendas de Construmart (cargadas vía discover_stores en el primer run; este
# listado es un fallback estático con los valores observados manualmente).
# Equivalencias Construmart store_id → lista de Sodimac/"Easy" store_id(s).
# Derivado del análisis por comuna/región (★ misma comuna · ◆ conurbación · ○ misma región).
SODIMAC_EQUIVALENTS = {
    "41":  ["E619"],          # IQUIQUE     → Arica (no hay tienda en Iquique)
    "44":  ["E534"],          # ANTOFAGASTA → Antofagasta
    "47":  ["E614"],          # CALAMA      → Calama
    "50":  ["E760"],          # COPIAPO     → Copiapó
    "56":  ["E521"],          # LA SERENA   → La Serena
    "53":  ["E521"],          # OVALLE      → La Serena
    "59":  ["E781"],          # EL BELLOTO  → El Belloto (Quilpué)
    "62":  ["E508", "E520"],  # LLO-LLEO    → Viña + Valparaíso
    "20":  ["E843"],          # BUIN ERRAZURIZ   → San Bernardo
    "23":  ["E843"],          # BUIN RODRIGUEZ   → San Bernardo
    "29":  ["E503"],          # CAMINO MELIPILLA → Maipú
    "32":  ["E988"],          # COLINA           → Chicureo (misma comuna)
    "35":  ["E513", "E643"],  # DEPARTAMENTAL    → El Llano + Ochagavía
    "17":  ["E510", "E874"],  # LA FLORIDA       → Florida + Santa Amalia
    "14":  ["E502"],          # LAS CONDES       → Kennedy
    "11":  ["E522"],          # LAS REJAS        → Cerrillos
    "26":  ["E503"],          # MELIPILLA        → Maipú
    "38":  ["E503"],          # PEÑAFLOR         → Maipú
    "65":  ["E504"],          # RANCAGUA         → Rancagua
    "71":  ["E592"],          # CURICO           → Curicó
    "68":  ["E524"],          # LINARES          → Linares
    "74":  ["E592"],          # MOLINA           → Curicó
    "77":  ["E591"],          # TALCA            → Talca
    "104": ["E525"],          # CHILLAN          → Chillán
    "86":  ["E633"],          # HUALPÉN          → Bio Bio (Hualpén)
    "80":  ["E529"],          # LOS ANGELES      → Los Ángeles
    "89":  ["E633", "E983"],  # SAN PEDRO DE LA PAZ → Bio Bio + Coronel
    "83":  ["E633"],          # TALCAHUANO       → Bio Bio (Hualpén)
    "92":  ["E517", "E506"],  # TEMUCO           → Temuco + Portal Temuco
    "95":  ["E585", "E748"],  # OSORNO           → Osorno + Portal Osorno
    "98":  ["E507"],          # PUERTO MONTT     → Puerto Montt
}


def sodimac_equivalent(store_id):
    return "|".join(SODIMAC_EQUIVALENTS.get(str(store_id), []))


STATIC_STORES = [
    {"id": "41",  "name": "IQUIQUE",              "region": "TARAPACÁ"},
    {"id": "44",  "name": "ANTOFAGASTA",          "region": "ANTOFAGASTA"},
    {"id": "47",  "name": "CALAMA",               "region": "ANTOFAGASTA"},
    {"id": "50",  "name": "COPIAPO",              "region": "ATACAMA"},
    {"id": "56",  "name": "LA SERENA",            "region": "COQUIMBO"},
    {"id": "53",  "name": "OVALLE",               "region": "COQUIMBO"},
    {"id": "59",  "name": "EL BELLOTO",           "region": "VALPARAISO"},
    {"id": "62",  "name": "LLO-LLEO",             "region": "VALPARAISO"},
    {"id": "20",  "name": "BUIN ERRAZURIZ",       "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "23",  "name": "BUIN RODRIGUEZ",       "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "29",  "name": "CAMINO MELIPILLA",     "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "32",  "name": "COLINA",               "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "35",  "name": "DEPARTAMENTAL",        "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "17",  "name": "LA FLORIDA",           "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "14",  "name": "LAS CONDES",           "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "11",  "name": "LAS REJAS",            "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "26",  "name": "MELIPILLA",            "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "38",  "name": "PEÑAFLOR",             "region": "REGIÓN METROPOLITANA DE SANTIAGO"},
    {"id": "65",  "name": "RANCAGUA",             "region": "O'HIGGINS"},
    {"id": "71",  "name": "CURICO",               "region": "MAULE"},
    {"id": "68",  "name": "LINARES",              "region": "MAULE"},
    {"id": "74",  "name": "MOLINA",               "region": "MAULE"},
    {"id": "77",  "name": "TALCA",                "region": "MAULE"},
    {"id": "104", "name": "CHILLAN",              "region": "ÑUBLE"},
    {"id": "86",  "name": "HUALPÉN",              "region": "BIOBÍO"},
    {"id": "80",  "name": "LOS ANGELES",          "region": "BIOBÍO"},
    {"id": "89",  "name": "SAN PEDRO DE LA PAZ",  "region": "BIOBÍO"},
    {"id": "83",  "name": "TALCAHUANO",           "region": "BIOBÍO"},
    {"id": "92",  "name": "TEMUCO",               "region": "LA ARAUCANÍA"},
    {"id": "95",  "name": "OSORNO",               "region": "LOS LAGOS"},
    {"id": "98",  "name": "PUERTO MONTT",         "region": "LOS LAGOS"},
]

QSTYLE = QStyle([
    ("qmark", "fg:#e30613 bold"),
    ("question", "bold"),
    ("answer", "fg:#00d97e bold"),
    ("pointer", "fg:#e30613 bold"),
    ("highlighted", "fg:#e30613 bold"),
    ("selected", "fg:#00d97e"),
    ("instruction", "fg:#888888"),
])


# ─────────────────────────────────────────  Persistence  ───────────────────

class PartialWriter:
    """Append-only JSONL writer para sobrevivir crashes en runs largos."""
    def __init__(self, run_id, append=False):
        self.path = PARTIAL_DIR / f"{run_id}.jsonl"
        mode = "a" if append else "w"
        self._fh = open(self.path, mode, encoding="utf-8")
        self.count = 0

    def write(self, row):
        self._fh.write(json.dumps(row, ensure_ascii=False) + "\n")
        self._fh.flush()
        self.count += 1

    def close(self):
        try:
            self._fh.close()
        except Exception:
            pass

    @staticmethod
    def load(path):
        rows = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
        return rows


# ─────────────────────────────────────────  Store popup  ───────────────────

async def discover_stores(page):
    """Carga el listado completo de regiones → tiendas del popup #store-popup.

    Devuelve list[dict(id, name, region)].
    """
    # Borrar cookie que oculta el popup tras la primera visita
    try:
        await page.context.clear_cookies(name="store_popup_shown")
    except Exception:
        pass
    await page.goto(f"{HOST}/", wait_until="domcontentloaded", timeout=60000)

    # Esperar hasta que el #region-selector-popup tenga opciones (poll hasta 20s)
    regions = []
    for _ in range(40):
        await page.wait_for_timeout(500)
        regions = await page.evaluate("""() => {
            const sel = document.querySelector('#region-selector-popup');
            if (!sel) return [];
            return [...sel.querySelectorAll('option')].map(o => o.value).filter(v => v);
        }""")
        if regions:
            break
    if not regions:
        # Forzar mostrar el popup si está oculto
        await page.evaluate("""() => {
            const p = document.querySelector('#store-popup');
            if (p) { p.classList.add('active'); p.style.display = 'block'; }
        }""")
        for _ in range(20):
            await page.wait_for_timeout(500)
            regions = await page.evaluate("""() => {
                const sel = document.querySelector('#region-selector-popup');
                if (!sel) return [];
                return [...sel.querySelectorAll('option')].map(o => o.value).filter(v => v);
            }""")
            if regions:
                break
    if not regions:
        return []
    stores = []
    for region in regions:
        opts = await page.evaluate(
            """async (region) => {
                const sel = document.querySelector('#region-selector-popup');
                sel.value = region;
                sel.dispatchEvent(new Event('change', {bubbles: true}));
                // Esperar a que se pueblen opciones (poll hasta 6s)
                const store = document.querySelector('#store-selector-popup');
                let opts = [];
                for (let i = 0; i < 30; i++) {
                    await new Promise(r => setTimeout(r, 200));
                    opts = [...store.querySelectorAll('option')]
                        .map(o => ({id: o.value, name: o.innerText.trim()}))
                        .filter(o => o.id);
                    if (opts.length > 0 && !store.disabled) break;
                }
                return opts;
            }""", region,
        )
        for o in opts:
            stores.append({"id": o["id"], "name": o["name"], "region": region})
    return stores


async def set_store(page, store_id, region=None):
    """Setea la tienda activa abriendo el popup #store-popup.

    Si el popup no aparece (ya estaba seteada otra tienda en cookie), navega a
    la home con un query especial para forzarlo. Devuelve True si éxito.
    """
    # Forzar popup borrando la cookie de "already shown"
    try:
        await page.context.clear_cookies(name="store_popup_shown")
    except Exception:
        pass

    await page.goto(f"{HOST}/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2500)

    # Si el popup no está visible, inyectarlo via CSS (a veces queda oculto si
    # ya hay store seleccionada en la sesión).
    visible = await page.evaluate(
        "() => { const p = document.querySelector('#store-popup'); "
        "return p && (p.classList.contains('active') || p.offsetHeight > 0); }"
    )
    if not visible:
        await page.evaluate("""() => {
            const p = document.querySelector('#store-popup');
            if (p) { p.classList.add('active'); p.style.display = 'block'; }
        }""")

    if region:
        ok = await page.evaluate(
            """async (region) => {
                const sel = document.querySelector('#region-selector-popup');
                if (!sel) return false;
                sel.value = region;
                sel.dispatchEvent(new Event('change', {bubbles: true}));
                await new Promise(r => setTimeout(r, 1500));
                return true;
            }""", region,
        )
        if not ok:
            return False

    ok = await page.evaluate(
        """async (storeId) => {
            const store = document.querySelector('#store-selector-popup');
            if (!store) return false;
            // Si la región no fue seteada, intentar deducirla del id buscando
            // entre todas las regiones.
            if (store.disabled || [...store.options].every(o => o.value !== storeId)) {
                const reg = document.querySelector('#region-selector-popup');
                const regions = [...reg.options].map(o => o.value).filter(v => v);
                for (const r of regions) {
                    reg.value = r;
                    reg.dispatchEvent(new Event('change', {bubbles: true}));
                    await new Promise(res => setTimeout(res, 1200));
                    if ([...store.options].some(o => o.value === storeId)) break;
                }
            }
            store.value = storeId;
            store.dispatchEvent(new Event('change', {bubbles: true}));
            await new Promise(r => setTimeout(r, 400));
            const btn = document.querySelector('#enter-store-button-popup');
            if (!btn || btn.disabled) return false;
            btn.click();
            return true;
        }""", store_id,
    )
    if not ok:
        return False
    await page.wait_for_timeout(2500)
    return True


async def set_store_with_retry(page, store_id, region=None, retries=2):
    for attempt in range(retries + 1):
        try:
            ok = await set_store(page, store_id, region)
        except Exception:
            ok = False
        if ok:
            return True
        if attempt < retries:
            await page.wait_for_timeout(1500 * (attempt + 1))
    return False


# ─────────────────────────────────────────  Discovery  ─────────────────────

EXCLUDED_TOP_NAMES_RE = re.compile(
    r'^(construdatos|ayuda|contacta|mis compras|catálogo)$',
    re.I,
)


async def discover_sections(page):
    """Devuelve [(section_name, [(group_name, group_url, [(leaf_name, leaf_url), ...]), ...]), ...]

    Lee #mainMenu del DOM (lo abre primero si está colapsado).
    """
    await page.goto(f"{HOST}/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2000)

    # Abrir el panel de categorías
    await page.evaluate("""() => {
        const t = document.querySelector('[data-action="toggle-nav"]');
        if (t) t.click();
    }""")
    await page.wait_for_timeout(800)

    raw = await page.evaluate("""() => {
        const tops = [...document.querySelectorAll('#mainMenu > li.item-title')];
        return tops.map(li => {
            const topA = li.querySelector(':scope > .label > a, :scope > a, :scope .label a');
            // Estructura del sub-menu: cada grupo (L2) contiene un titular y
            // hojas (L3). Usamos una heurística por niveles de path en la URL.
            const subLinks = [...li.querySelectorAll('.sub-menu a')]
                .map(a => ({text:(a.innerText||'').trim(), href:a.href}))
                .filter(x => x.text
                    && x.href.startsWith('http')
                    && x.href.includes('construmart.cl')
                    && !/^ver todo/i.test(x.text));
            return {
                top: (topA?.innerText||'').trim(),
                topHref: topA?.href || '',
                subs: subLinks,
            };
        });
    }""")

    sections = []
    seen_section_names = set()
    for entry in raw or []:
        title = (entry.get("top") or "").strip()
        top_href = (entry.get("topHref") or "").strip()
        if not title or title in seen_section_names:
            continue
        if EXCLUDED_TOP_NAMES_RE.match(title):
            continue
        # Determinar segmentos del path del top para clasificar L2/L3
        try:
            from urllib.parse import urlparse
            top_path = urlparse(top_href).path.strip("/")
        except Exception:
            top_path = ""
        top_segments = top_path.count("/") + 1 if top_path else 0

        # Agrupar: una hoja con path len == top_segments+1 es L2, len == +2 es L3.
        groups = {}      # name -> {url, leaves: [(name, url)]}
        group_order = []
        for s in entry.get("subs") or []:
            from urllib.parse import urlparse as _up
            href = s.get("href") or ""
            text = (s.get("text") or "").strip()
            if not text or not href:
                continue
            path = _up(href).path.strip("/")
            if top_path and not path.startswith(top_path):
                continue
            segs = path.split("/")
            depth = len(segs)
            if depth == top_segments + 1:
                # L2 group
                if text not in groups:
                    group_order.append(text)
                    groups[text] = {"url": href, "leaves": []}
                else:
                    groups[text]["url"] = groups[text]["url"] or href
            elif depth >= top_segments + 2:
                # L3 leaf: parent is the L2 (segs[top_segments])
                # Pero no sabemos el nombre del L2; lo identificamos vía path.
                # Buscamos en groups el que tenga la misma url-prefix de L2.
                l2_prefix = "/" + "/".join(segs[:top_segments + 1]) + "/"
                # Match groups by url prefix path
                from urllib.parse import urlparse as _up2
                parent_name = None
                for gn, gv in groups.items():
                    gp = _up2(gv["url"]).path
                    if gp.rstrip("/") == l2_prefix.rstrip("/"):
                        parent_name = gn
                        break
                if parent_name is None:
                    # Crear un grupo placeholder si aún no apareció
                    parent_name = segs[top_segments].replace("-", " ").title()
                    if parent_name not in groups:
                        group_order.append(parent_name)
                        groups[parent_name] = {
                            "url": HOST + l2_prefix.rstrip("/"),
                            "leaves": [],
                        }
                groups[parent_name]["leaves"].append((text, href))

        # Si un grupo no tiene hojas, su única hoja es él mismo.
        out_groups = []
        for gn in group_order:
            gv = groups[gn]
            leaves = gv["leaves"]
            if not leaves:
                leaves = [(gn, gv["url"])]
            # Quitar duplicados manteniendo orden
            seen_u = set()
            uniq = []
            for ln, lu in leaves:
                if lu in seen_u:
                    continue
                seen_u.add(lu)
                uniq.append((ln, lu))
            out_groups.append((gn, gv["url"], uniq))

        if not out_groups:
            # Sección sin sub-menú: usar el top como única hoja
            if top_href:
                out_groups = [(title, top_href, [(title, top_href)])]
            else:
                continue

        seen_section_names.add(title)
        sections.append((title, out_groups))

    return sections


# ─────────────────────────────────────────  Product extraction  ────────────

def _to_int_clp(v):
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    digits = re.sub(r"\D", "", str(v))
    return int(digits) if digits else None


async def _extract_products_from_dom(page):
    """Devuelve la lista de productos extraídos del DOM de la PLP actual."""
    return await page.evaluate("""() => {
        // Decodifica filename del amlabel a un texto legible.
        const decodeAmlabel = (src) => {
            const fn = (src || '').split('/').pop().toLowerCase()
                .replace(/\\.(png|jpg|jpeg|webp|svg)$/, '');
            // 10x9, 4x3, 3x2, etc. → "NxM"
            const nxm = fn.match(/(\\d+)\\s*x\\s*(\\d+)/);
            if (nxm) return `${nxm[1]}x${nxm[2]}`;
            // -3_x_4un, 3_10 (= -3% desde 10 un), 3_4, etc.
            const vol = fn.match(/-?(\\d+)[\\-_x ]+(\\d+)\\s*un?/);
            if (vol) return `-${vol[1]}% desde ${vol[2]} un`;
            const vol2 = fn.match(/^-?(\\d+)[_\\-](\\d+)$/);
            if (vol2) return `-${vol2[1]}% desde ${vol2[2]} un`;
            if (fn.includes('ahorro_profesional') || fn.includes('flags_ahorro')) return 'Precio exclusivo web';
            if (fn.includes('exclusivo')) return 'Precio exclusivo web';
            if (fn.includes('liquid')) return 'Liquidación';
            if (fn.includes('oferta')) return 'Oferta';
            if (fn.includes('nuevo')) return 'Nuevo';
            return fn;
        };
        const mainList = document.querySelector('.products.list.items.product-items');
        if (!mainList) return [];
        const items = [...mainList.querySelectorAll(':scope > li.product-item')];
        return items.map(li => {
            const link = li.querySelector('a.product-item-link');
            const brand = li.querySelector('.atributte-brand');
            const finalEl = li.querySelector('[data-price-type="finalPrice"]');
            const oldEl = li.querySelector('[data-price-type="oldPrice"]');
            const finalAmt = finalEl?.dataset?.priceAmount || '';
            const oldAmt = oldEl?.dataset?.priceAmount || '';
            const stockEl = li.querySelector('.stock');
            const stockText = stockEl ? (stockEl.innerText||'').trim() : '';
            const inStock = stockEl ? stockEl.classList.contains('available') : !!finalAmt;
            const img = li.querySelector('img.product-image-photo');
            const imgSrc = img?.src || '';
            const skuBtn = li.querySelector('[data-sku]');
            // En las PLP de búsqueda (catalogsearch) NO viene data-sku en los
            // cards; sólo viene data-product-id (id interno de Magento, distinto
            // del SKU). Fallback: extraer del slug de la URL (ej. .../foo-bar-23899).
            const linkHref = link?.href || '';
            const urlSku = (linkHref.match(/-(\\d+)(?:[\\/?#]|$)/) || [])[1] || '';
            const sku = skuBtn?.dataset?.sku || urlSku || '';
            const pidEl = li.querySelector('[data-product-id]');
            const productId = pidEl?.dataset?.productId || '';
            // Descuento visible (.discount-badge → "-12%")
            const discountBadge = li.querySelector('.discount-badge')?.innerText?.trim() || '';
            // Badges/promos (amasty label images)
            const amBadges = [...li.querySelectorAll('img[src*="amlabel"], img[src*="amasty"]')]
                .map(i => decodeAmlabel(i.src))
                .filter(t => t);
            return {
                name: (link?.innerText||'').trim(),
                url: link?.href || '',
                brand: (brand?.innerText||'').trim(),
                priceFinal: finalAmt,
                priceOld: oldAmt,
                stockText,
                inStock,
                imageUrl: imgSrc,
                sku,
                productId,
                discountBadge,
                badges: amBadges,
            };
        });
    }""")

def _extract_product_row(raw, section_name, subcat_name, store_id, store_name):
    pct = ""
    n_old = _to_int_clp(raw.get("priceOld"))
    n_final = _to_int_clp(raw.get("priceFinal"))
    if n_old and n_final and n_old > n_final:
        pct = f"-{round((n_old - n_final) / n_old * 100)}%"

    # Si el sitio muestra un .discount-badge ("-12%"), preferirlo
    badge_pct = (raw.get("discountBadge") or "").strip()
    if badge_pct:
        pct = badge_pct

    badges = raw.get("badges") or []
    return {
        "Tienda": sodimac_equivalent(store_id),
        "Nombre Tienda": store_name,
        "Sección": section_name,
        "Subcategoría": subcat_name,
        "Marca": raw.get("brand") or "",
        "SKU": raw.get("sku") or raw.get("productId") or "",
        "Product ID": raw.get("productId") or "",
        "Descripción Producto": raw.get("name") or "",
        "Precio Normal": n_old or "",
        "Precio Internet": n_final or "",
        "% Descuento": pct,
        "Promos": " | ".join(badges),
        "Stock": raw.get("stockText") or ("EN STOCK" if raw.get("inStock") else "SIN STOCK"),
        "En Stock": "Si" if raw.get("inStock") else "No",
        "URL": raw.get("url") or "",
    }


# ─────────────────────────────────────────  Image download  ────────────────

def _download_image(url, dest_path, timeout=10):
    if not url:
        return False
    if dest_path.exists() and dest_path.stat().st_size > 0:
        return True
    try:
        import requests
        r = requests.get(url, timeout=timeout, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        })
        if r.status_code != 200 or not r.content:
            return False
        dest_path.write_bytes(r.content)
        return True
    except Exception:
        return False


# ─────────────────────────────────────────  Scraping  ──────────────────────

async def _safe_goto(page, url, retries=2, timeout=60000):
    for attempt in range(retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            return True
        except Exception:
            if attempt < retries:
                await page.wait_for_timeout(1500 * (attempt + 1))
    return False


def _build_page_url(base_url, page_num):
    sep = "&" if "?" in base_url else "?"
    if page_num == 1:
        return f"{base_url}{sep}product_list_limit={PER_PAGE}"
    return f"{base_url}{sep}p={page_num}&product_list_limit={PER_PAGE}"


async def scrape_subcat(page, section_name, subcat_name, subcat_url, store,
                        progress=None, page_task=None, download_images=True):
    """Pagina una subcategoría y extrae todos los productos."""
    result = {"rows": [], "pages": 0, "truncated": False, "failed": False, "empty": False}
    seen = set()
    base_url = re.sub(r"([?&])p=\d+&?", r"\1", subcat_url).rstrip("?&")
    if base_url.endswith("?"):
        base_url = base_url[:-1]

    page_num = 1
    total_amount = None
    while True:
        url = _build_page_url(base_url, page_num)
        if not await _safe_goto(page, url):
            if page_num == 1:
                result["failed"] = True
            break
        await page.wait_for_timeout(1000)

        if total_amount is None:
            try:
                total_amount = await page.evaluate(
                    "() => { const t = document.querySelector('.toolbar-amount'); "
                    "if (!t) return 0; const m = (t.innerText||'').match(/(\\d[\\d.,]*)/); "
                    "return m ? parseInt(m[1].replace(/[.,]/g, '')) : 0; }"
                )
            except Exception:
                total_amount = 0

        items = await _extract_products_from_dom(page)
        if not items:
            if page_num == 1:
                result["empty"] = True
            break

        # Screenshots: capturar cada card completo (img + marca + nombre + precio + promos)
        card_paths = {}
        if download_images:
            cards = await page.query_selector_all(
                ".products.list.items.product-items > li.product-item"
            )
            for i, card in enumerate(cards):
                if i >= len(items):
                    break
                sku = items[i].get("sku") or items[i].get("productId") or f"unknown_{page_num}_{i}"
                img_path = SCREENSHOT_DIR / f"{sku}.jpg"
                if img_path.exists() and img_path.stat().st_size > 0:
                    card_paths[sku] = str(img_path)
                    continue
                try:
                    await card.scroll_into_view_if_needed()
                    # Asegurar que lazy images carguen antes del screenshot
                    await page.wait_for_timeout(120)
                    await card.screenshot(path=str(img_path), type="jpeg", quality=80)
                    card_paths[sku] = str(img_path)
                except Exception:
                    pass

        new_in_page = 0
        for raw in items:
            sku = raw.get("sku") or raw.get("productId")
            if not sku or sku in seen:
                continue
            seen.add(sku)
            row = _extract_product_row(raw, section_name, subcat_name, store["id"], store["name"])
            row["Image Path"] = card_paths.get(sku, "")
            result["rows"].append(row)
            new_in_page += 1

        result["pages"] = page_num
        if progress is not None and page_task is not None:
            try:
                progress.update(page_task,
                    description=f"  └─ {subcat_name} · pág {page_num} · {len(result['rows'])} cards")
            except Exception:
                pass

        # ¿Hay más páginas?
        if total_amount:
            max_page = (total_amount + PER_PAGE - 1) // PER_PAGE
        else:
            max_page = MAX_PAGES_PER_SUBCAT

        if page_num >= max_page or new_in_page == 0:
            break
        page_num += 1
        if page_num > MAX_PAGES_PER_SUBCAT:
            result["truncated"] = True
            break

    return result


# ─────────────────────────────────────────  Store picker  ──────────────────

def pick_stores(stores):
    rm = [s for s in stores if "METROPOLITANA" in (s["region"] or "").upper()]
    presets = [
        ("Solo LAS CONDES (default — más rápido)",
         [s for s in stores if s["id"] == DEFAULT_STORE_ID] or [stores[0]]),
        (f"Todas RM ({len(rm)} tiendas)", rm),
        (f"Todas Chile ({len(stores)} tiendas)", stores),
        ("Personalizado (elegir manualmente)", None),
    ]
    choice = questionary.select(
        "¿Qué tiendas querés scrapear?",
        choices=[questionary.Choice(title=t, value=i) for i, (t, _) in enumerate(presets)],
        style=QSTYLE, instruction="(↑↓ para moverte, Enter para elegir)",
    ).ask()
    if choice is None:
        return None
    _, picked = presets[choice]
    if picked is not None:
        return picked
    chs = [
        questionary.Choice(
            title=f"{s['id']:>3}  {s['name']:<22}  ({s['region']})",
            value=s,
            checked=(s["id"] == DEFAULT_STORE_ID),
        )
        for s in stores
    ]
    selected = questionary.checkbox(
        "Marcá las tiendas con [Espacio], confirmá con [Enter]:",
        choices=chs, style=QSTYLE,
    ).ask()
    return selected or []


# ─────────────────────────────────────────  Excel  ─────────────────────────

def write_excel(rows, output_file):
    if not rows:
        return False
    df = pd.DataFrame(rows)
    leading = [c for c in ("Tienda", "Nombre Tienda") if c in df.columns]
    desired = [
        "Sección", "Subcategoría", "Marca", "SKU", "Product ID",
        "Descripción Producto",
        "Precio Normal", "Precio Internet", "% Descuento", "Promos",
        "Stock", "En Stock",
        "URL", "Image Path",
    ]
    cols = [c for c in desired if c in df.columns and c not in leading]
    extras = [c for c in df.columns if c not in leading + cols]
    df = df[leading + cols + extras]
    df["Imagen"] = ""
    df.to_excel(output_file, index=False)
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    try:
        ipi = list(df.columns).index("Image Path") + 1
        ii = list(df.columns).index("Imagen") + 1
        ws.column_dimensions[openpyxl.utils.get_column_letter(ii)].width = 30
        for ri, rd in enumerate(rows, start=2):
            ws.row_dimensions[ri].height = 260
            ip = rd.get("Image Path", "")
            if ip and os.path.exists(ip):
                try:
                    img = OpenpyxlImage(ip); img.width = 200; img.height = 250
                    img.anchor = TwoCellAnchor(
                        editAs="oneCell",
                        _from=AnchorMarker(col=ii - 1, colOff=0, row=ri - 1, rowOff=0),
                        to=AnchorMarker(col=ii, colOff=0, row=ri, rowOff=0),
                    )
                    ws.add_image(img)
                except Exception:
                    pass
        ws.delete_cols(ipi)
        from ._excel_utils import apply_clean_style
        for _sn in wb.sheetnames:
            apply_clean_style(wb[_sn])
        wb.save(output_file)
    except Exception as e:
        console.print(f"[yellow]Aviso embebiendo imágenes: {e}[/]")
    return True


# ─────────────────────────────────────────  UI  ────────────────────────────

def banner():
    console.print()
    console.print(Panel.fit(
        "[bold #e30613]CONSTRUMART SECTION SCRAPER[/]\n"
        "[dim]Captura productos de Construmart por sección y tienda[/]",
        border_style="#e30613", padding=(1, 4),
    ))
    console.print()


def section_summary(rows, section_name, output_file):
    t = Table(title=f"Resumen — {section_name}", box=box.SIMPLE_HEAVY,
              show_header=True, header_style="bold #e30613")
    t.add_column("Métrica", style="white")
    t.add_column("Valor", style="bold green", justify="right")
    t.add_row("Filas totales", str(len(rows)))
    t.add_row("SKUs únicos", str(len({r["SKU"] for r in rows if r.get("SKU")})))
    t.add_row("Tiendas en el dataset", str(len({r["Tienda"] for r in rows if r.get("Tienda")})))
    t.add_row("Subcategorías cubiertas", str(len({r["Subcategoría"] for r in rows if r.get("Subcategoría")})))
    t.add_row("En stock", str(sum(1 for r in rows if r.get("En Stock") == "Si")))
    t.add_row("Con % Descuento", str(sum(1 for r in rows if r.get("% Descuento"))))
    t.add_row("Con Imagen", str(sum(1 for r in rows if r.get("Image Path"))))
    console.print(t)
    console.print(f"\n[bold green]✓ Excel guardado en:[/] [cyan]{output_file}[/]")


# ─────────────────────────────────────────  Main  ──────────────────────────

async def main():
    banner()

    async with Stealth().use_async(async_playwright()) as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900}, color_scheme="light",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await ctx.new_page()

        # Descubrir tiendas reales (fallback al estático si falla)
        with console.status("[cyan]Descubriendo tiendas…[/]", spinner="dots"):
            try:
                stores = await discover_stores(page)
            except Exception:
                stores = []
        if not stores:
            stores = STATIC_STORES
            console.print("[yellow]Usando listado estático de tiendas (fallback).[/]")
        else:
            console.print(f"[green]✓[/] {len(stores)} tiendas detectadas")

        picked_stores = pick_stores(stores)
        if not picked_stores:
            console.print("[yellow]No seleccionaste tiendas. Cancelado.[/]")
            await browser.close(); return

        download_images = questionary.confirm(
            "¿Descargar imágenes de productos y embeberlas en el Excel?",
            default=True, style=QSTYLE,
        ).ask()
        if download_images is None:
            await browser.close(); return

        # Setear primera tienda y descubrir categorías
        first = picked_stores[0]
        with console.status(f"[cyan]Configurando tienda {first['name']}…[/]", spinner="dots"):
            ok = await set_store_with_retry(page, first["id"], first.get("region"))
        if not ok:
            console.print(f"[yellow]No pude fijar la tienda inicial ({first['name']}). Continúo igual.[/]")
        else:
            console.print(f"[green]✓[/] Tienda inicial: {first['name']}")

        with console.status("[cyan]Descubriendo secciones del menú…[/]", spinner="dots"):
            sections = await discover_sections(page)
        if not sections:
            console.print("[red]No pude leer las secciones del menú.[/]")
            await browser.close(); return
        console.print(f"[green]✓[/] {len(sections)} secciones detectadas")

        choices = [
            questionary.Choice(
                title=f"{n}  [dim]({sum(len(l) for _, _, l in groups)} subcat. en {len(groups)} grupos)[/]",
                value=n,
            )
            for n, groups in sections
        ]
        section_name = questionary.select(
            "¿Qué sección querés scrapear?",
            choices=choices, style=QSTYLE,
            instruction="(↑↓ para moverte, Enter para elegir)",
        ).ask()
        if not section_name:
            await browser.close(); return
        groups = next(g for n, g in sections if n == section_name)

        choice_items = []
        for g_name, _g_url, leaves in groups:
            choice_items.append(questionary.Separator(f"── {g_name} ──"))
            for ln, lu in leaves:
                choice_items.append(questionary.Choice(title=ln, value=(ln, lu), checked=True))
        selected = questionary.checkbox(
            f"Subcategorías a scrapear de [{section_name}]",
            choices=choice_items, style=QSTYLE,
            instruction="(Espacio: marcar/desmarcar, A: todas, I: invertir, Enter: confirmar)",
        ).ask()
        if not selected:
            console.print("[yellow]No marcaste ninguna subcategoría. Cancelado.[/]")
            await browser.close(); return
        subcats = selected
        console.print(f"[green]✓[/] Sección: [bold]{section_name}[/] · {len(subcats)} subcategorías")

        all_rows = []
        skipped_zone = []
        failed_subcats = []
        truncated_subcats = []

        run_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{re.sub(r'[^A-Za-z0-9]+', '', section_name) or 'run'}"
        partial = PartialWriter(run_id)
        console.print(f"[dim]Persistencia incremental: {partial.path}[/]")

        with Progress(
            SpinnerColumn(),
            TextColumn("[bold #e30613]{task.description}"),
            BarColumn(bar_width=30),
            MofNCompleteColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TimeElapsedColumn(),
            console=console,
        ) as progress:
            store_task = progress.add_task("Tiendas", total=len(picked_stores))
            cat_task = progress.add_task("Subcategorías", total=len(subcats))
            page_task = progress.add_task("Esperando…", total=None)

            for st_idx, store in enumerate(picked_stores):
                progress.update(store_task, description=f"Tienda: {store['name']}")
                if st_idx > 0:
                    ok = await set_store_with_retry(page, store["id"], store.get("region"))
                    if not ok:
                        skipped_zone.append(store["id"])
                        progress.advance(store_task)
                        continue

                seen_skus = set()
                progress.reset(cat_task, total=len(subcats))
                for sc_name, sc_url in subcats:
                    progress.update(cat_task, description=f"  └─ {sc_name}")
                    res = await scrape_subcat(
                        page, section_name, sc_name, sc_url, store,
                        progress, page_task,
                        download_images=download_images,
                    )
                    if res["failed"]:
                        failed_subcats.append((store["id"], sc_name))
                    if res["truncated"]:
                        truncated_subcats.append((store["id"], sc_name))
                    for r in res["rows"]:
                        sku = r.get("SKU")
                        if not sku or sku in seen_skus:
                            continue
                        seen_skus.add(sku)
                        all_rows.append(r)
                        partial.write(r)
                    progress.advance(cat_task)
                progress.advance(store_task)
            progress.remove_task(page_task)

        await browser.close()
        partial.close()

    if skipped_zone:
        console.print(f"[yellow]Tiendas saltadas por falla de set_store: {', '.join(skipped_zone)}[/]")
    if failed_subcats:
        console.print(f"[yellow]Subcategorías que fallaron: {len(failed_subcats)}[/]")
        for store_id, sc in failed_subcats[:10]:
            console.print(f"   · {store_id} · {sc}")
    if truncated_subcats:
        console.print(f"[yellow]Subcategorías truncadas en {MAX_PAGES_PER_SUBCAT} pág: {len(truncated_subcats)}[/]")

    if not all_rows:
        console.print("[yellow]No se encontraron productos.[/]")
        return

    unique_skus = len({r["SKU"] for r in all_rows if r.get("SKU")})
    console.print(f"[dim]{len(all_rows)} filas totales · {unique_skus} SKUs únicos[/]")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    safe = re.sub(r'[^\w\s-]', '', section_name).strip().replace(' ', '_')
    suffix = "_con_imgs" if download_images else ""
    output = PROJECT_DIR / f"construmart_{safe}{suffix}_{timestamp}.xlsx"
    with console.status("[cyan]Escribiendo Excel con imágenes…[/]", spinner="dots"):
        write_excel(all_rows, str(output))

    section_summary(all_rows, section_name, output)

# ============== MK6 layer ==============

"""Construmart MK6 — busca un listado de SKUs en todas (o algunas) tiendas
de Construmart y devuelve precios/marcas/imágenes en un Excel multi-tienda.

Equivalente a Scrapper MK6 de Sodimac. Estrategia de búsqueda:
  - Magento de Construmart soporta multi-término en /catalogsearch/result/?q=
    siempre que los términos sean SKUs reales (los que aparecen en la URL/imagen
    del producto, NO los product_id internos).
  - Se agrupan en batches de BATCH_SIZE (def. 16) por tienda.
  - Si un batch devuelve 1 sólo resultado, Magento redirige al PDP — manejado.
  - Output: Excel 2 hojas (Datos + Con fotos), 1 fila por (SKU × tienda).
"""

import os, re, sys, json, asyncio
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

# En Colab, scraper_construmart ya fue ejecutado arriba — definimos proxy y dirs.
THIS_DIR = Path.cwd()
class _ScProxy: pass
sc = _ScProxy()
for _name in ("set_store_with_retry", "_extract_products_from_dom",
              "discover_stores", "discover_sections", "STATIC_STORES",
              "sodimac_equivalent", "SCREENSHOT_DIR"):
    setattr(sc, _name, globals().get(_name))

from playwright.async_api import async_playwright
from playwright_stealth import Stealth

HOST = "https://www.construmart.cl"
SEARCH_URL = HOST + "/catalogsearch/result/?q={q}"

SCREENSHOT_DIR = THIS_DIR / "mk6_screenshots"
SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)

# Aliases de columnas del input (compatible con MK6 Sodimac)
SKU_COL_ALIASES = [
    "SKU Construmart", "SKU", "SKU Producto", "Cód. Producto",
    "Código Producto", "Cod Producto",
]
EASY_COL_ALIASES = ["SKU Easy", "Cód. Easy", "Codigo Easy", "Cod Easy", "Easy"]
DESC_COL_ALIASES = ["Desc. Producto", "Desc Producto", "Descripción", "Descripcion",
                    "Descripcion Easy", "Descripción Easy"]


# ─────────────────────────────────────────  Input  ─────────────────────────

def _norm_sku(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("por definir", "pordefinir", "nan", "none"):
        return ""
    # Quitar decimales tipo "26950.0"
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s


def _pick_col(df, aliases):
    cols = {c.lower().strip(): c for c in df.columns}
    for a in aliases:
        if a.lower().strip() in cols:
            return cols[a.lower().strip()]
    return None


def read_input(path):
    """Lee Excel/CSV. Devuelve (df, desc_col, sku_col, easy_col).
    Requiere al menos la columna SKU."""
    p = Path(path)
    if p.suffix.lower() in (".csv", ".tsv"):
        df = pd.read_csv(p, dtype=str)
    else:
        df = pd.read_excel(p, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    sku_col = _pick_col(df, SKU_COL_ALIASES)
    easy_col = _pick_col(df, EASY_COL_ALIASES)
    desc_col = _pick_col(df, DESC_COL_ALIASES)
    if not sku_col:
        raise ValueError(
            f"Input no tiene columna de SKU. Aliases aceptados: {SKU_COL_ALIASES}\n"
            f"Columnas presentes: {list(df.columns)}"
        )
    df[sku_col] = df[sku_col].apply(_norm_sku).astype(str).str.strip()
    df = df[df[sku_col] != ""].reset_index(drop=True)
    if not easy_col:
        df["SKU Easy"] = ""
        easy_col = "SKU Easy"
    else:
        df[easy_col] = df[easy_col].apply(_norm_sku).astype(str)
    if not desc_col:
        df["Desc. Producto"] = ""
        desc_col = "Desc. Producto"
    return df, desc_col, sku_col, easy_col


# ─────────────────────────────────────────  Lookup engine  ─────────────────

async def _extract_from_pdp(page, sku):
    """Extrae datos cuando la búsqueda redirigió al PDP del producto."""
    data = await page.evaluate("""(sku) => {
        const decodeAmlabel = (src) => {
            const fn = (src || '').split('/').pop().toLowerCase()
                .replace(/\\.(png|jpg|jpeg|webp|svg)$/, '');
            const nxm = fn.match(/(\\d+)\\s*x\\s*(\\d+)/);
            if (nxm) return `${nxm[1]}x${nxm[2]}`;
            const vol = fn.match(/-?(\\d+)[\\-_x ]+(\\d+)\\s*un?/);
            if (vol) return `-${vol[1]}% desde ${vol[2]} un`;
            const vol2 = fn.match(/^-?(\\d+)[_\\-](\\d+)$/);
            if (vol2) return `-${vol2[1]}% desde ${vol2[2]} un`;
            if (fn.includes('ahorro_profesional') || fn.includes('flags_ahorro')) return 'Precio exclusivo web';
            if (fn.includes('exclusivo')) return 'Precio exclusivo web';
            return fn;
        };
        const finalEl = document.querySelector('.product-info-main [data-price-type="finalPrice"], [data-price-type="finalPrice"]');
        const oldEl = document.querySelector('.product-info-main [data-price-type="oldPrice"], [data-price-type="oldPrice"]');
        const finalAmt = finalEl?.dataset?.priceAmount || '';
        const oldAmt = oldEl?.dataset?.priceAmount || '';
        // Marca: buscar en attribute "Marca"
        let brand = '';
        const attrRows = [...document.querySelectorAll('.product.attribute, .product-info-main [class*="brand"], [data-th="Marca"], .product-brand, .atributte-brand')];
        for (const a of attrRows) {
            const t = (a.innerText||'').trim();
            if (t && t.length < 60 && !/precio|stock|sku/i.test(t)) { brand = t; break; }
        }
        // h1 page title
        const name = (document.querySelector('h1.page-title, h1[itemprop="name"], .page-title-wrapper h1')?.innerText||'').trim();
        // Imagen del producto: fotorama
        const img = document.querySelector('.fotorama__img, .fotorama img, .product.media img, img[class*="product-image"]');
        const imageUrl = img?.src || '';
        const disc = document.querySelector('.discount-badge')?.innerText?.trim() || '';
        const stockEl = document.querySelector('.stock');
        const stockText = stockEl ? (stockEl.innerText||'').trim() : '';
        const inStock = stockEl ? !stockEl.classList.contains('unavailable') : !!finalAmt;
        const amBadges = [...document.querySelectorAll('img[src*="amlabel"], img[src*="amasty"]')]
            .map(i => decodeAmlabel(i.src)).filter(Boolean);
        return {
            sku: String(sku),
            name, brand,
            priceFinal: finalAmt, priceOld: oldAmt,
            discountBadge: disc, badges: amBadges,
            stockText, inStock,
            url: location.href, imageUrl,
            source: 'pdp',
        };
    }""", sku)
    return data


async def _extract_from_plp(page, sku):
    items = await sc._extract_products_from_dom(page)
    for it in items:
        item_sku = str(it.get("sku") or it.get("productId") or "").strip()
        if item_sku == str(sku):
            it["source"] = "plp"
            return it
    return None


BATCH_SIZE = 16

# SKUs "guarda" — productos populares con stock amplio en la mayoría de tiendas.
# Se incluyen en cada query batch para forzar render del grid y mejorar la
# robustez ante búsquedas con muy pocos resultados. Se filtran del output.
#   198498 = Látex Constructor 4 GL Blanco SOQUINA
#   195637 = Rollo Lana De Vidrio Libre R94 40 mm VOLCÁN
#   23602  = Yeso Cartón Volcanita RH Borde Rebajado 12.5 mm
DEFAULT_GUARD_SKUS = ["198498", "195637", "23602"]


async def lookup_batch_in_store(page, skus, screenshot_dir=None, _retry=False):
    """Busca varios SKUs juntos. Devuelve dict {sku: raw_data}.

    /catalogsearch/result/?q=SKU1+SKU2+...&product_list_limit=40 — Magento
    devuelve PLP con hasta 40 matches en una sola página (sin esto serían
    10 y los SKUs caerían en páginas 2+). Si hay 1 sólo, redirige al PDP
    (manejado). Si quedan SKUs no encontrados después del batch, hace retry
    individual.

    Incluye SKUs guarda (DEFAULT_GUARD_SKUS) en cada query para forzar render
    del grid y mejorar robustez. Los guards se filtran del output.
    """
    found = {}
    if not skus:
        return found
    # Agregar guards (sólo en el primer intento, no en retry single-SKU)
    guards_used = []
    if not _retry:
        guards_used = [g for g in DEFAULT_GUARD_SKUS if str(g) not in [str(s) for s in skus]]
    query_skus = guards_used + [str(s) for s in skus]
    query = "+".join(query_skus)
    # `product_list_limit=40` es el max permitido por el limiter del sitio.
    url = SEARCH_URL.format(q=query) + "&product_list_limit=40"
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=45000)
    except Exception:
        return found
    await page.wait_for_timeout(1500)
    # Si redirigió a PDP, esperar a que el precio esté en el DOM
    if "/catalogsearch/" not in page.url:
        try:
            await page.wait_for_selector('[data-price-type="finalPrice"]', timeout=4000)
        except Exception:
            pass
    final_url = page.url

    if "/catalogsearch/" not in final_url:
        # Redirigió al PDP — sólo 1 match en todo el batch
        data = await _extract_from_pdp(page, None)
        # Si priceFinal quedó vacío, esperar y re-extraer una vez más
        if data and not data.get("priceFinal"):
            try:
                await page.wait_for_selector(
                    '[data-price-type="finalPrice"][data-price-amount]',
                    timeout=4000,
                )
            except Exception:
                pass
            data = await _extract_from_pdp(page, None)
        if data and data.get("priceFinal"):
            # Determinar a qué SKU del batch corresponde por el URL slug
            slug_sku = ""
            m = re.search(r"-(\d+)$", final_url)
            if m:
                slug_sku = m.group(1)
            matched = None
            for s in skus:
                if str(s) == slug_sku:
                    matched = str(s); break
            if not matched:
                # fallback: si sólo un SKU consultado, asumir que es ese
                if len(skus) == 1:
                    matched = str(skus[0])
            if matched:
                data["sku"] = matched
                if screenshot_dir:
                    try:
                        target = await page.query_selector(".product-info-main") \
                                 or await page.query_selector(".column.main")
                        if target:
                            shot = screenshot_dir / f"{matched}.jpg"
                            if not shot.exists():
                                await target.scroll_into_view_if_needed()
                                await page.wait_for_timeout(120)
                                await target.screenshot(path=str(shot),
                                                         type="jpeg", quality=78)
                            data["screenshot_path"] = str(shot)
                    except Exception:
                        pass
                found[matched] = data
        return found

    # PLP: extraer todos los cards de pág 1
    items = await sc._extract_products_from_dom(page)
    by_sku = {}
    for it in items:
        item_sku = str(it.get("sku") or it.get("productId") or "").strip()
        if item_sku:
            by_sku[item_sku] = it
    for s in skus:
        if str(s) in by_sku:
            data = by_sku[str(s)]
            data["source"] = "plp"
            found[str(s)] = data

    # Si la query con guards generó >40 productos (relevance fuzzy), algunos
    # SKUs pueden caer en página 2+. Paginamos hasta encontrar los faltantes
    # o agotar páginas (máx 4 para no eternizarnos).
    missing_now = [str(s) for s in skus if str(s) not in found]
    if missing_now:
        try:
            total = await page.evaluate(
                "() => { const t = document.querySelector('.toolbar-amount'); "
                "if (!t) return 0; const m = (t.innerText||'').match(/(\\d[\\d.,]*)/); "
                "return m ? parseInt(m[1].replace(/[.,]/g, '')) : 0; }"
            )
        except Exception:
            total = 0
        max_pages = min(4, (total + 39) // 40) if total else 1
        for pg in range(2, max_pages + 1):
            if not missing_now:
                break
            page_url = url + f"&p={pg}"
            try:
                await page.goto(page_url, wait_until="domcontentloaded", timeout=45000)
            except Exception:
                break
            await page.wait_for_timeout(900)
            more_items = await sc._extract_products_from_dom(page)
            for it in more_items:
                isku = str(it.get("sku") or it.get("productId") or "").strip()
                if isku and isku not in by_sku:
                    by_sku[isku] = it
                    items.append(it)
            for s in list(missing_now):
                if s in by_sku:
                    data = by_sku[s]
                    data["source"] = f"plp(p{pg})"
                    found[s] = data
                    missing_now.remove(s)

    # Screenshots por card matched
    if screenshot_dir and found:
        cards = await page.query_selector_all(
            ".products.list.items.product-items > li.product-item"
        )
        # Reconstruir mapa index → sku via items order
        for i, card in enumerate(cards):
            if i >= len(items):
                break
            item_sku = str(items[i].get("sku") or items[i].get("productId") or "").strip()
            if item_sku not in found:
                continue
            shot = screenshot_dir / f"{item_sku}.jpg"
            if shot.exists() and shot.stat().st_size > 0:
                found[item_sku]["screenshot_path"] = str(shot)
                continue
            try:
                await card.scroll_into_view_if_needed()
                await page.wait_for_timeout(80)
                await card.screenshot(path=str(shot), type="jpeg", quality=78)
                found[item_sku]["screenshot_path"] = str(shot)
            except Exception:
                pass


    # Filtrar guards del output (sólo el usuario debe ver sus SKUs)
    for g in guards_used:
        found.pop(g, None)

    return found

# ─────────────────────────────────────────  Search loop  ───────────────────

def _to_int(v):
    if not v:
        return ""
    if isinstance(v, (int, float)):
        return int(v)
    d = re.sub(r"\D", "", str(v))
    return int(d) if d else ""


def _row_from_data(raw, sku, store, easy, desc):
    pct = (raw.get("discountBadge") or "").strip()
    n_old = _to_int(raw.get("priceOld"))
    n_final = _to_int(raw.get("priceFinal"))
    if not pct and n_old and n_final and n_old > n_final:
        pct = f"-{round((n_old - n_final) / n_old * 100)}%"
    badges = raw.get("badges") or []
    return {
        "Tienda": sc.sodimac_equivalent(store["id"]),
        "Nombre Tienda": store["name"],
        "SKU Easy": easy or "",
        "Desc. Producto": desc or "",
        "SKU Construmart": sku,
        "Marca": raw.get("brand") or "",
        "Descripción Producto": raw.get("name") or "",
        "Precio Normal": n_old or "",
        "Precio Internet": n_final or "",
        "% Descuento": pct,
        "Promos": " | ".join(badges),
        "Stock": raw.get("stockText") or ("EN STOCK" if raw.get("inStock") else ""),
        "En Stock": "Si" if raw.get("inStock") else "No",
        "URL": raw.get("url") or "",
        "Source": raw.get("source") or "",
        "Image Path": raw.get("screenshot_path") or "",
    }


def _empty_row(sku, store, easy, desc, note=""):
    return {
        "Tienda": sc.sodimac_equivalent(store["id"]),
        "Nombre Tienda": store["name"],
        "SKU Easy": easy or "",
        "Desc. Producto": desc or "",
        "SKU Construmart": sku,
        "Marca": "", "Descripción Producto": "",
        "Precio Normal": "", "Precio Internet": "", "% Descuento": "",
        "Promos": "", "Stock": "", "En Stock": "",
        "URL": "", "Source": note or "no encontrado", "Image Path": "",
    }


async def search_skus(skus_with_meta, stores, screenshot=True, headless=True,
                      progress_cb=None, skip_pairs=None, on_row=None, status_cb=None):
    """
    skus_with_meta: list of dicts {sku, easy, desc}
    stores: list of {id, name, region}
    skip_pairs: set de (store_id, sku) ya completados (para resume); se saltan.
    on_row: callable(row_dict) — invocado por cada fila apenas se genera
        (para checkpoint incremental).
    Devuelve: list[dict row].
    """
    all_rows = []
    skip_pairs = skip_pairs or set()
    async with Stealth().use_async(async_playwright()) as pw:
        browser = await pw.chromium.launch(headless=headless)
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900}, color_scheme="light",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await ctx.new_page()
        for si, store in enumerate(stores):
            store_id = str(store["id"])
            # Si TODOS los SKUs de esta tienda ya están done → saltar set_store
            pending = [m for m in skus_with_meta
                       if (store_id, str(m["sku"])) not in skip_pairs]
            if not pending:
                if progress_cb:
                    progress_cb(si + 1, len(stores), store, len(skus_with_meta), len(all_rows))
                continue

            if status_cb: status_cb(f"🔧 Cambiando tienda a {store.get('name','?')}…")
            ok = await sc.set_store_with_retry(page, store["id"], store.get("region"))
            if not ok:
                for meta in pending:
                    row = _empty_row(meta["sku"], store, meta.get("easy"),
                                      meta.get("desc"), note="no_se_pudo_setear_tienda")
                    all_rows.append(row)
                    if on_row: on_row(row)
                if progress_cb:
                    progress_cb(si + 1, len(stores), store, None, len(all_rows))
                continue

            store_screenshot_dir = SCREENSHOT_DIR / store["id"]
            store_screenshot_dir.mkdir(parents=True, exist_ok=True)

            meta_by_sku = {m["sku"]: m for m in pending}
            skus_all = [m["sku"] for m in pending]
            done_skus = 0
            for batch_start in range(0, len(skus_all), BATCH_SIZE):
                batch = skus_all[batch_start: batch_start + BATCH_SIZE]
                if status_cb:
                    bn = batch_start // BATCH_SIZE + 1
                    btot = (len(skus_all) + BATCH_SIZE - 1) // BATCH_SIZE
                    status_cb(f"🔍 {store.get('name','?')} · buscando lote {bn}/{btot} ({len(batch)} SKUs)…")
                found = await lookup_batch_in_store(
                    page, batch,
                    screenshot_dir=store_screenshot_dir if screenshot else None,
                )
                for sku in batch:
                    meta = meta_by_sku[sku]
                    raw = found.get(sku)
                    if raw:
                        row = _row_from_data(raw, sku, store,
                                              meta.get("easy"), meta.get("desc"))
                    else:
                        row = _empty_row(sku, store, meta.get("easy"), meta.get("desc"))
                    all_rows.append(row)
                    if on_row: on_row(row)
                done_skus += len(batch)
                if progress_cb:
                    progress_cb(si + 1, len(stores), store, done_skus, len(all_rows))
        await browser.close()
    return all_rows

# ─────────────────────────────────────────  Output  ────────────────────────

OUTPUT_COLS = [
    "Tienda", "Nombre Tienda", "SKU Easy", "Desc. Producto", "SKU Construmart",
    "Marca", "Descripción Producto",
    "Precio Normal", "Precio Internet", "% Descuento", "Promos",
    "En Stock", "URL",
]


def write_output(rows, output_path):
    """Excel 2 hojas: 'Datos' (sin imágenes, filtrable) y 'Con fotos' (screenshots embebidos)."""
    if not rows:
        raise ValueError("No hay filas para escribir.")
    from ._excel_utils import filter_and_reorder, apply_url_truncation

    df = pd.DataFrame(rows)
    df_data = filter_and_reorder(df, OUTPUT_COLS)
    df_data.to_excel(output_path, index=False, sheet_name="Datos")

    wb = openpyxl.load_workbook(output_path)
    ws1 = wb["Datos"]
    ws2 = wb.create_sheet("Con fotos")
    photo_cols = OUTPUT_COLS + ["Imagen"]
    ws2.append(photo_cols)
    img_col_idx = len(photo_cols)
    ws2.column_dimensions[openpyxl.utils.get_column_letter(img_col_idx)].width = 30

    for ri, rd in enumerate(rows, start=2):
        for ci, col in enumerate(OUTPUT_COLS, start=1):
            ws2.cell(row=ri, column=ci, value=rd.get(col, ""))
        ip = rd.get("Image Path", "")
        if ip and os.path.exists(ip):
            ws2.row_dimensions[ri].height = 220
            try:
                img = OpenpyxlImage(ip); img.width = 170; img.height = 200
                img.anchor = TwoCellAnchor(
                    editAs="oneCell",
                    _from=AnchorMarker(col=img_col_idx - 1, colOff=0, row=ri - 1, rowOff=0),
                    to=AnchorMarker(col=img_col_idx, colOff=0, row=ri, rowOff=0),
                )
                ws2.add_image(img)
            except Exception:
                pass

    # URL truncado en ambas hojas
    url_col_idx_photos = photo_cols.index("URL") + 1
    url_col_idx_datos = OUTPUT_COLS.index("URL") + 1
    apply_url_truncation(ws2, url_col_idx_photos, img_col_idx, url_width=40, total_rows=len(rows) + 1)
    apply_url_truncation(ws1, url_col_idx_datos, url_col_idx_datos + 1, url_width=40, total_rows=len(rows) + 1)

    # Forzar columnas SKU como texto en ambas hojas
    for sheet in (ws1, ws2):
        cols_in_sheet = OUTPUT_COLS if sheet is ws1 else photo_cols
        for col_name in ("SKU Easy", "SKU Construmart"):
            if col_name in cols_in_sheet:
                cidx = cols_in_sheet.index(col_name) + 1
                for ri in range(2, len(rows) + 2):
                    c = sheet.cell(row=ri, column=cidx)
                    c.number_format = "@"
                    if c.value is not None:
                        c.value = str(c.value)

    from ._excel_utils import apply_clean_style
    for _sn in wb.sheetnames:
        apply_clean_style(wb[_sn])
    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────  CLI  ───────────────────────────

def _print(*a):
    print(*a, flush=True)


async def _amain(input_path, store_ids=None, screenshot=True, output_path=None,
                 headless=True):
    df, desc_col, sku_col, easy_col = read_input(input_path)
    _print(f"[{datetime.now():%H:%M:%S}] input: {input_path}")
    _print(f"  columnas: SKU='{sku_col}'  Easy='{easy_col}'  Desc='{desc_col}'")
    skus_with_meta = []
    seen = set()
    for _, r in df.iterrows():
        sku = str(r[sku_col]).strip()
        if not sku or sku in seen:
            continue
        seen.add(sku)
        skus_with_meta.append({
            "sku": sku,
            "easy": r.get(easy_col, ""),
            "desc": r.get(desc_col, ""),
        })
    _print(f"  {len(skus_with_meta)} SKUs únicos.")

    # Decidir tiendas
    if store_ids:
        store_lookup = {s["id"]: s for s in sc.STATIC_STORES}
        # complementar con dinámico
        try:
            async with Stealth().use_async(async_playwright()) as pw:
                br = await pw.chromium.launch(headless=True)
                ct = await br.new_context()
                pg = await ct.new_page()
                discovered = await sc.discover_stores(pg)
                await br.close()
            for s in discovered:
                store_lookup[s["id"]] = s
        except Exception:
            pass
        stores = [store_lookup[sid] for sid in store_ids if sid in store_lookup]
        if not stores:
            raise SystemExit(f"Ningún store_id válido en {store_ids}")
    else:
        # Default: solo LAS CONDES (14)
        stores = [s for s in sc.STATIC_STORES if s["id"] == "14"] or sc.STATIC_STORES[:1]
    _print(f"  tiendas: {[s['name'] for s in stores]}")

    t0 = datetime.now()
    def _cb(si, n_st, store, ki, total_rows):
        if ki is None:
            _print(f"  [{datetime.now():%H:%M:%S}] {si}/{n_st} {store['name']}: set_store falló")
        elif ki % 10 == 0 or ki == len(skus_with_meta):
            _print(f"  [{datetime.now():%H:%M:%S}] {si}/{n_st} {store['name']} "
                   f"sku {ki}/{len(skus_with_meta)}  ({total_rows} filas)")

    rows = await search_skus(skus_with_meta, stores, screenshot=screenshot,
                             headless=headless, progress_cb=_cb)
    dt = (datetime.now() - t0).total_seconds()
    found = sum(1 for r in rows if r.get("Precio Internet"))
    _print(f"[{datetime.now():%H:%M:%S}] {len(rows)} filas · {found} con precio · {dt:.0f}s")

    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = THIS_DIR / f"construmart_mk6_{ts}.xlsx"
    write_output(rows, str(output_path))
    _print(f"✓ Excel: {output_path}")
    return output_path


def main(argv=None):
    import argparse
    ap = argparse.ArgumentParser(description="Construmart MK6 — busca SKUs en varias tiendas.")
    ap.add_argument("input", help="Excel/CSV con columna SKU (y opcionales SKU Easy, Desc.)")
    ap.add_argument("--stores", nargs="*", default=None,
                    help="IDs de tiendas Construmart (ej: 14 17 65). Default: solo 14 LAS CONDES.")
    ap.add_argument("--no-screenshot", action="store_true",
                    help="No capturar screenshot de cards.")
    ap.add_argument("--output", default=None, help="Ruta del Excel de salida.")
    ap.add_argument("--no-headless", action="store_true", help="Mostrar el browser.")
    args = ap.parse_args(argv)
    asyncio.run(_amain(
        args.input, store_ids=args.stores,
        screenshot=not args.no_screenshot,
        output_path=args.output, headless=not args.no_headless,
    ))

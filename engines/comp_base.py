# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
"""Base compartida de los engines de "Competidores".

Todos los engines de Competidores (comp_*.py) hablan el mismo contrato para que
un solo launcher (launchers/competidores.py) los maneje sin saber de qué retailer
se trata:

  RETAILER_NAME : str
  USES_BROWSER  : bool      # los de API HTTP = False; los de DOM/Playwright = True
  discover_sections() -> [(section_name, [(subcat_name, subcat_ref), ...]), ...]
  scrape_section(subcats, *, on_row=None, progress_cb=None, limit=None) -> [row, ...]
  write_excel(rows, path) -> None

`subcat_ref` es opaco (lo interpreta cada engine: id de categoría VTEX, slug Woo,
dimensión Endeca, etc.). Una `row` es un dict con las claves de OUTPUT_COLS.

Este módulo NO depende de `requests` (usa urllib stdlib) para que los engines de
API sean testeables fuera de Colab. La escritura de Excel reusa la estética
unificada `apply_clean_style` del proyecto.
"""
from __future__ import annotations

import json
import ssl
import time
import urllib.request
import urllib.error

# Columnas unificadas de salida — pensadas para comparar competidores lado a lado.
OUTPUT_COLS = [
    "Tienda", "Sección", "Subcategoría", "Marca", "SKU",
    "Descripción Producto", "Precio Normal", "Precio Internet",
    "% Descuento", "URL",
]

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
_CTX = ssl.create_default_context()
_CTX.check_hostname = False
_CTX.verify_mode = ssl.CERT_NONE


def http_json(url, *, headers=None, timeout=30, retries=3, backoff=1.5):
    """GET `url` y parsea JSON. Reintenta ante errores transitorios.

    Devuelve el objeto parseado o lanza la última excepción tras agotar retries.
    """
    hdr = {"User-Agent": _UA, "Accept": "application/json"}
    if headers:
        hdr.update(headers)
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=hdr)
            with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
                return json.loads(r.read())
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as e:
            last = e
            if attempt < retries - 1:
                time.sleep(backoff * (attempt + 1))
    raise last


def http_text(url, *, headers=None, timeout=30, retries=3, backoff=1.5):
    """GET `url` y devuelve el texto crudo (para engines de DOM/HTML)."""
    hdr = {"User-Agent": _UA}
    if headers:
        hdr.update(headers)
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=hdr)
            with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
                return r.read().decode("utf-8", "replace")
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            last = e
            if attempt < retries - 1:
                time.sleep(backoff * (attempt + 1))
    raise last


def pct_discount(normal, internet):
    """% de descuento entero (0–100) entre precio normal e internet. '' si no aplica."""
    try:
        n = float(normal); i = float(internet)
        if n > 0 and i > 0 and i < n:
            return round(100 * (n - i) / n)
    except (TypeError, ValueError):
        pass
    return ""


def make_row(*, tienda, seccion="", subcat="", marca="", sku="",
             descripcion="", precio_normal="", precio_internet="", url="", img=""):
    """Construye una row con las claves de OUTPUT_COLS (y calcula % Descuento).

    `img` (URL de la imagen del producto) se guarda en la clave interna `_img`
    para embeberla en el Excel si se pide `with_images=True` (no es una columna
    de OUTPUT_COLS; se agrega como "Imagen" sólo cuando corresponde).
    """
    return {
        "Tienda": tienda, "Sección": seccion, "Subcategoría": subcat,
        "Marca": marca or "", "SKU": sku or "",
        "Descripción Producto": descripcion or "",
        "Precio Normal": precio_normal if precio_normal not in (None, "") else "",
        "Precio Internet": precio_internet if precio_internet not in (None, "") else "",
        "% Descuento": pct_discount(precio_normal, precio_internet),
        "URL": url or "",
        "_img": img or "",
    }


def _download_bytes(url, *, timeout=20):
    """Descarga bytes (imágenes). Devuelve None ante cualquier fallo."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": _UA})
        with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
            return r.read()
    except Exception:
        return None


def _embed_images(ws, rows, img_col_idx, *, cell_px=90, max_images=1500):
    """Descarga y embebe las imágenes (`_img`) en la columna `img_col_idx`.

    Defensivo: si falta PIL/openpyxl.drawing o una imagen falla, se salta esa
    fila sin romper el guardado. Las imágenes se bajan en paralelo (hilos) y se
    redimensionan a `cell_px`. Cap `max_images` para no explotar el Excel.
    """
    try:
        import io
        from concurrent.futures import ThreadPoolExecutor
        from openpyxl.drawing.image import Image as XLImage
        try:
            from PIL import Image as PILImage
        except Exception:
            PILImage = None
    except Exception:
        return

    urls = [(i, r.get("_img")) for i, r in enumerate(rows) if r.get("_img")]
    urls = urls[:max_images]
    if not urls:
        return

    def _fetch(item):
        i, u = item
        return i, _download_bytes(u)

    blobs = {}
    with ThreadPoolExecutor(max_workers=8) as ex:
        for i, data in ex.map(_fetch, urls):
            if data:
                blobs[i] = data

    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(img_col_idx)
    ws.column_dimensions[col_letter].width = cell_px / 7.0
    for i, data in blobs.items():
        try:
            bio = io.BytesIO(data)
            if PILImage is not None:
                im = PILImage.open(bio).convert("RGB")
                im.thumbnail((cell_px, cell_px))
                bio = io.BytesIO()
                im.save(bio, format="PNG")
                bio.seek(0)
            xl = XLImage(bio)
            row_excel = i + 2  # +1 header, +1 a 1-based
            ws.row_dimensions[row_excel].height = cell_px * 0.78
            ws.add_image(xl, f"{col_letter}{row_excel}")
        except Exception:
            continue


def write_excel(rows, path, *, sheet_name="Datos", columns=None, with_images=False):
    """Escribe las rows a un .xlsx con la estética unificada del proyecto.

    Usa openpyxl directo (no pandas) para no agregar dependencia. Aplica
    apply_clean_style (cabecera celeste, auto-filtro, anchos, URL truncada).

    Si `with_images=True`, agrega una columna "Imagen" (última) y embebe la
    imagen del producto (clave interna `_img` de cada row) descargándola y
    redimensionándola. `Imagen` siempre es la última columna (convención del
    proyecto, ver _excel_utils).
    """
    import openpyxl
    from engines._excel_utils import apply_clean_style

    cols = list(columns or OUTPUT_COLS)
    if with_images:
        cols = cols + ["Imagen"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])  # "Imagen" queda vacía; se embebe después
    apply_clean_style(ws)
    if with_images:
        _embed_images(ws, rows, len(cols))  # última columna = Imagen
    wb.save(path)

# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
# === Scrapper SKU — código del módulo ===
# (versión sincronizada desde scrapper_mk6.py)

"""Scrapper MK6 — buscador batch de SKU Tienda en Sodimac.

Idea:
  - Subís un Excel con columnas 'Desc. Producto' y 'SKU Producto'.
  - Para cada zona (en cascada), agrupamos los SKUs pendientes en batches
    de tamaño N (default 17) y forzamos la grilla agregando 3 SKUs guardia
    conocidos al inicio de cada query → URL: /buscar?Ntt=G1+G2+G3+S1+S2+...
  - Sodimac renderiza la página con SSR; los productos vienen estructurados
    en el script <__NEXT_DATA__> bajo props.pageProps.results.
  - De cada result extraemos:
        - skuId          → el SKU que matchea con el input
        - productId      → SKU TIENDA (el target)
        - sellerSkuId    → código interno adicional
        - brand, displayName, sellerName, prices, promotions, etc.
  - Los SKUs que no aparezcan en el JSON quedan en la lista pendiente para
    la siguiente zona.
"""

import re
from pathlib import Path

import pandas as pd
from playwright.async_api import async_playwright, Page
from playwright_stealth import Stealth


BASE_URL = "https://www.sodimac.cl/sodimac-cl"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Default guard SKUs — Sodimac products validated to be searchable in every
# zone of Chile (or fail in at most 1 zone each). They are prepended to every
# query so the search results page renders the product grid (single-SKU queries
# can return blank in headless). With 4 redundant guards, the grid renders even
# if 1-2 guards happen to be unavailable in the zone.
#   110038221 — UBERMANN martillo carpintero 20oz   (42/42 zonas)
#   130607328 — BAUKER set destornilladores 31 pzs  (42/42 zonas)
#   110229015 — ENERGIZER pilas AA x4               (41/42, falla solo en Arica)
#   110096085 — AQUACHEM Cloro granulado 2kg        (41/42, falla solo en Osorno)
DEFAULT_GUARD_SKUS = ["110038221", "130607328", "110229015", "110096085"]

# How many SKUs per query, including guards. Sodimac PLP page size is ~24, so
# 4 guards + 16 real = 20 fits comfortably in one page (no pagination needed).
DEFAULT_BATCH_SIZE = 16

ALL_STORES = [
    {"id": "E534", "name": "Antofagasta",   "region": "Antofagasta",   "comuna": "Antofagasta"},
    {"id": "E619", "name": "Arica",         "region": "Arica",         "comuna": "Arica"},
    {"id": "E633", "name": "Bio Bio",       "region": "Biobío",        "comuna": "Hualpén"},
    {"id": "E614", "name": "Calama",        "region": "Antofagasta",   "comuna": "Calama"},
    {"id": "E522", "name": "Cerrillos",     "region": "Metropolitana", "comuna": "Cerrillos"},
    {"id": "E988", "name": "Chicureo",      "region": "Metropolitana", "comuna": "Colina"},
    {"id": "E990", "name": "Chiguayante",   "region": "Biobío",        "comuna": "Chiguayante"},
    {"id": "E525", "name": "Chillán",       "region": "Ñuble",         "comuna": "Chillán"},
    {"id": "E760", "name": "Copiapó",       "region": "Atacama",       "comuna": "Copiapó"},
    {"id": "E983", "name": "Coronel",       "region": "Biobío",        "comuna": "Coronel"},
    {"id": "E511", "name": "Costanera",     "region": "Metropolitana", "comuna": "Providencia"},
    {"id": "E592", "name": "Curicó",        "region": "Maule",         "comuna": "Curicó"},
    {"id": "E781", "name": "El Belloto",    "region": "Valparaíso",    "comuna": "Quilpué"},
    {"id": "E513", "name": "El Llano",      "region": "Metropolitana", "comuna": "San Miguel"},
    {"id": "E510", "name": "Florida",       "region": "Metropolitana", "comuna": "La Florida"},
    {"id": "E502", "name": "Kennedy",       "region": "Metropolitana", "comuna": "Las Condes"},
    {"id": "E514", "name": "La Dehesa",     "region": "Metropolitana", "comuna": "Lo Barnechea"},
    {"id": "E512", "name": "La Reina",      "region": "Metropolitana", "comuna": "La Reina"},
    {"id": "E521", "name": "La Serena",     "region": "Coquimbo",      "comuna": "La Serena"},
    {"id": "E744", "name": "La Unión",      "region": "Ríos",          "comuna": "La Unión"},
    {"id": "E524", "name": "Linares",       "region": "Maule",         "comuna": "Linares"},
    {"id": "E900", "name": "Los Andes",     "region": "Valparaíso",    "comuna": "Los Andes"},
    {"id": "E529", "name": "Los Ángeles",   "region": "Biobío",        "comuna": "Los Ángeles"},
    {"id": "E503", "name": "Maipú",         "region": "Metropolitana", "comuna": "Maipú"},
    {"id": "E643", "name": "Ochagavía",     "region": "Metropolitana", "comuna": "Pedro Aguirre Cerda"},
    {"id": "E585", "name": "Osorno",        "region": "Lagos",         "comuna": "Osorno"},
    {"id": "E775", "name": "Portal Ñuñoa",  "region": "Metropolitana", "comuna": "Ñuñoa"},
    {"id": "E748", "name": "Portal Osorno", "region": "Lagos",         "comuna": "Osorno"},
    {"id": "E506", "name": "Portal Temuco", "region": "Araucanía",     "comuna": "Temuco"},
    {"id": "E659", "name": "Puente Alto",   "region": "Metropolitana", "comuna": "Puente Alto"},
    {"id": "E507", "name": "Puerto Montt",  "region": "Lagos",         "comuna": "Puerto Montt"},
    {"id": "E655", "name": "Quilicura",     "region": "Metropolitana", "comuna": "Quilicura"},
    {"id": "E518", "name": "Quilín",        "region": "Metropolitana", "comuna": "Peñalolén"},
    {"id": "E646", "name": "Quillota",      "region": "Valparaíso",    "comuna": "Quillota"},
    {"id": "E504", "name": "Rancagua",      "region": "O'Higgins",     "comuna": "Rancagua"},
    {"id": "E843", "name": "San Bernardo",  "region": "Metropolitana", "comuna": "San Bernardo"},
    {"id": "E874", "name": "Santa Amalia",  "region": "Metropolitana", "comuna": "La Florida"},
    {"id": "E591", "name": "Talca",         "region": "Maule",         "comuna": "Talca"},
    {"id": "E517", "name": "Temuco",        "region": "Araucanía",     "comuna": "Temuco"},
    {"id": "E520", "name": "Valparaíso",    "region": "Valparaíso",    "comuna": "Valparaíso"},
    {"id": "E830", "name": "Villarrica",    "region": "Araucanía",     "comuna": "Villarrica"},
    {"id": "E508", "name": "Viña del Mar",  "region": "Valparaíso",    "comuna": "Viña del Mar"},
]


# ─────────────────────────────────────────  Zone (autocomplete)  ───────────

async def _type_autocomplete(page: Page, placeholder: str, value: str) -> bool:
    sel = f'input[placeholder="{placeholder}"]'
    for _ in range(20):
        st = await page.evaluate(
            """(s) => { const i = document.querySelector(s);
                return i ? {present: true, disabled: i.disabled, hidden: i.offsetHeight === 0} : {present: false}; }""",
            sel,
        )
        if st.get("present") and not st.get("disabled") and not st.get("hidden"):
            break
        await page.wait_for_timeout(500)
    else:
        return False
    inp = page.locator(sel).first
    await page.evaluate("""() => {
        document.querySelectorAll(
            '[id*="onetrust"], [class*="onetrust"], '
            + '[id^="cookie"], [class^="cookie"], '
            + '#CybotCookiebotDialog, [class*="CookieConsent"]'
        ).forEach(e => { try { e.remove(); } catch (_) {} });
    }""")
    try:
        await inp.click(timeout=5000)
    except Exception:
        try:
            await inp.click(force=True, timeout=5000)
        except Exception:
            await page.evaluate("(s) => { const el = document.querySelector(s); if (el) el.focus(); }", sel)
    try:
        await inp.fill("", timeout=5000)
    except Exception:
        await page.evaluate("(s) => { const el = document.querySelector(s); if (el) el.value = ''; }", sel)
    await page.keyboard.type(value, delay=60)

    # Buscar match exacto/endsWith/contains. Si no aparece, ir borrando chars
    # del final (suggestions se relistan con menos filtro) hasta encontrar el
    # target real. Esto resuelve el caso "Calama" (Sodimac filtra agresivo y
    # esconde la opción correcta si tipeas el nombre completo).
    PICK_JS = """(target) => {
        const lis = [...document.querySelectorAll('li[class*="Autocomplete-module_suggestion"]')]
            .filter(e => e.offsetHeight > 0 && (e.innerText || '').trim());
        const norm = (s) => s.normalize("NFD").replace(/[\u0300-\u036f]/g, "").toLowerCase();
        const t = norm(target);
        const exact = lis.find(e => norm(e.innerText.trim()) === t);
        const endsWith = lis.find(e => {
            const x = norm(e.innerText.trim());
            return x === t || x.endsWith(" - " + t);
        });
        const contains = lis.find(e => norm(e.innerText.trim()).includes(t));
        const pick = exact || endsWith || contains;
        if (!pick) return null;
        const fire = (type) => pick.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
        fire('mousedown'); fire('mouseup'); fire('click');
        return pick.innerText.trim();
    }"""

    picked = None
    chars_left = len(value)
    while chars_left >= 3:
        for _ in range(12):
            await page.wait_for_timeout(250)
            has = await page.evaluate(
                """() => [...document.querySelectorAll('li[class*="Autocomplete-module_suggestion"]')]
                        .some(e => e.offsetHeight > 0 && (e.innerText||'').trim())"""
            )
            if has:
                break
        picked = await page.evaluate(PICK_JS, value)
        if picked:
            break
        await page.keyboard.press("Backspace")
        chars_left -= 1
        await page.wait_for_timeout(400)

    if picked:
        await page.wait_for_timeout(700)
        return True
    return False


async def set_zone(page: Page, region: str, comuna: str) -> bool:
    await page.goto(f"{BASE_URL}/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2000)
    opened = await page.evaluate("""() => {
        // 1) Selector específico nuevo (UI 2026): p dentro de Zone-module_zone-lable
        const p = document.querySelector('p[class*="Zone-module_zone-lable"]');
        if (p && p.offsetHeight > 0) { p.click(); return true; }
        // 2) Fallback al método legacy: buscar elemento con texto exacto
        const el = [...document.querySelectorAll('*')].find(e => {
            const t = (e.innerText || '').trim();
            return e.offsetHeight > 0 && (
                t === 'Ingresa tu ubicación'
                || /^Entrega en/.test(t)
                || /^Despacha en/.test(t)
                || /^Envía a/.test(t)
            );
        });
        if (el) { el.click(); return true; } return false;
    }""")
    if not opened:
        return False
    await page.wait_for_timeout(1500)
    if not await _type_autocomplete(page, "Ingresa una Región", region):
        return False
    if not await _type_autocomplete(page, "Ingresa una Comuna", comuna):
        return False
    await page.evaluate("""() => {
        const btn = [...document.querySelectorAll('button')].find(b => b.innerText.trim() === 'Guardar' && !b.disabled);
        if (btn) btn.click();
    }""")
    await page.wait_for_timeout(3000)
    return True


# ─────────────────────────────────────────  Warm-up  ───────────────────────

async def warmup_session(page: Page) -> None:
    """Visit a couple of pages so Sodimac fixes the session tokens.

    Without this, /buscar?Ntt=... served from a cold Chromium returns the
    home page instead of the product grid (anti-bot heuristic).
    """
    await page.goto(f"{BASE_URL}/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(3500)
    for _ in range(3):
        await page.evaluate("window.scrollBy(0, 350)")
        await page.wait_for_timeout(700)


# ─────────────────────────────────────────  Batch search  ──────────────────

async def _fetch_results_json(page: Page, url: str) -> list | None:
    """Navigate and return props.pageProps.results, or None on failure.

    Polls in-page (via evaluate) for the __NEXT_DATA__ script's results array.
    Avoids `wait_for_function` because long polling loops can race with
    Sodimac's client-side redirects, leaving page.content() out of sync.
    """
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
    except Exception:
        return None
    # Poll up to ~10s for the results array to be populated. On each tick
    # we evaluate and extract directly — no HTML snapshot, no race.
    deadline_ticks = 20  # 20 * 500ms = 10s
    extract_js = """() => {
        const s = document.getElementById('__NEXT_DATA__');
        if (!s) return {state: 'no-script'};
        try {
            const d = JSON.parse(s.textContent);
            const r = d?.props?.pageProps?.results;
            if (!Array.isArray(r)) return {state: 'no-results-key'};
            return {state: 'ok', results: r};
        } catch (e) { return {state: 'parse-error', err: String(e)}; }
    }"""
    last_results = None
    for _ in range(deadline_ticks):
        try:
            payload = await page.evaluate(extract_js)
        except Exception:
            payload = None
        if payload and payload.get("state") == "ok":
            r = payload.get("results") or []
            last_results = r
            if len(r) > 0:
                return r
        await page.wait_for_timeout(500)
    return last_results  # may be [] (genuine empty) or None (never had key)


_EXTRACT_CARD_JS = r"""(card) => {
    // DOM-based extraction — same logic as scraper_seccion_interactivo.
    const text = (sel) => {
        const el = card.querySelector(sel);
        return el ? (el.innerText || '').replace(/\s+/g, ' ').trim() : "";
    };
    const marca = text('.pod-title');
    const desc  = text('.pod-subTitle');
    const sellerEl = card.querySelector('.pod-sellerText, [class*="pod-seller"], [class*="sellerText"]');
    let vendedor = sellerEl ? (sellerEl.innerText || '').replace(/\s+/g, ' ').replace(/^Por\s+/i, '').trim() : "";
    if (!vendedor) {
        const all = card.innerText || "";
        const m = all.match(/Por\s+([A-Z0-9 ]{2,30})/);
        if (m) vendedor = m[1].trim();
    }
    const a = card.querySelector('a[href*="/articulo/"], a[href*="/product/"]');
    const href = a ? a.href : "";
    let sku_dom = "";
    if (href) {
        const parts = href.split('?')[0].split('/').filter(Boolean);
        sku_dom = parts[parts.length - 1];
    }
    let precios_congelados = "No";
    const badges = card.querySelectorAll('.pod-badges-item, [class*="Badge"]');
    for (const b of badges) {
        if ((b.innerText || '').includes("Precios Congelados")) { precios_congelados = "Si"; break; }
    }
    let precio_mayorista = "", descuento_mayorista = "";
    const wc = card.querySelector('.wholesale-container, [class*="wholesale"]');
    if (wc) {
        const bt = wc.querySelector('.bottom-text, [class*="bottom-text"]');
        descuento_mayorista = bt ? (bt.innerText || '').replace(/\s+/g, ' ').trim() : "";
        const pm = wc.querySelector('.prices-0 span[class*="copy10"], span[class*="copy10"]');
        precio_mayorista = pm ? (pm.innerText || '').replace(/\s+/g, ' ').trim() : "";
    }
    const priceSpans = [...card.querySelectorAll('.prices-0 span[class*="copy"], [class*="prices"] span[class*="copy"]')];
    const allPrices = priceSpans.map(p => (p.innerText || '').replace(/\s+/g, ' ').trim()).filter(t => /\$/.test(t));
    let precio_internet = "";
    for (const p of priceSpans) {
        if (wc && wc.contains(p)) continue;
        const t = (p.innerText || '').trim();
        if (/\$/.test(t)) { precio_internet = t.replace(/\s+/g, ' '); break; }
    }
    let precio_normal = "";
    const strike = card.querySelector('s, del, [class*="line-through"], [class*="crossed"], [class*="strikethrough"]');
    if (strike) precio_normal = (strike.innerText || '').replace(/\s+/g, ' ').trim();
    let precio_cmr = "";
    const cmrLabel = [...card.querySelectorAll('*')].find(e => {
        const t = (e.innerText || '').trim();
        return t && t.length < 50 && /\bCMR\b/.test(t) && /\$/.test(t);
    });
    if (cmrLabel) {
        const m = (cmrLabel.innerText || '').match(/\$\s*[\d.,]+/);
        if (m) {
            const candidate = m[0].replace(/\s+/g, ' ').trim();
            const intDigits = (precio_internet || '').replace(/\D/g, '');
            const cmrDigits = candidate.replace(/\D/g, '');
            if (cmrDigits && cmrDigits !== intDigits) precio_cmr = candidate;
        }
    }
    let pct_descuento = "";
    const pctRegex = /-?\d{1,3}\s*%/;
    const discBadge = card.querySelector('[class*="discount-badge-item"], [class*="discount-badge"]');
    if (discBadge) {
        const t = (discBadge.innerText || '').trim();
        const m = t.match(pctRegex);
        if (m) pct_descuento = m[0];
    }
    if (!pct_descuento) {
        for (const b of badges) {
            const t = (b.innerText || '').trim();
            if (pctRegex.test(t) && !/desde/i.test(t)) { pct_descuento = t.match(pctRegex)[0]; break; }
        }
    }
    return {
        vendedor: vendedor,
        marca: marca,
        sku_dom: sku_dom,
        descripcion: desc,
        precios_congelados: precios_congelados,
        precio_normal: precio_normal,
        precio_internet: precio_internet,
        pct_descuento: pct_descuento,
        precio_cmr: precio_cmr,
        precio_mayorista: precio_mayorista,
        descuento_mayorista: descuento_mayorista,
        todos_los_precios: allPrices.join(" | "),
        url: href,
    };
}"""


async def search_batch(
    page: Page,
    skus_query: list[str],
    screenshot_dir: Path = None,
    *,
    max_attempts: int = 3,
    expect_results: bool = True,
) -> dict:
    """Run /buscar?Ntt={s1+s2+...} and return parsed results.

    Strategy:
      1. Parse __NEXT_DATA__ to get the productId↔skuId mapping (we need this
         because our INPUT SKU often differs from the productId rendered in the
         DOM card's URL).
      2. Iterate the DOM cards and run the same JS extraction the seccion
         scraper uses — gives richer fields (Precio CMR, Descuento Mayorista
         literal, Todos los Precios, etc.) than the JSON snapshot.
    """
    if not skus_query:
        return {}
    q = "+".join(skus_query)
    url = f"{BASE_URL}/buscar?Ntt={q}"

    results = None
    for attempt in range(1, max_attempts + 1):
        results = await _fetch_results_json(page, url)
        if results:
            break
        if not expect_results:
            break
        await page.wait_for_timeout(int(1500 * (2 ** (attempt - 1))))
    if not results:
        return {}

    # The DOM href looks like /articulo/{productId}/{slug}/{skuId} — its last
    # numeric segment IS the input SKU. We use the JSON only as a "did the
    # page render at all?" sanity check; the actual matching is DOM-based.
    # We still grab productId from the JSON in case the caller wants it.
    skuid_to_pid = {}
    for r in results:
        pid = str(r.get("productId", "")).strip()
        sid = str(r.get("skuId", "")).strip()
        if pid and sid:
            skuid_to_pid[sid] = pid

    input_skus = set(skus_query)

    # Strip overlays + lazy-load all cards before extracting.
    await page.evaluate("""() => {
        document.querySelectorAll('[data-testid="overlay"], [class*="overlay"]')
            .forEach(o => { try { o.remove(); } catch (_) {} });
    }""")
    for _ in range(5):
        await page.evaluate("window.scrollBy(0, window.innerHeight)")
        await page.wait_for_timeout(250)
    await page.evaluate("window.scrollTo(0, 0)")
    await page.wait_for_timeout(400)

    # OPCIÓN A: esperar a que los precios terminen de renderizar (client-side)
    # antes de extraer. Causa raíz de las filas en blanco: el __NEXT_DATA__ (JSON)
    # llega antes que los precios visibles del DOM, y la extracción ganaba la carrera.
    # Estrategia: poll hasta que el nº de tarjetas con precio se ESTABILICE
    # (2 ticks iguales y >0) o se alcance el timeout (~8s). La estabilización
    # tolera productos sin stock que legítimamente nunca muestran precio.
    _price_ready_js = r"""() => {
        const cards = [...document.querySelectorAll('div[class*="grid-pod"]')];
        let withPrice = 0;
        for (const c of cards) {
            const spans = [...c.querySelectorAll('.prices-0 span[class*="copy"], [class*="prices"] span[class*="copy"]')];
            if (spans.some(s => /\$/.test((s.innerText || '')))) withPrice++;
        }
        return {total: cards.length, withPrice};
    }"""
    _prev_count, _stable_ticks = -1, 0
    for _ in range(16):  # 16 * 500ms = 8s máx
        try:
            _stat = await page.evaluate(_price_ready_js)
        except Exception:
            _stat = None
        _cur = _stat.get("withPrice", 0) if _stat else 0
        if _cur > 0 and _cur == _prev_count:
            _stable_ticks += 1
            if _stable_ticks >= 2:
                break
        else:
            _stable_ticks = 0
        _prev_count = _cur
        await page.wait_for_timeout(500)

    if screenshot_dir:
        screenshot_dir = Path(screenshot_dir)
        screenshot_dir.mkdir(parents=True, exist_ok=True)

    cards = await page.query_selector_all('div[class*="grid-pod"]')
    out: dict = {}
    for card in cards:
        try:
            data = await card.evaluate(_EXTRACT_CARD_JS)
        except Exception:
            continue
        sid = (data.get("sku_dom") or "").strip()
        if not sid or sid not in input_skus:
            continue
        if sid in out:
            # First match wins — guards against duplicates (sponsored slots).
            continue
        data["sku_input"] = sid
        data["product_id"] = skuid_to_pid.get(sid, "")
        data["screenshot_path"] = ""
        if screenshot_dir:
            safe_sid = re.sub(r"[^A-Za-z0-9_.-]", "_", sid)
            img_path = screenshot_dir / f"{safe_sid}.jpg"
            try:
                await card.scroll_into_view_if_needed()
                await page.wait_for_timeout(120)
                await card.screenshot(path=str(img_path), type="jpeg", quality=78)
                if img_path.exists() and img_path.stat().st_size > 0:
                    data["screenshot_path"] = str(img_path)
            except Exception:
                pass
        out[sid] = data
    return out


# ─────────────────────────────────────────  Multi-zone orchestrator  ──────

async def search_skus_mk6(
    skus: list[str],
    stores: list[dict],
    *,
    guards: list[str] = None,
    batch_size: int = DEFAULT_BATCH_SIZE,
    headless: bool = True,
    screenshot_dir=None,
    progress_cb=None,
    skip_store_ids=None,
    on_match=None,
):
    """Searches every SKU in EVERY zone (no cascading early-termination).

    For each input SKU and each zone, if the SKU appears in Sodimac's search
    response for that zone, an entry is recorded. A SKU may produce 0..N entries
    where N = number of zones it's stocked in.

    Returns list[dict]; each dict is one (SKU × zone) match with all the
    flattened result fields plus `store_id` and `store_found` (zone name).
    SKUs never matched are NOT in the returned list — caller can detect them
    by comparing to the input SKU set.

    Progress events:
      {"event": "warmup_start", "store": store | None}
      {"event": "warmup_done",  "store": store | None}
      {"event": "zone_start", "store": store, "n_skus": int}
      {"event": "batch_done", "store": store, "batch_skus": [...],
                              "found_in_batch": int,
                              "batches_done_in_zone": int,
                              "total_batches_in_zone": int}
      {"event": "zone_end", "store": store, "found_in_zone": int,
                            "zone_failed": bool, "retried": bool}
      {"event": "complete", "matches": list, "stats": dict}
    """
    guards = list(guards) if guards else list(DEFAULT_GUARD_SKUS)
    matches_all: list[dict] = []
    skip_store_ids = set(skip_store_ids or ())
    n_skus = len(skus)
    n_batches_per_zone = max(1, (n_skus + batch_size - 1) // batch_size)

    async def _run_zone(browser, store, *, retried: bool) -> bool:
        """Process one zone in a fresh context. Returns True on success.

        A fresh context per zone isolates session state — if a prior zone
        corrupted cookies / left a modal pegado, the next zone starts clean.
        Re-warming inside this context also refreshes the anti-bot tokens
        that otherwise expire halfway through a long run.
        """
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900}, color_scheme="light",
            user_agent=USER_AGENT,
        )
        page = await ctx.new_page()
        try:
            if progress_cb: progress_cb({"event": "warmup_start", "store": store})
            await warmup_session(page)
            if progress_cb: progress_cb({"event": "warmup_done", "store": store})

            if progress_cb:
                progress_cb({"event": "zone_start", "store": store, "n_skus": n_skus})

            ok = await set_zone(page, store["region"], store["comuna"])
            if not ok:
                ok = await set_zone(page, store["region"], store["comuna"])
            if not ok:
                if progress_cb:
                    progress_cb({"event": "zone_end", "store": store,
                                 "found_in_zone": 0, "zone_failed": True,
                                 "retried": retried})
                return False

            zone_shots = (Path(screenshot_dir) / store["id"]) if screenshot_dir else None
            found_in_zone = 0

            for i in range(0, n_skus, batch_size):
                chunk = skus[i:i + batch_size]
                guards_for_query = [g for g in guards if g not in chunk]
                query_skus = guards_for_query + chunk
                try:
                    batch_matches = await search_batch(page, query_skus, screenshot_dir=zone_shots)
                except Exception:
                    batch_matches = {}
                for g in guards_for_query:
                    batch_matches.pop(g, None)
                found_here = 0
                for sku in chunk:
                    if sku in batch_matches:
                        _row = dict(
                            batch_matches[sku],
                            store_found=store["name"],
                            store_id=store["id"],
                        )
                        matches_all.append(_row)
                        if on_match: on_match(_row)
                        found_here += 1
                found_in_zone += found_here
                if progress_cb:
                    progress_cb({"event": "batch_done", "store": store,
                                 "batch_skus": chunk,
                                 "found_in_batch": found_here,
                                 "batches_done_in_zone": (i // batch_size) + 1,
                                 "total_batches_in_zone": n_batches_per_zone})

            if progress_cb:
                progress_cb({"event": "zone_end", "store": store,
                             "found_in_zone": found_in_zone,
                             "zone_failed": False, "retried": retried})
            return True
        finally:
            try:
                await ctx.close()
            except Exception:
                pass

    if progress_cb: progress_cb({"event": "browser_launching"})
    # In Colab/Linux/root environments Chromium needs --no-sandbox or it hangs
    # silently at launch. --disable-dev-shm-usage avoids /dev/shm exhaustion.
    launch_args = ["--no-sandbox", "--disable-dev-shm-usage"]
    async with Stealth().use_async(async_playwright()) as pw:
        try:
            # Playwright's `timeout` kwarg is in ms — 60s.
            browser = await pw.chromium.launch(
                headless=headless, args=launch_args, timeout=60_000,
            )
        except Exception as e:
            if progress_cb: progress_cb({"event": "browser_error",
                                          "stage": "launch", "msg": str(e)})
            return []
        if progress_cb: progress_cb({"event": "browser_ready"})
        failed_stores: list[dict] = []
        for store in stores:
            if str(store["id"]) in skip_store_ids:
                continue  # ya completada en run anterior (resume)
            try:
                ok = await _run_zone(browser, store, retried=False)
            except Exception:
                ok = False
            if not ok:
                failed_stores.append(store)

        # Second pass over zones that failed setZone / blew up the context.
        for store in failed_stores:
            try:
                await _run_zone(browser, store, retried=True)
            except Exception:
                pass

        await browser.close()

    distinct_skus_matched = len({m.get("sku_input", "") for m in matches_all})
    stats = {
        "total_input_skus":      n_skus,
        "distinct_skus_matched": distinct_skus_matched,
        "skus_not_found":        n_skus - distinct_skus_matched,
        "total_rows":            len(matches_all),
        "n_zones":               len(stores),
    }
    if progress_cb:
        progress_cb({"event": "complete", "matches": matches_all, "stats": stats})
    return matches_all


# ─────────────────────────────────────────  IO helpers  ────────────────────

DESC_COL_ALIASES = ["Desc. Producto", "Desc Producto", "Desc", "Descripción", "Descripcion", "Descripción Producto", "Descripcion Easy"]
SKU_COL_ALIASES  = ["SKU Sodimac", "SKU Producto", "SKU", "Cód. Producto", "Código Producto", "Cod Producto"]
EASY_COL_ALIASES = ["SKU Easy", "Cód. Easy", "Codigo Easy", "Cod Easy", "Easy"]


def _pick_col(df, aliases):
    cols = {c.strip().lower(): c for c in df.columns}
    for a in aliases:
        if a.strip().lower() in cols:
            return cols[a.strip().lower()]
    return None


def read_input(path):
    """Reads input Excel/CSV. Returns (df, desc_col, sku_col, easy_col).

    Expects something like 'SKU Easy, Desc, SKU Sodimac'. SKU column is required;
    description and Easy code are optional (synthesized if missing).
    """
    p = Path(path)
    if p.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(p)
    else:
        df = pd.read_csv(p, dtype=str)
    desc_col = _pick_col(df, DESC_COL_ALIASES)
    sku_col  = _pick_col(df, SKU_COL_ALIASES)
    easy_col = _pick_col(df, EASY_COL_ALIASES)
    if not sku_col:
        raise ValueError(
            f"No encontré columna de SKU. Columnas presentes: {list(df.columns)}. "
            f"Esperaba alguna de: {SKU_COL_ALIASES}."
        )
    def _norm(v):
        if pd.isna(v):
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    df[sku_col] = df[sku_col].apply(_norm).str.strip()
    if easy_col:
        df[easy_col] = df[easy_col].apply(_norm).str.strip()
    else:
        df["SKU Easy"] = ""
        easy_col = "SKU Easy"
    if not desc_col:
        df["Desc. Producto"] = ""
        desc_col = "Desc. Producto"
    return df, desc_col, sku_col, easy_col


def write_output(df, desc_col, sku_col, easy_col, matches, output_path, stores=None):
    """Combine matches into a long-format Excel.

    Matches now come from the DOM extraction (mirrors scraper_seccion_interactivo)
    so the columns are aligned with that scraper.

    Column order (one row per input × zone):
      1. Tienda                    (store ID, e.g. E522)
      2. Nombre Tienda             (store name, e.g. Cerrillos)
      3. SKU Easy                  (passthrough from input)
      4. Desc. Producto            (passthrough from input)
      5. SKU Sodimac               (input SKU, = sku_col)
      6. Vendedor
      7. Marca
      8. Descripción Producto
      9. Precio Normal
     10. Precio Internet
     11. % Descuento
     12. Precio CMR
     13. Precio Mayorista
     14. Descuento Mayorista
     15. Todos los Precios
     16. URL
     17. Screenshot tarjeta        (embedded image, if available)
    """
    if stores is None:
        stores = ALL_STORES

    EMPTY = {
        "vendedor": "", "marca": "", "descripcion": "",
        "precios_congelados": "", "precio_normal": "", "precio_internet": "",
        "pct_descuento": "", "precio_cmr": "", "precio_mayorista": "",
        "descuento_mayorista": "", "todos_los_precios": "",
        "url": "", "screenshot_path": "",
    }

    # Group matches by (SKU, Store) — match.sku_input is the input SKU.
    by_sku_store: dict[tuple[str, str], dict] = {}
    for m in matches:
        by_sku_store[(str(m.get("sku_input", "")), str(m.get("store_id", "")))] = m

    # Columnas opcionales: si el archivo de entrada no trae easy_col/desc_col
    # (o el nombre no se detectó), usamos "" en vez de crashear con KeyError.
    cols_present = set(df.columns)
    has_easy = easy_col in cols_present
    has_desc = desc_col in cols_present

    # Build long-format rows
    rows: list[dict] = []
    for _, src in df.iterrows():
        sku  = str(src[sku_col])
        easy = src[easy_col] if has_easy else ""
        desc = src[desc_col] if has_desc else ""
        for store in stores:
            m = by_sku_store.get((sku, store["id"]))
            base = {"SKU Easy": easy, "Desc. Producto": desc, "SKU Sodimac": sku,
                    "store_id": store["id"], "store_found": store["name"]}
            if m:
                base.update({k: m.get(k, "") for k in EMPTY})
            else:
                base.update(EMPTY)
            rows.append(base)

    cols = {}
    cols["Tienda"]                   = pd.Series([r["store_id"] for r in rows], dtype=object)
    cols["Nombre Tienda"]            = [r["store_found"] for r in rows]
    cols["SKU Easy"]                 = pd.Series([r["SKU Easy"] for r in rows], dtype=object)
    cols["Desc. Producto"]           = pd.Series([r["Desc. Producto"] for r in rows], dtype=object)
    cols["SKU Sodimac"]              = pd.Series([r["SKU Sodimac"] for r in rows], dtype=object)
    cols["Vendedor"]                 = [r["vendedor"] for r in rows]
    cols["Marca"]                    = [r["marca"] for r in rows]
    cols["Descripción Producto"]     = [r["descripcion"] for r in rows]
    cols["Precio Normal"]            = [r["precio_normal"] for r in rows]
    cols["Precio Internet"]          = [r["precio_internet"] for r in rows]
    cols["% Descuento"]              = [r["pct_descuento"] for r in rows]
    cols["Precio CMR"]               = [r["precio_cmr"] for r in rows]
    cols["Precio Mayorista"]         = [r["precio_mayorista"] for r in rows]
    cols["Descuento Mayorista"]      = [r["descuento_mayorista"] for r in rows]
    cols["Todos los Precios"]        = [r["todos_los_precios"] for r in rows]
    cols["URL"]                      = [r["url"] for r in rows]

    # Build dataframe SIN columna Imagen para la hoja 1 ("Datos", limpia para filtrar).
    out = pd.DataFrame(cols)
    screenshot_paths = [r["screenshot_path"] for r in rows]

    # Dos hojas: 'Datos' (sin imágenes) y 'Con fotos' (con Imagen como última columna).
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Datos", index=False)
        out_with_imgs = out.copy()
        out_with_imgs["Imagen"] = ""  # última columna
        out_with_imgs.to_excel(writer, sheet_name="Con fotos", index=False)

    # Post-process: SKUs como texto + embed screenshots + URL truncation
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
        from openpyxl.utils import get_column_letter
        from ._excel_utils import apply_url_truncation

        wb = openpyxl.load_workbook(output_path)

        def _force_text_skus(ws):
            headers = {cell.value: cell.column for cell in ws[1]}
            for col_name in ("SKU Easy", "SKU Sodimac"):
                col_idx = headers.get(col_name)
                if not col_idx:
                    continue
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is None or cell.value == "":
                            continue
                        s = str(cell.value)
                        if s.endswith(".0") and s[:-2].isdigit():
                            s = s[:-2]
                        cell.value = s
                        cell.number_format = "@"

        for sheet_name in ("Datos", "Con fotos"):
            if sheet_name in wb.sheetnames:
                _force_text_skus(wb[sheet_name])

        # Embed screenshots solo en 'Con fotos'
        if "Con fotos" in wb.sheetnames:
            ws = wb["Con fotos"]
            headers = {cell.value: cell.column for cell in ws[1]}
            img_col_idx = headers.get("Imagen")
            url_col_idx = headers.get("URL")
            if img_col_idx:
                ws.column_dimensions[get_column_letter(img_col_idx)].width = 26
                for i, path in enumerate(screenshot_paths):
                    if not path or not Path(path).exists():
                        continue
                    ri = i + 2
                    ws.row_dimensions[ri].height = 160
                    try:
                        img = OpenpyxlImage(path); img.width = 170; img.height = 200
                        img.anchor = TwoCellAnchor(
                            editAs="oneCell",
                            _from=AnchorMarker(col=img_col_idx - 1, colOff=0, row=ri - 1, rowOff=0),
                            to=AnchorMarker(col=img_col_idx, colOff=0, row=ri, rowOff=0),
                        )
                        ws.add_image(img)
                    except Exception:
                        pass

            # URL truncation en 'Con fotos' (next=Imagen)
            if url_col_idx and img_col_idx:
                apply_url_truncation(ws, url_col_idx, img_col_idx, url_width=40, total_rows=len(rows) + 1)

        # URL truncation en 'Datos' (next = columna siguiente cualquiera)
        if "Datos" in wb.sheetnames:
            ws_d = wb["Datos"]
            headers_d = {cell.value: cell.column for cell in ws_d[1]}
            url_col_d = headers_d.get("URL")
            if url_col_d:
                apply_url_truncation(ws_d, url_col_d, url_col_d + 1, url_width=40, total_rows=len(rows) + 1)

        wb.save(output_path)
    except Exception:
        pass
    return output_path

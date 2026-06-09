# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
"""Helpers compartidos para el formato de los Excel de salida.

Lo que ofrece:
  - filter_and_reorder(df, columns): filtra el DataFrame a las columnas especificadas
    en ese orden. Columnas faltantes en df se crean vacías.
  - apply_url_truncation(ws, url_col_idx, next_col_idx): le da a la columna URL
    el estilo "texto clipeado" (no wrap, no shrink, no overflow). Para que Excel
    no derrame el texto a la columna siguiente, escribe " " en cualquier celda
    contigua que esté vacía.
  - download_once(path, files_module): dispara files.download() evitando descargas
    duplicadas del mismo archivo dentro de una ventana de debounce (Colab a veces
    re-dispara el JS de descarga al re-renderizar un Output widget).
"""
from __future__ import annotations

import time as _time

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Estética compartida de TODOS los Excel de salida (look unificado tipo Ferni):
# cabecera celeste con texto blanco negrita, auto-filtro y anchos auto-ajustados.
# NO inmoviliza paneles (freeze) — decisión de diseño del 2026-06-02.
HEADER_FILL = "2E86C1"  # celeste neutro


def apply_clean_style(ws, *, header_fill: str = HEADER_FILL,
                      skip_width=("URL", "Imagen"), min_w: int = 8,
                      max_w: int = 48, sample_rows: int = 300) -> None:
    """Aplica el estilo limpio unificado a una hoja ya escrita.

    - Cabecera: fondo `header_fill`, texto blanco negrita, centrado, alto 30.
    - Auto-filtro sobre toda la tabla.
    - Anchos auto-ajustados al contenido (cap [min_w, max_w]); salta las columnas
      en `skip_width` (URL / Imagen tienen su propio ancho fijo en cada engine).
    - SIN freeze_panes (los paneles inmovilizados se excluyeron a propósito).

    Defensivo: nunca lanza — si algo falla, deja la hoja como estaba (no debe
    romper el guardado del workbook, que suele traer imágenes embebidas)."""
    try:
        last_col, last_row = ws.max_column, ws.max_row
        if last_col < 1 or last_row < 1:
            return
        fill = PatternFill("solid", fgColor=header_fill)
        font = Font(color="FFFFFF", bold=True, size=11)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        names = {}
        for c in range(1, last_col + 1):
            cell = ws.cell(row=1, column=c)
            try:
                cell.fill = fill
                cell.font = font
                cell.alignment = align
            except Exception:
                pass
            names[c] = str(cell.value or "")
        try:
            ws.row_dimensions[1].height = 30
            ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
        except Exception:
            pass
        skip = set(skip_width)
        scan_to = min(last_row, 1 + max(0, sample_rows))
        for c in range(1, last_col + 1):
            if names[c] in skip:
                continue
            maxlen = len(names[c])
            for r in range(2, scan_to + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    ln = len(str(v))
                    if ln > maxlen:
                        maxlen = ln
            try:
                ws.column_dimensions[get_column_letter(c)].width = max(min_w, min(max_w, maxlen + 2))
            except Exception:
                pass
        # URL: ancho fijo 40 + texto clipeado (sin wrap) + " " en la celda
        # contigua vacía para que el texto largo NO se derrame a la derecha.
        try:
            url_c = next((c for c, n in names.items() if n == "URL"), None)
            if url_c:
                ws.column_dimensions[get_column_letter(url_c)].width = 40
                clip = Alignment(horizontal="left", vertical="center",
                                 wrap_text=False, shrink_to_fit=False)
                has_next = url_c < last_col
                for r in range(2, last_row + 1):
                    ws.cell(row=r, column=url_c).alignment = clip
                    if has_next:
                        nxt = ws.cell(row=r, column=url_c + 1)
                        if nxt.value in (None, ""):
                            nxt.value = " "
        except Exception:
            pass
    except Exception:
        pass


# path -> timestamp de la última descarga disparada
_LAST_DOWNLOADS: dict[str, float] = {}


def download_once(path, files_module, *, debounce_secs: float = 6.0) -> bool:
    """Dispara files_module.download(path) una sola vez por ventana de debounce.

    Devuelve True si disparó la descarga, False si la omitió por duplicado.
    El botón "descargar de nuevo" funciona igual porque el click manual ocurre
    bien después de la ventana de debounce.
    """
    if files_module is None:
        return False
    p = str(path)
    now = _time.time()
    last = _LAST_DOWNLOADS.get(p, 0.0)
    if now - last < debounce_secs:
        return False
    _LAST_DOWNLOADS[p] = now
    files_module.download(p)
    return True


def filter_and_reorder(df, columns: list[str]):
    """Devuelve df solo con las columnas dadas, en ese orden. Crea vacías las faltantes."""
    for c in columns:
        if c not in df.columns:
            df[c] = ""
    return df[columns].copy()


def apply_url_truncation(ws, url_col_idx: int, next_col_idx: int | None, *,
                         url_width: int = 40, total_rows: int | None = None) -> None:
    """Aplica el estilo de texto clipeado a la columna URL.

    - url_col_idx, next_col_idx: 1-based.
    - url_width: ancho de la columna URL (en unidades de Excel).
    - Si next_col_idx está dado, escribe " " en celdas vacías de esa columna
      para evitar overflow visual de URL.
    """
    align = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=False)
    ws.column_dimensions[get_column_letter(url_col_idx)].width = url_width

    if total_rows is None:
        total_rows = ws.max_row

    for ri in range(2, total_rows + 1):
        ws.cell(row=ri, column=url_col_idx).alignment = align
        if next_col_idx is not None:
            nxt = ws.cell(row=ri, column=next_col_idx)
            if nxt.value is None or nxt.value == "":
                nxt.value = " "


def _force_text_cols(ws, text_cols):
    """Fuerza columnas (por nombre) a formato texto, quitando el .0 de floats."""
    if not text_cols:
        return
    headers = {cell.value: cell.column for cell in ws[1]}
    for name in text_cols:
        ci = headers.get(name)
        if not ci:
            continue
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row:
                if cell.value in (None, ""):
                    continue
                s = str(cell.value)
                if s.endswith(".0") and s[:-2].isdigit():
                    s = s[:-2]
                cell.value = s
                cell.number_format = "@"


def write_two_sheets_df(df_no_img, rows, output_file, *, with_images=False,
                        image_path_key="Image Path", image_col="Imagen",
                        sheet_data="Datos", sheet_photos="Con fotos",
                        text_cols=(), img_w=160, img_h=200, row_h=160):
    """Escribe el Excel en formato 2-hojas (estilo MK7) para engines de DOM.

    - `df_no_img`: DataFrame SIN la columna de imagen.
    - `rows`: lista de dicts originales (mismo orden que df) — de cada uno se saca
      la ruta local del screenshot en `image_path_key`.
    - `with_images=False` → 1 hoja `sheet_data` (sin columna Imagen).
    - `with_images=True`  → 2 hojas: `sheet_data` (sin fotos) + `sheet_photos`
      (mismos datos + columna Imagen con el screenshot embebido).

    Aplica `apply_clean_style` a todas las hojas y URL truncada. Defensivo: si el
    embebido falla, no rompe el guardado. Devuelve True.
    """
    import os as _os
    import pandas as _pd  # noqa: F401  (asegura pandas en el entorno)
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

    with _pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_no_img.to_excel(writer, sheet_name=sheet_data, index=False)
        if with_images:
            di = df_no_img.copy()
            di[image_col] = ""  # última columna
            di.to_excel(writer, sheet_name=sheet_photos, index=False)

    wb = openpyxl.load_workbook(output_file)
    try:
        _force_text_cols(wb[sheet_data], text_cols)
        if with_images and sheet_photos in wb.sheetnames:
            ws = wb[sheet_photos]
            _force_text_cols(ws, text_cols)
            cols = list(df_no_img.columns) + [image_col]
            ii = cols.index(image_col) + 1
            ws.column_dimensions[get_column_letter(ii)].width = 25
            for r_i, rd in enumerate(rows, start=2):
                ip = rd.get(image_path_key, "") if isinstance(rd, dict) else ""
                if ip and _os.path.exists(ip):
                    ws.row_dimensions[r_i].height = row_h
                    try:
                        img = OpenpyxlImage(ip)
                        img.width, img.height = img_w, img_h
                        img.anchor = TwoCellAnchor(
                            editAs="oneCell",
                            _from=AnchorMarker(col=ii - 1, colOff=0, row=r_i - 1, rowOff=0),
                            to=AnchorMarker(col=ii, colOff=0, row=r_i, rowOff=0))
                        ws.add_image(img)
                    except Exception:
                        pass
            if "URL" in cols:
                apply_url_truncation(ws, cols.index("URL") + 1, ii,
                                     url_width=40, total_rows=len(rows) + 1)
        wsd = wb[sheet_data]
        dcols = list(df_no_img.columns)
        if "URL" in dcols:
            uc = dcols.index("URL") + 1
            apply_url_truncation(wsd, uc, uc + 1, url_width=40,
                                 total_rows=len(df_no_img) + 1)
        for sn in wb.sheetnames:
            apply_clean_style(wb[sn])
        wb.save(output_file)
    except Exception:
        pass
    return True

# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
"""Scraper interactivo de productos vendidos por Falabella, por sección.

Flujo:
  1. Setea una zona (Metropolitana / Las Condes por defecto). A diferencia de
     Sodimac, Falabella tiene precios nacionales, así que no se itera por tienda.
  2. Descubre las secciones del megamenú leyendo el __NEXT_DATA__ de la home
     (con fallback al scrape DOM del menú hamburguesa).
  3. El usuario elige una sección con un menú interactivo.
  4. Para cada subcategoría:
       a. Navega a /category/{catId}/{slug}
       b. Extrae __NEXT_DATA__.props.pageProps.results y .pagination
       c. Pagina con ?page=N hasta cubrir pagination.count
  5. Filtra por sellerName == "Falabella" (opcional, default sí).
  6. Descarga imagen de mediaUrls[0] de cada producto y la embebe en el Excel.
"""

import os
import re
import json
import asyncio
from pathlib import Path
from datetime import datetime

import pandas as pd
from playwright.async_api import async_playwright
from playwright_stealth import Stealth
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

import questionary
from questionary import Style as QStyle
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import (
    Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn, MofNCompleteColumn,
)
from rich import box

console = Console()

HOST = "https://www.falabella.com"
BASE_URL = "https://www.falabella.com/falabella-cl"

try:
    PROJECT_DIR = Path(__file__).resolve().parent
except NameError:
    PROJECT_DIR = Path.cwd()
SCREENSHOT_DIR = PROJECT_DIR / "cards_screenshots"
PARTIAL_DIR = PROJECT_DIR / "partial_runs"
PROJECT_DIR.mkdir(parents=True, exist_ok=True)
SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
PARTIAL_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_REGION = "Metropolitana"
DEFAULT_COMUNA = "Las Condes"
MAX_PAGES_PER_SUBCAT = 350
PER_PAGE_DEFAULT = 48

# Falabella tiene precios nacionales, así que iterar por "zona" raramente cambia
# el resultado. Mantenemos la misma lista que Sodimac (mismo grupo empresarial,
# mismo geofinder) para que el usuario tenga la misma UX de selección y, si
# quiere validar diferencias de stock/envío por región, pueda hacerlo.
ALL_STORES = [
    {"id": "E534", "name": "Antofagasta",   "region": "Antofagasta",   "comuna": "Antofagasta"},
    {"id": "E619", "name": "Arica",         "region": "Arica",         "comuna": "Arica"},
    {"id": "E633", "name": "Bio Bio",       "region": "Biobío",        "comuna": "Hualpén"},
    {"id": "E614", "name": "Calama",        "region": "Antofagasta",   "comuna": "Calama"},
    {"id": "E522", "name": "Cerrillos",     "region": "Metropolitana", "comuna": "Cerrillos"},
    {"id": "E988", "name": "Chicureo",      "region": "Metropolitana", "comuna": "Colina"},
    {"id": "E990", "name": "Chiguayante",   "region": "Biobío",        "comuna": "Chiguayante"},
    {"id": "E525", "name": "Chillán",       "region": "Ñuble",         "comuna": "Chillán"},
    {"id": "E760", "name": "Copiapó",       "region": "Atacama",       "comuna": "Copiapó"},
    {"id": "E983", "name": "Coronel",       "region": "Biobío",        "comuna": "Coronel"},
    {"id": "E511", "name": "Costanera",     "region": "Metropolitana", "comuna": "Providencia"},
    {"id": "E592", "name": "Curicó",        "region": "Maule",         "comuna": "Curicó"},
    {"id": "E781", "name": "El Belloto",    "region": "Valparaíso",    "comuna": "Quilpué"},
    {"id": "E513", "name": "El Llano",      "region": "Metropolitana", "comuna": "San Miguel"},
    {"id": "E510", "name": "Florida",       "region": "Metropolitana", "comuna": "La Florida"},
    {"id": "E502", "name": "Kennedy",       "region": "Metropolitana", "comuna": "Las Condes"},
    {"id": "E514", "name": "La Dehesa",     "region": "Metropolitana", "comuna": "Lo Barnechea"},
    {"id": "E512", "name": "La Reina",      "region": "Metropolitana", "comuna": "La Reina"},
    {"id": "E521", "name": "La Serena",     "region": "Coquimbo",      "comuna": "La Serena"},
    {"id": "E744", "name": "La Unión",      "region": "Ríos",          "comuna": "La Unión"},
    {"id": "E524", "name": "Linares",       "region": "Maule",         "comuna": "Linares"},
    {"id": "E900", "name": "Los Andes",     "region": "Valparaíso",    "comuna": "Los Andes"},
    {"id": "E529", "name": "Los Ángeles",   "region": "Biobío",        "comuna": "Los Ángeles"},
    {"id": "E503", "name": "Maipú",         "region": "Metropolitana", "comuna": "Maipú"},
    {"id": "E643", "name": "Ochagavía",     "region": "Metropolitana", "comuna": "Pedro Aguirre Cerda"},
    {"id": "E585", "name": "Osorno",        "region": "Lagos",         "comuna": "Osorno"},
    {"id": "E775", "name": "Portal Ñuñoa",  "region": "Metropolitana", "comuna": "Ñuñoa"},
    {"id": "E748", "name": "Portal Osorno", "region": "Lagos",         "comuna": "Osorno"},
    {"id": "E506", "name": "Portal Temuco", "region": "Araucanía",     "comuna": "Temuco"},
    {"id": "E659", "name": "Puente Alto",   "region": "Metropolitana", "comuna": "Puente Alto"},
    {"id": "E507", "name": "Puerto Montt",  "region": "Lagos",         "comuna": "Puerto Montt"},
    {"id": "E655", "name": "Quilicura",     "region": "Metropolitana", "comuna": "Quilicura"},
    {"id": "E518", "name": "Quilín",        "region": "Metropolitana", "comuna": "Peñalolén"},
    {"id": "E646", "name": "Quillota",      "region": "Valparaíso",    "comuna": "Quillota"},
    {"id": "E504", "name": "Rancagua",      "region": "O'Higgins",     "comuna": "Rancagua"},
    {"id": "E843", "name": "San Bernardo",  "region": "Metropolitana", "comuna": "San Bernardo"},
    {"id": "E874", "name": "Santa Amalia",  "region": "Metropolitana", "comuna": "La Florida"},
    {"id": "E591", "name": "Talca",         "region": "Maule",         "comuna": "Talca"},
    {"id": "E517", "name": "Temuco",        "region": "Araucanía",     "comuna": "Temuco"},
    {"id": "E520", "name": "Valparaíso",    "region": "Valparaíso",    "comuna": "Valparaíso"},
    {"id": "E830", "name": "Villarrica",    "region": "Araucanía",     "comuna": "Villarrica"},
    {"id": "E508", "name": "Viña del Mar",  "region": "Valparaíso",    "comuna": "Viña del Mar"},
]

QSTYLE = QStyle([
    ("qmark", "fg:#a4123f bold"),
    ("question", "bold"),
    ("answer", "fg:#00d97e bold"),
    ("pointer", "fg:#a4123f bold"),
    ("highlighted", "fg:#a4123f bold"),
    ("selected", "fg:#00d97e"),
    ("instruction", "fg:#888888"),
])


# ─────────────────────────────────────────  Persistence  ───────────────────

class PartialWriter:
    """Append-only JSONL writer so scrapes longer than minutes survive crashes."""
    def __init__(self, run_id, append=False):
        self.path = PARTIAL_DIR / f"{run_id}.jsonl"
        mode = "a" if append else "w"
        self._fh = open(self.path, mode, encoding="utf-8")
        self.count = 0

    def write(self, row):
        self._fh.write(json.dumps(row, ensure_ascii=False) + "\n")
        self._fh.flush()
        self.count += 1

    def close(self):
        try:
            self._fh.close()
        except Exception:
            pass

    @staticmethod
    def load(path):
        rows = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
        return rows


# ─────────────────────────────────────────  Zone & autocomplete  ───────────

async def _type_autocomplete(page, placeholder, value):
    sel = f'input[placeholder="{placeholder}"]'
    for _ in range(20):
        st = await page.evaluate(
            """(s) => { const i = document.querySelector(s);
                return i ? {present: true, disabled: i.disabled, hidden: i.offsetHeight === 0} : {present: false}; }""",
            sel,
        )
        if st.get("present") and not st.get("disabled") and not st.get("hidden"):
            break
        await page.wait_for_timeout(500)
    else:
        return False
    inp = page.locator(sel).first
    await page.evaluate("""() => {
        document.querySelectorAll(
            '[id*="onetrust"], [class*="onetrust"], '
            + '[id^="cookie"], [class^="cookie"], '
            + '#CybotCookiebotDialog, [class*="CookieConsent"]'
        ).forEach(e => { try { e.remove(); } catch (_) {} });
    }""")
    try:
        await inp.click(timeout=5000)
    except Exception:
        try:
            await inp.click(force=True, timeout=5000)
        except Exception:
            await page.evaluate(
                "(s) => { const el = document.querySelector(s); if (el) el.focus(); }",
                sel,
            )
    try:
        await inp.fill("", timeout=5000)
    except Exception:
        await page.evaluate(
            "(s) => { const el = document.querySelector(s); if (el) el.value = ''; }",
            sel,
        )
    await page.keyboard.type(value, delay=60)

    # Buscar match exacto/endsWith/contains. Si no aparece, ir borrando chars
    # del final (suggestions se relistan con menos filtro) hasta encontrar el
    # target real. Esto resuelve el caso "Calama" (Sodimac filtra agresivo y
    # esconde la opción correcta si tipeas el nombre completo).
    PICK_JS = """(target) => {
        const lis = [...document.querySelectorAll('li[class*="Autocomplete-module_suggestion"]')]
            .filter(e => e.offsetHeight > 0 && (e.innerText || '').trim());
        const norm = (s) => s.normalize("NFD").replace(/[\u0300-\u036f]/g, "").toLowerCase();
        const t = norm(target);
        const exact = lis.find(e => norm(e.innerText.trim()) === t);
        const endsWith = lis.find(e => {
            const x = norm(e.innerText.trim());
            return x === t || x.endsWith(" - " + t);
        });
        const contains = lis.find(e => norm(e.innerText.trim()).includes(t));
        const pick = exact || endsWith || contains;
        if (!pick) return null;
        const fire = (type) => pick.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
        fire('mousedown'); fire('mouseup'); fire('click');
        return pick.innerText.trim();
    }"""

    picked = None
    chars_left = len(value)
    while chars_left >= 3:
        for _ in range(12):
            await page.wait_for_timeout(250)
            has = await page.evaluate(
                """() => [...document.querySelectorAll('li[class*="Autocomplete-module_suggestion"]')]
                        .some(e => e.offsetHeight > 0 && (e.innerText||'').trim())"""
            )
            if has:
                break
        picked = await page.evaluate(PICK_JS, value)
        if picked:
            break
        await page.keyboard.press("Backspace")
        chars_left -= 1
        await page.wait_for_timeout(400)

    if picked:
        await page.wait_for_timeout(700)
        return True
    return False


async def _accept_cookies(page):
    """Falabella muestra un banner de cookies en el footer; lo cerramos si está."""
    try:
        await page.evaluate("""() => {
            const cands = [...document.querySelectorAll('button, a, [role="button"]')];
            const btn = cands.find(e => {
                const t = (e.innerText || '').trim().toLowerCase();
                return e.offsetHeight > 0 && (t === 'aceptar' || t === 'acepto' || t === 'aceptar todo' || t === 'aceptar todas');
            });
            if (btn) btn.click();
            document.querySelectorAll(
                '[id*="onetrust"], [class*="onetrust"], '
                + '[id^="cookie"], [class^="cookie"], '
                + '#CybotCookiebotDialog, [class*="CookieConsent"]'
            ).forEach(e => { try { e.remove(); } catch (_) {} });
        }""")
    except Exception:
        pass


async def set_zone(page, region, comuna):
    await page.goto(f"{BASE_URL}/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2500)
    await _accept_cookies(page)
    opened = await page.evaluate("""() => {
        // 1) Selector específico nuevo (UI 2026): p dentro de Zone-module_zone-lable
        const p = document.querySelector('p[class*="Zone-module_zone-lable"]');
        if (p && p.offsetHeight > 0) { p.click(); return true; }
        // 2) Fallback al método legacy: buscar elemento con texto exacto
        const el = [...document.querySelectorAll('*')].find(e => {
            const t = (e.innerText || '').trim();
            return e.offsetHeight > 0 && (
                t === 'Ingresa tu ubicación'
                || /^Entrega en/.test(t)
                || /^Despacha en/.test(t)
                || /^Envía a/.test(t)
            );
        });
        if (el) { el.click(); return true; } return false;
    }""")
    if not opened:
        return False
    await page.wait_for_timeout(1500)
    if not await _type_autocomplete(page, "Ingresa una Región", region): return False
    if not await _type_autocomplete(page, "Ingresa una Comuna", comuna): return False
    await page.evaluate("""() => {
        const btn = [...document.querySelectorAll('button')].find(b => b.innerText.trim() === 'Guardar' && !b.disabled);
        if (btn) btn.click();
    }""")
    await page.wait_for_timeout(3000)
    return True


async def set_zone_with_retry(page, region, comuna, retries=2):
    for attempt in range(retries + 1):
        try:
            ok = await set_zone(page, region, comuna)
        except Exception:
            ok = False
        if ok:
            return True
        if attempt < retries:
            await page.wait_for_timeout(1500 * (attempt + 1))
    return False


# ─────────────────────────────────────────  __NEXT_DATA__ helpers  ────────

def _extract_next_data(html):
    m = re.search(
        r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>',
        html, re.DOTALL,
    )
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None


CATEGORY_URL_RE = re.compile(r"/category/(cat\d+)/", re.I)


def _norm_category_url(url):
    """Asegura URL absoluta de Falabella para una categoría."""
    if not url:
        return ""
    if url.startswith("http"):
        return url
    if url.startswith("/falabella-cl/"):
        return HOST + url
    if url.startswith("/category/"):
        return BASE_URL + url
    if url.startswith("/"):
        return HOST + url
    return url


# ─────────────────────────────────────────  Discovery  ─────────────────────

EXCLUDED_SECTIONS_RE = re.compile(
    r'^lo último$|últim|mejores marcas|^ofertas$|^regalos$|cyber|black|liquidación|^outlet$|^home$|^inicio$|experiencias y servicios',
    re.I,
)


async def discover_sections(page):
    """Devuelve la estructura del megamenú de Falabella en 3 niveles:

        [(section_name, [
            (group_name, group_url, [
                (leaf_name, leaf_url), ...
            ]), ...
        ]), ...]

    - Si un grupo no tiene third_level_categories útiles, su única hoja es él
      mismo (leaf_name = group_name, leaf_url = group_url).
    - Se filtran entradas "Ver todo X" (is_ver_todo=true) que duplican el grupo.
    - Los L2 marketing-only sin hijos válidos se descartan.

    Lee taxonomy.entry.all_accesses.categories del __NEXT_DATA__ de la home.
    """
    await page.goto(f"{BASE_URL}/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2500)
    html = await page.content()
    m = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>', html, re.DOTALL)
    if not m:
        return []
    try:
        data = json.loads(m.group(1))
        cats = data["props"]["pageProps"]["serverData"]["headerData"]["taxonomy"]["entry"]["all_accesses"]["categories"]
    except Exception:
        return []

    def _ok_url(u):
        return bool(u) and u.startswith("http") and "falabella.com" in u and u not in ("#",)

    sections = []
    seen_names = set()
    for c in cats:
        title = (c.get("item_name") or "").strip()
        if not title or title in seen_names:
            continue
        if EXCLUDED_SECTIONS_RE.search(title):
            continue

        groups = []
        for s in c.get("second_level_categories", []) or []:
            g_name = (s.get("item_name") or "").strip()
            g_url = (s.get("item_url") or "").strip()
            # Saltar "Ver todo Hombre" / "Ver todo Tecnología", etc.
            if not g_name or re.match(r'^ver todo\b', g_name, re.I):
                continue

            leaves = []
            for t in s.get("third_level_categories", []) or []:
                if t.get("is_ver_todo"):
                    continue
                ln = (t.get("item_name") or "").strip()
                lu = (t.get("item_url") or "").strip()
                if ln and _ok_url(lu):
                    leaves.append((ln, lu))

            # Si el grupo no tiene hojas válidas pero tiene URL propio, usarlo como hoja.
            if not leaves:
                if _ok_url(g_url):
                    leaves = [(g_name, g_url)]
                else:
                    continue

            groups.append((g_name, g_url if _ok_url(g_url) else leaves[0][1], leaves))

        if not groups:
            continue
        seen_names.add(title)
        sections.append((title, groups))

    sections.sort(key=lambda x: x[0])
    return sections


def flatten_groups(groups):
    """Aplana [(group, group_url, [(leaf, leaf_url), ...]), ...] a lista de hojas
    con nombre 'Grupo / Hoja' (o solo 'Grupo' si la hoja es el grupo).

    Devuelve: [(subcat_name, subcat_url), ...]
    """
    out = []
    seen = set()
    for g_name, _g_url, leaves in groups:
        for ln, lu in leaves:
            label = ln if ln == g_name else f"{g_name} / {ln}"
            if lu in seen:
                continue
            seen.add(lu)
            out.append((label, lu))
    return out


# ─────────────────────────────────────────  Product extraction  ────────────

def _first_price(prices, ptype):
    """Devuelve el primer precio de prices[] cuyo type == ptype."""
    if not isinstance(prices, list):
        return ""
    for p in prices:
        if not isinstance(p, dict):
            continue
        if (p.get("type") or "").lower() == ptype.lower():
            v = p.get("price")
            if isinstance(v, list) and v:
                return v[0]
            if isinstance(v, (int, float, str)):
                return v
    return ""


def _to_int_clp(v):
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    digits = re.sub(r"\D", "", str(v))
    return int(digits) if digits else None


def _extract_product_from_json(result, section_name, subcat_name):
    """Transforma un item de results[] al formato de fila output."""
    prod_id = str(result.get("productId") or "").strip()
    sku_id = str(result.get("skuId") or prod_id).strip()
    display = (result.get("displayName") or "").strip()
    brand = ""
    b = result.get("brand")
    if isinstance(b, dict):
        brand = (b.get("name") or "").strip()
    elif isinstance(b, str):
        brand = b.strip()

    prices = result.get("prices") or []
    p_normal = _first_price(prices, "normalPrice")
    p_internet = _first_price(prices, "internetPrice")
    p_cmr = _first_price(prices, "cmrPrice")
    # Fallback de Precio Internet: muchos productos en promoción usan
    # `eventPrice` (precio de evento/oferta) en lugar de `internetPrice`.
    # También algunos usan `salePrice` o `crossedPrice` para el mismo.
    if not p_internet:
        for alt in ("eventPrice", "salePrice", "crossedPrice"):
            v = _first_price(prices, alt)
            if v:
                p_internet = v
                break
    # Fallback de Precio Normal
    if not p_normal:
        p_normal = _first_price(prices, "crossedPrice")

    pct = ""
    n_i = _to_int_clp(p_normal)
    n_int = _to_int_clp(p_internet)
    if n_i and n_int and n_i > n_int:
        pct = f"-{round((n_i - n_int) / n_i * 100)}%"

    badges = result.get("badges") or []
    badge_labels = []
    for bd in badges:
        if isinstance(bd, dict):
            lbl = (bd.get("label") or bd.get("text") or "").strip()
            if lbl:
                badge_labels.append(lbl)

    media = result.get("mediaUrls") or []
    img_url = media[0] if media else ""
    # Algunas variantes traen "images" en vez de mediaUrls
    if not img_url:
        imgs = result.get("images") or []
        if imgs and isinstance(imgs[0], dict):
            img_url = imgs[0].get("url") or ""
        elif imgs and isinstance(imgs[0], str):
            img_url = imgs[0]

    url = result.get("url") or ""
    if url and not url.startswith("http"):
        if url.startswith("/falabella-cl/"):
            url = HOST + url
        elif url.startswith("/"):
            url = HOST + url

    seller = (result.get("sellerName") or result.get("seller") or "").strip()
    if not seller:
        sid = (result.get("sellerId") or "").strip()
        if sid:
            seller = sid

    return {
        "Sección": section_name,
        "Subcategoría": subcat_name,
        "Vendedor": seller,
        "Marca": brand,
        "SKU": sku_id,
        "Product ID": prod_id,
        "Descripción Producto": display,
        "Precio Normal": _to_int_clp(p_normal) or "",
        "Precio Internet": _to_int_clp(p_internet) or "",
        "Precio CMR": _to_int_clp(p_cmr) or "",
        "% Descuento": pct,
        "Rating": result.get("rating") or "",
        "Reviews": result.get("totalReviews") or "",
        "En Stock": "No" if result.get("isOutOfStock") else "Si",
        "Badges": " | ".join(badge_labels),
        "URL": url,
        "Imagen URL": img_url,
    }

# ─────────────────────────────────────────  Image download  ────────────────

def _download_image(url, dest_path, timeout=10):
    """Descarga una imagen a dest_path. Devuelve True si OK."""
    if not url:
        return False
    if dest_path.exists() and dest_path.stat().st_size > 0:
        return True
    try:
        import requests
        r = requests.get(url, timeout=timeout, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        })
        if r.status_code != 200 or not r.content:
            return False
        dest_path.write_bytes(r.content)
        return True
    except Exception:
        return False


# ─────────────────────────────────────────  Scraping  ──────────────────────

async def _safe_goto(page, url, retries=2, timeout=60000):
    for attempt in range(retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            return True
        except Exception:
            if attempt < retries:
                await page.wait_for_timeout(1500 * (attempt + 1))
    return False


async def _extract_results_and_pagination(page):
    """Lee __NEXT_DATA__ del DOM actual y devuelve (results, pagination_dict)."""
    data = await page.evaluate("""() => {
        const el = document.getElementById('__NEXT_DATA__');
        if (!el) return null;
        try { return JSON.parse(el.textContent); } catch (e) { return null; }
    }""")
    if not data:
        return [], {}
    try:
        pp = data["props"]["pageProps"]
    except Exception:
        return [], {}
    results = pp.get("results") or []
    if not isinstance(results, list):
        results = []
    pagination = pp.get("pagination") or {}
    return results, pagination


async def scrape_subcat(page, section_name, subcat_name, subcat_url, progress=None, page_task=None,
                        download_images=True, only_falabella=True):
    """Scrapea una subcategoría paginando por ?page=N.

    Devuelve dict: {rows, pages, truncated, failed, empty}
    """
    result = {"rows": [], "pages": 0, "truncated": False, "failed": False, "empty": False}

    # Construir URL base sin page param para poder añadirla
    base_url = re.sub(r"([?&])page=\d+&?", r"\1", subcat_url).rstrip("?&")
    # Quitar trailing ? si quedó vacío
    if base_url.endswith("?"):
        base_url = base_url[:-1]

    seen_in_subcat = set()
    total_count = None
    per_page = PER_PAGE_DEFAULT
    page_num = 1

    while True:
        sep = "&" if "?" in base_url else "?"
        page_url = base_url if page_num == 1 else f"{base_url}{sep}page={page_num}"

        if not await _safe_goto(page, page_url):
            if page_num == 1:
                result["failed"] = True
            break

        await page.wait_for_timeout(1200)
        results, pagination = await _extract_results_and_pagination(page)

        if not results:
            if page_num == 1:
                result["empty"] = True
            break

        if total_count is None:
            try:
                total_count = int(pagination.get("count") or 0)
                per_page = int(pagination.get("perPage") or PER_PAGE_DEFAULT) or PER_PAGE_DEFAULT
            except Exception:
                total_count = 0

        new_in_page = 0
        for r in results:
            if not isinstance(r, dict):
                continue
            row = _extract_product_from_json(r, section_name, subcat_name)
            sku = row.get("SKU") or row.get("Product ID")
            if not sku or sku in seen_in_subcat:
                continue
            seen_in_subcat.add(sku)

            if download_images and row.get("Imagen URL"):
                ext = ".jpg"
                img_path = SCREENSHOT_DIR / f"{sku}{ext}"
                if _download_image(row["Imagen URL"], img_path):
                    row["Image Path"] = str(img_path)
                else:
                    row["Image Path"] = ""
            else:
                row["Image Path"] = ""

            result["rows"].append(row)
            new_in_page += 1

        result["pages"] = page_num

        if progress is not None and page_task is not None:
            try:
                progress.update(page_task, description=f"  └─ {subcat_name} · pág {page_num} · {len(result['rows'])} cards")
            except Exception:
                pass

        # ¿Hay más páginas?
        if total_count and per_page:
            max_page = (total_count + per_page - 1) // per_page
        else:
            max_page = MAX_PAGES_PER_SUBCAT

        if page_num >= max_page:
            break
        if new_in_page == 0:
            break

        page_num += 1
        if page_num > MAX_PAGES_PER_SUBCAT:
            result["truncated"] = True
            break

    return result


# ─────────────────────────────────────────  Zone picker  ──────────────────

def pick_stores():
    """Devuelve la lista de zonas seleccionadas por el usuario.

    Falabella tiene precios nacionales, así que en la práctica una sola zona
    alcanza. Se incluye la misma lista que Sodimac por consistencia de UX y
    por si el usuario quiere validar disponibilidad/envío por región.
    """
    rm_stores = [s for s in ALL_STORES if s["region"] == "Metropolitana"]
    presets = [
        ("Solo Kennedy / Las Condes (default — más rápido)", [s for s in ALL_STORES if s["id"] == "E502"]),
        (f"Todas RM ({len(rm_stores)} zonas)", rm_stores),
        (f"Todas Chile ({len(ALL_STORES)} zonas)", ALL_STORES),
        ("Personalizado (elegir manualmente)", None),
    ]
    choice = questionary.select(
        "¿Qué zonas querés scrapear?",
        choices=[questionary.Choice(title=t, value=i) for i, (t, _) in enumerate(presets)],
        style=QSTYLE, instruction="(↑↓ para moverte, Enter para elegir)",
    ).ask()
    if choice is None:
        return None
    _, stores = presets[choice]
    if stores is not None:
        return stores
    chs = [
        questionary.Choice(
            title=f"{s['id']}  {s['name']:<14}  ({s['region']} / {s['comuna']})",
            value=s,
            checked=(s["id"] == "E502"),
        )
        for s in ALL_STORES
    ]
    selected = questionary.checkbox(
        "Marcá las zonas con [Espacio], confirmá con [Enter]:",
        choices=chs, style=QSTYLE,
    ).ask()
    return selected or []


# ─────────────────────────────────────────  Excel  ─────────────────────────

def write_excel(rows, output_file):
    if not rows:
        return False
    df = pd.DataFrame(rows)
    # Si hay columnas de zona, van primero (igual que Sodimac).
    leading = [c for c in ("Tienda", "Nombre Tienda") if c in df.columns]
    # Orden de columnas deseado para el resto
    desired = [
        "Sección", "Subcategoría", "Vendedor", "Marca", "SKU", "Product ID",
        "Descripción Producto",
        "Precio Normal", "Precio Internet", "Precio CMR", "% Descuento",
        "Rating", "Reviews", "En Stock", "Badges",
        "URL", "Imagen URL", "Image Path",
    ]
    cols = [c for c in desired if c in df.columns and c not in leading]
    extras = [c for c in df.columns if c not in leading + cols]
    df = df[leading + cols + extras]
    df["Imagen"] = ""
    df.to_excel(output_file, index=False)
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    try:
        ipi = list(df.columns).index("Image Path") + 1
        ii = list(df.columns).index("Imagen") + 1
        ws.column_dimensions[openpyxl.utils.get_column_letter(ii)].width = 25
        for ri, rd in enumerate(rows, start=2):
            ws.row_dimensions[ri].height = 160
            ip = rd.get("Image Path", "")
            if ip and os.path.exists(ip):
                try:
                    img = OpenpyxlImage(ip); img.width = 160; img.height = 200
                    img.anchor = TwoCellAnchor(
                        editAs="oneCell",
                        _from=AnchorMarker(col=ii - 1, colOff=0, row=ri - 1, rowOff=0),
                        to=AnchorMarker(col=ii, colOff=0, row=ri, rowOff=0),
                    )
                    ws.add_image(img)
                except Exception:
                    pass
        ws.delete_cols(ipi)
        from ._excel_utils import apply_clean_style
        for _sn in wb.sheetnames:
            apply_clean_style(wb[_sn])
        wb.save(output_file)
    except Exception as e:
        console.print(f"[yellow]Aviso embebiendo imágenes: {e}[/]")
    return True


# ─────────────────────────────────────────  UI  ────────────────────────────

def banner():
    console.print()
    console.print(Panel.fit(
        "[bold #a4123f]FALABELLA SELLER SCRAPER[/]\n"
        "[dim]Captura productos vendidos por Falabella, por sección[/]",
        border_style="#a4123f", padding=(1, 4),
    ))
    console.print()


def section_summary(rows, section_name, output_file):
    t = Table(title=f"Resumen — {section_name}", box=box.SIMPLE_HEAVY, show_header=True, header_style="bold #a4123f")
    t.add_column("Métrica", style="white")
    t.add_column("Valor", style="bold green", justify="right")
    t.add_row("Filas totales", str(len(rows)))
    t.add_row("SKUs únicos", str(len({r["SKU"] for r in rows if r.get("SKU")})))
    if any(r.get("Tienda") for r in rows):
        t.add_row("Zonas en el dataset", str(len({r["Tienda"] for r in rows if r.get("Tienda")})))
    t.add_row("Subcategorías cubiertas", str(len({r["Subcategoría"] for r in rows if r.get("Subcategoría")})))
    t.add_row("Con Precio CMR", str(sum(1 for r in rows if r.get("Precio CMR"))))
    t.add_row("Con % Descuento", str(sum(1 for r in rows if r.get("% Descuento"))))
    t.add_row("Con Rating", str(sum(1 for r in rows if r.get("Rating"))))
    t.add_row("Con Imagen", str(sum(1 for r in rows if r.get("Image Path"))))
    console.print(t)
    console.print(f"\n[bold green]✓ Excel guardado en:[/] [cyan]{output_file}[/]")


# ─────────────────────────────────────────  Main  ──────────────────────────

async def main():
    banner()

    stores = pick_stores()
    if not stores:
        console.print("[yellow]No seleccionaste zonas. Cancelado.[/]")
        return
    console.print(f"[green]✓[/] {len(stores)} zona(s) seleccionada(s): " +
                  ", ".join(s["name"] for s in stores))

    only_fala = questionary.confirm(
        "¿Filtrar solo productos vendidos por Falabella?",
        default=True, style=QSTYLE,
    ).ask()
    if only_fala is None:
        return

    download_images = questionary.confirm(
        "¿Descargar imágenes de productos y embeberlas en el Excel?",
        default=True, style=QSTYLE,
    ).ask()
    if download_images is None:
        return

    async with Stealth().use_async(async_playwright()) as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900}, color_scheme="light",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await ctx.new_page()

        first = stores[0]
        with console.status(f"[cyan]Configurando zona {first['region']} / {first['comuna']}…[/]", spinner="dots"):
            ok = await set_zone_with_retry(page, first["region"], first["comuna"])
        if not ok:
            console.print(f"[yellow]No pude fijar zona inicial ({first['name']}). Sigo (Falabella tiene precios nacionales).[/]")
        else:
            console.print(f"[green]✓[/] Zona inicial: {first['name']} ({first['comuna']})")

        with console.status("[cyan]Descubriendo secciones del menú de Falabella…[/]", spinner="dots"):
            sections = await discover_sections(page)
        if not sections:
            console.print("[red]No pude leer las secciones.[/]")
            await browser.close(); return
        console.print(f"[green]✓[/] {len(sections)} secciones detectadas")

        choices = [
            questionary.Choice(
                title=f"{n}  [dim]({sum(len(l) for _, _, l in groups)} subcat. en {len(groups)} grupos)[/]",
                value=n,
            )
            for n, groups in sections
        ]
        section_name = questionary.select(
            "¿Qué sección querés scrapear?",
            choices=choices, style=QSTYLE,
            instruction="(↑↓ para moverte, Enter para elegir)",
        ).ask()
        if not section_name:
            await browser.close(); return
        groups = next(g for n, g in sections if n == section_name)

        # Pickear subcategorías agrupadas por grupo (L2). Separadores muestran el grupo.
        group_choices = [questionary.Separator(f"── {g_name} ──") for g_name, _, _ in []]
        choice_items = []
        for g_name, _g_url, leaves in groups:
            choice_items.append(questionary.Separator(f"── {g_name} ──"))
            for ln, lu in leaves:
                label = ln if ln == g_name else f"{g_name} / {ln}"
                choice_items.append(questionary.Choice(title=ln, value=(label, lu), checked=True))
        selected = questionary.checkbox(
            f"Subcategorías a scrapear de [{section_name}]",
            choices=choice_items, style=QSTYLE,
            instruction="(Espacio: marcar/desmarcar, A: todas, I: invertir, Enter: confirmar)",
        ).ask()
        if not selected:
            console.print("[yellow]No marcaste ninguna subcategoría. Cancelado.[/]")
            await browser.close(); return
        subcats = selected
        console.print(f"[green]✓[/] Sección: [bold]{section_name}[/] · {len(subcats)} subcategorías seleccionadas")

        all_rows = []
        non_seller = 0
        skipped_zone = []
        failed_subcats = []
        truncated_subcats = []

        run_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{re.sub(r'[^A-Za-z0-9]+', '', section_name) or 'run'}"
        partial = PartialWriter(run_id)
        console.print(f"[dim]Persistencia incremental: {partial.path}[/]")

        with Progress(
            SpinnerColumn(),
            TextColumn("[bold #a4123f]{task.description}"),
            BarColumn(bar_width=30),
            MofNCompleteColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TimeElapsedColumn(),
            console=console,
        ) as progress:
            store_task = progress.add_task("Zonas", total=len(stores))
            cat_task = progress.add_task("Subcategorías", total=len(subcats))
            page_task = progress.add_task("Esperando…", total=None)

            for st_idx, store in enumerate(stores):
                progress.update(store_task, description=f"Zona: {store['name']} ({store['comuna']})")
                if st_idx > 0:
                    ok = await set_zone_with_retry(page, store["region"], store["comuna"])
                    if not ok:
                        skipped_zone.append(store["id"])
                        progress.advance(store_task)
                        continue

                seen_skus_in_store = set()
                progress.reset(cat_task, total=len(subcats))
                for sc_name, sc_url in subcats:
                    progress.update(cat_task, description=f"  └─ {sc_name}")
                    res = await scrape_subcat(
                        page, section_name, sc_name, sc_url,
                        progress, page_task,
                        download_images=download_images, only_falabella=only_fala,
                    )
                    if res["failed"]:
                        failed_subcats.append((store["id"], sc_name))
                    if res["truncated"]:
                        truncated_subcats.append((store["id"], sc_name))
                    for r in res["rows"]:
                        if only_fala:
                            vendor = (r.get("Vendedor") or "").strip().upper()
                            if "FALABELLA" not in vendor:
                                non_seller += 1
                                continue
                        sku = r.get("SKU")
                        if not sku or sku in seen_skus_in_store:
                            continue
                        seen_skus_in_store.add(sku)
                        r_copy = {"Tienda": store["id"], "Nombre Tienda": store["name"], **r}
                        all_rows.append(r_copy)
                        partial.write(r_copy)
                    progress.advance(cat_task)
                progress.advance(store_task)
            progress.remove_task(page_task)

        await browser.close()
        partial.close()

    if non_seller:
        console.print(f"[dim]Cards descartadas (no vendidas por Falabella): {non_seller}[/]")
    if skipped_zone:
        console.print(f"[yellow]Zonas saltadas por falla de set_zone: {', '.join(skipped_zone)}[/]")
    if failed_subcats:
        console.print(f"[yellow]Subcategorías que fallaron al cargar: {len(failed_subcats)}[/]")
        for store_id, sc in failed_subcats[:10]:
            console.print(f"   · {store_id} · {sc}")
    if truncated_subcats:
        console.print(f"[yellow]Subcategorías truncadas en {MAX_PAGES_PER_SUBCAT} pág: {len(truncated_subcats)}[/]")
        for store_id, sc in truncated_subcats[:10]:
            console.print(f"   · {store_id} · {sc}")

    if not all_rows:
        console.print("[yellow]No se encontraron productos.[/]")
        return

    unique_skus = len({r["SKU"] for r in all_rows if r.get("SKU")})
    console.print(f"[dim]{len(all_rows)} filas totales · {unique_skus} SKUs únicos[/]")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    safe = re.sub(r'[^\w\s-]', '', section_name).strip().replace(' ', '_')
    suffix = "" if only_fala else "_all_sellers"
    suffix += "_con_imgs" if download_images else ""
    output = PROJECT_DIR / f"falabella_{safe}{suffix}_{timestamp}.xlsx"
    with console.status("[cyan]Escribiendo Excel con imágenes…[/]", spinner="dots"):
        write_excel(all_rows, str(output))

    section_summary(all_rows, section_name, output)

# ============== MK6 layer ==============

"""Falabella MK6 — busca un listado de SKUs en Falabella y devuelve precios
en un Excel.

Diferencias vs MK6 Sodimac/Construmart:
  - Falabella tiene **precios nacionales** → por default 1 sola zona. Multi-zona
    sólo afecta disponibilidad/envío.
  - Búsqueda batch via /falabella-cl/search?Ntt=SKU1+SKU2+… que devuelve un
    __NEXT_DATA__ con props.pageProps.results (mismo shape que las PLPs).
  - Output: Excel 2 hojas (Datos + Con fotos), 1 fila por (SKU × zona).
"""
import os, re, sys, asyncio
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

# En Colab, scraper_falabella fue ejecutado arriba — definimos proxy y dirs.
THIS_DIR = Path.cwd()
class _FaProxy: pass
fa = _FaProxy()
for _name in ("set_zone_with_retry", "_extract_results_and_pagination",
              "_extract_product_from_json", "ALL_STORES"):
    setattr(fa, _name, globals().get(_name))


from playwright.async_api import async_playwright
from playwright_stealth import Stealth

HOST = "https://www.falabella.com"
BASE_URL = "https://www.falabella.com/falabella-cl"
SEARCH_URL = BASE_URL + "/search?Ntt={q}"

SCREENSHOT_DIR = THIS_DIR / "mk6_screenshots"
SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)

BATCH_SIZE = 30  # /search soporta muchos términos; Falabella muestra 48/pag

SKU_COL_ALIASES = [
    "SKU Falabella", "SKU", "SKU Producto", "Cód. Producto",
    "Código Producto", "Cod Producto", "skuId", "productId",
]
EASY_COL_ALIASES = ["SKU Easy", "Cód. Easy", "Codigo Easy", "Cod Easy", "Easy"]
DESC_COL_ALIASES = ["Desc. Producto", "Desc Producto", "Descripción", "Descripcion",
                    "Descripcion Easy", "Descripción Easy"]


# ─────────────────────────────────────────  Input  ─────────────────────────

def _norm_sku(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() in ("por definir", "pordefinir", "nan", "none"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s


def _pick_col(df, aliases):
    cols = {c.lower().strip(): c for c in df.columns}
    for a in aliases:
        if a.lower().strip() in cols:
            return cols[a.lower().strip()]
    return None


def read_input(path):
    p = Path(path)
    if p.suffix.lower() in (".csv", ".tsv"):
        df = pd.read_csv(p, dtype=str)
    else:
        df = pd.read_excel(p, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    sku_col = _pick_col(df, SKU_COL_ALIASES)
    easy_col = _pick_col(df, EASY_COL_ALIASES)
    desc_col = _pick_col(df, DESC_COL_ALIASES)
    if not sku_col:
        raise ValueError(f"Input no tiene columna de SKU. Aliases: {SKU_COL_ALIASES}\n"
                         f"Columnas presentes: {list(df.columns)}")
    df[sku_col] = df[sku_col].apply(_norm_sku).astype(str).str.strip()
    df = df[df[sku_col] != ""].reset_index(drop=True)
    if not easy_col:
        df["SKU Easy"] = ""; easy_col = "SKU Easy"
    else:
        df[easy_col] = df[easy_col].apply(_norm_sku).astype(str)
    if not desc_col:
        df["Desc. Producto"] = ""; desc_col = "Desc. Producto"
    return df, desc_col, sku_col, easy_col


# ─────────────────────────────────────────  Lookup engine  ─────────────────

async def lookup_batch(page, skus, screenshot_dir=None):
    """Busca múltiples SKUs juntos. Devuelve dict {sku: raw_row}."""
    found = {}
    if not skus:
        return found
    query = "+".join(str(s) for s in skus)
    url = SEARCH_URL.format(q=query)
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=45000)
    except Exception:
        return found
    await page.wait_for_timeout(1000)

    # Si nada hizo match, Falabella redirige a /noResult
    if "/noResult" in page.url:
        return found

    results, _pag = await fa._extract_results_and_pagination(page)
    by_sku = {}
    for r in results:
        if not isinstance(r, dict):
            continue
        sid = str(r.get("skuId") or r.get("productId") or "").strip()
        pid = str(r.get("productId") or "").strip()
        if sid:
            by_sku[sid] = r
        if pid and pid not in by_sku:
            by_sku[pid] = r

    for s in skus:
        if str(s) in by_sku:
            r = by_sku[str(s)]
            found[str(s)] = r

    # Screenshots: capturar las cards de los SKUs matched
    if screenshot_dir and found:
        cards = await page.query_selector_all(
            "[class*='pod'], [data-pod], li[id^='testId-pod']"
        )
        # Falabella PLP usa cards con class 'pod' o '_pod-style'.
        # Si los selectores no matchean, fallback: usar los grid items.
        if not cards:
            cards = await page.query_selector_all("ol > li, .grid-pod, .jsx-pod")
        # Match cards a results por orden
        n = min(len(cards), len(results))
        for i in range(n):
            r = results[i]
            sid = str(r.get("skuId") or r.get("productId") or "").strip()
            if sid not in found:
                continue
            shot = screenshot_dir / f"{sid}.jpg"
            if shot.exists() and shot.stat().st_size > 0:
                found[sid]["_screenshot"] = str(shot)
                continue
            try:
                await cards[i].scroll_into_view_if_needed()
                await page.wait_for_timeout(80)
                await cards[i].screenshot(path=str(shot), type="jpeg", quality=78)
                found[sid]["_screenshot"] = str(shot)
            except Exception:
                pass
    return found


# ─────────────────────────────────────────  Rows  ──────────────────────────

def _row_from_result(r, sku, zone, easy, desc):
    """Reusa fa._extract_product_from_json y agrega columnas MK6."""
    base = fa._extract_product_from_json(r, "", "")
    return {
        "Zona": zone["name"] if zone else "Nacional",
        "Comuna": (zone or {}).get("comuna", ""),
        "SKU Easy": easy or "",
        "Desc. Producto": desc or "",
        "SKU Falabella": sku,
        "Vendedor": base.get("Vendedor", ""),
        "Marca": base.get("Marca", ""),
        "Descripción Producto": base.get("Descripción Producto", ""),
        "Precio Normal": base.get("Precio Normal", ""),
        "Precio Internet": base.get("Precio Internet", ""),
        "Precio CMR": base.get("Precio CMR", ""),
        "% Descuento": base.get("% Descuento", ""),
        "Rating": base.get("Rating", ""),
        "Reviews": base.get("Reviews", ""),
        "En Stock": base.get("En Stock", ""),
        "Badges": base.get("Badges", ""),
        "URL": base.get("URL", ""),
        "Source": "search",
        "Image Path": r.get("_screenshot", ""),
    }


def _empty_row(sku, zone, easy, desc):
    return {
        "Zona": zone["name"] if zone else "Nacional",
        "Comuna": (zone or {}).get("comuna", ""),
        "SKU Easy": easy or "",
        "Desc. Producto": desc or "",
        "SKU Falabella": sku,
        "Vendedor": "", "Marca": "", "Descripción Producto": "",
        "Precio Normal": "", "Precio Internet": "", "Precio CMR": "",
        "% Descuento": "", "Rating": "", "Reviews": "", "En Stock": "",
        "Badges": "", "URL": "", "Source": "no encontrado", "Image Path": "",
    }


async def search_skus(skus_with_meta, zones=None, screenshot=True, headless=True,
                      progress_cb=None, skip_pairs=None, on_row=None, status_cb=None):
    """zones: lista de {id, name, region, comuna}. Si None → 1 zona default.
    skus_with_meta: list[{sku, easy, desc}].
    skip_pairs: set de (zone_id, sku) ya completados (para resume).
    on_row: callable(row_dict) — invocado por cada fila apenas se genera.
    """
    if not zones:
        zones = [{"id":"E502","name":"Kennedy","region":"Metropolitana","comuna":"Las Condes"}]
    all_rows = []
    skip_pairs = skip_pairs or set()
    async with Stealth().use_async(async_playwright()) as pw:
        browser = await pw.chromium.launch(headless=headless)
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900}, color_scheme="light",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await ctx.new_page()

        for zi, zone in enumerate(zones):
            zone_id = str(zone.get("id") or "")
            pending = [m for m in skus_with_meta
                       if (zone_id, str(m["sku"])) not in skip_pairs]
            if not pending:
                if progress_cb:
                    progress_cb(zi + 1, len(zones), zone, len(skus_with_meta), len(all_rows))
                continue

            if status_cb: status_cb(f"🔧 Cambiando zona a {zone.get('name','?')}…")
            ok = await fa.set_zone_with_retry(page, zone["region"], zone["comuna"])
            if not ok and len(zones) == 1:
                # Si falla la zona pero es la única, seguimos igual: precios nacionales
                pass

            zone_dir = SCREENSHOT_DIR / (zone.get("id") or "default")
            zone_dir.mkdir(parents=True, exist_ok=True)
            meta_by_sku = {m["sku"]: m for m in pending}
            skus_all = [m["sku"] for m in pending]
            done = 0
            for batch_start in range(0, len(skus_all), BATCH_SIZE):
                batch = skus_all[batch_start: batch_start + BATCH_SIZE]
                if status_cb:
                    bn = batch_start // BATCH_SIZE + 1
                    btot = (len(skus_all) + BATCH_SIZE - 1) // BATCH_SIZE
                    status_cb(f"🔍 {zone.get('name','?')} · buscando lote {bn}/{btot} ({len(batch)} SKUs)…")
                found = await lookup_batch(
                    page, batch,
                    screenshot_dir=zone_dir if screenshot else None,
                )
                for sku in batch:
                    meta = meta_by_sku[sku]
                    r = found.get(sku)
                    if r:
                        row = _row_from_result(r, sku, zone,
                                                meta.get("easy"), meta.get("desc"))
                    else:
                        row = _empty_row(sku, zone, meta.get("easy"), meta.get("desc"))
                    all_rows.append(row)
                    if on_row: on_row(row)
                done += len(batch)
                if progress_cb:
                    progress_cb(zi + 1, len(zones), zone, done, len(all_rows))
        await browser.close()
    return all_rows


# ─────────────────────────────────────────  Output  ────────────────────────

# Columnas del Excel de output (Imagen se agrega solo en la hoja "Con fotos").
OUTPUT_COLS = [
    "Zona", "Comuna", "SKU Easy", "Desc. Producto", "SKU Falabella",
    "Vendedor", "Marca", "Descripción Producto",
    "Precio Normal", "Precio Internet", "Precio CMR", "% Descuento",
    "URL",
]


def write_output(rows, output_path):
    """Excel 2 hojas: 'Datos' (sin imágenes, fácil de filtrar) y 'Con fotos' (screenshots embebidos)."""
    if not rows:
        raise ValueError("No hay filas para escribir.")
    from ._excel_utils import filter_and_reorder, apply_url_truncation

    df = pd.DataFrame(rows)
    df_data = filter_and_reorder(df, OUTPUT_COLS)
    df_data.to_excel(output_path, index=False, sheet_name="Datos")

    wb = openpyxl.load_workbook(output_path)
    ws1 = wb["Datos"]
    # Hoja "Con fotos" = mismas columnas + Imagen como última
    ws2 = wb.create_sheet("Con fotos")
    photo_cols = OUTPUT_COLS + ["Imagen"]
    ws2.append(photo_cols)
    img_col_idx = len(photo_cols)
    url_col_idx_photos = photo_cols.index("URL") + 1
    url_col_idx_datos = OUTPUT_COLS.index("URL") + 1
    ws2.column_dimensions[openpyxl.utils.get_column_letter(img_col_idx)].width = 30

    for ri, rd in enumerate(rows, start=2):
        for ci, col in enumerate(OUTPUT_COLS, start=1):
            ws2.cell(row=ri, column=ci, value=rd.get(col, ""))
        ip = rd.get("Image Path", "")
        if ip and os.path.exists(ip):
            ws2.row_dimensions[ri].height = 220
            try:
                img = OpenpyxlImage(ip); img.width = 170; img.height = 200
                img.anchor = TwoCellAnchor(
                    editAs="oneCell",
                    _from=AnchorMarker(col=img_col_idx - 1, colOff=0, row=ri - 1, rowOff=0),
                    to=AnchorMarker(col=img_col_idx, colOff=0, row=ri, rowOff=0),
                )
                ws2.add_image(img)
            except Exception:
                pass

    # URL truncado en ambas hojas (en "Datos" usa columna inexistente como next para forzar bloqueo
    # con espacio en blanco; en "Con fotos" usa Imagen como next).
    apply_url_truncation(ws2, url_col_idx_photos, img_col_idx, url_width=40, total_rows=len(rows) + 1)
    apply_url_truncation(ws1, url_col_idx_datos, url_col_idx_datos + 1, url_width=40, total_rows=len(rows) + 1)

    # Forzar SKUs como texto en ambas hojas
    for sheet in (ws1, ws2):
        cols_in_sheet = OUTPUT_COLS if sheet is ws1 else photo_cols
        for col_name in ("SKU Easy", "SKU Falabella"):
            if col_name in cols_in_sheet:
                cidx = cols_in_sheet.index(col_name) + 1
                for ri in range(2, len(rows) + 2):
                    c = sheet.cell(row=ri, column=cidx)
                    c.number_format = "@"
                    if c.value is not None:
                        c.value = str(c.value)
    from ._excel_utils import apply_clean_style
    for _sn in wb.sheetnames:
        apply_clean_style(wb[_sn])
    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────  CLI  ───────────────────────────

def _print(*a):
    print(*a, flush=True)


async def _amain(input_path, zone_ids=None, screenshot=True, output_path=None,
                 headless=True):
    df, desc_col, sku_col, easy_col = read_input(input_path)
    _print(f"[{datetime.now():%H:%M:%S}] input: {input_path}")
    _print(f"  cols: SKU='{sku_col}' Easy='{easy_col}' Desc='{desc_col}'")
    skus_with_meta = []
    seen = set()
    for _, r in df.iterrows():
        sku = str(r[sku_col]).strip()
        if not sku or sku in seen:
            continue
        seen.add(sku)
        skus_with_meta.append({"sku": sku, "easy": r.get(easy_col, ""),
                                "desc": r.get(desc_col, "")})
    _print(f"  {len(skus_with_meta)} SKUs únicos.")

    if zone_ids:
        zones = [z for z in fa.ALL_STORES if z["id"] in zone_ids]
        if not zones:
            raise SystemExit(f"Ningún zone_id válido en {zone_ids}")
    else:
        zones = [z for z in fa.ALL_STORES if z["id"] == "E502"] or [fa.ALL_STORES[0]]
    _print(f"  zonas: {[z['name'] for z in zones]}")

    t0 = datetime.now()
    def _cb(zi, n_z, zone, ki, total_rows):
        _print(f"  [{datetime.now():%H:%M:%S}] {zi}/{n_z} {zone['name']} "
               f"sku {ki}/{len(skus_with_meta)}  ({total_rows} filas)")

    rows = await search_skus(skus_with_meta, zones, screenshot=screenshot,
                              headless=headless, progress_cb=_cb)
    dt = (datetime.now() - t0).total_seconds()
    found = sum(1 for r in rows if r.get("Precio Internet"))
    _print(f"[{datetime.now():%H:%M:%S}] {len(rows)} filas · {found} con precio · {dt:.0f}s")

    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = THIS_DIR / f"falabella_mk6_{ts}.xlsx"
    write_output(rows, str(output_path))
    _print(f"✓ Excel: {output_path}")
    return output_path


def main(argv=None):
    import argparse
    ap = argparse.ArgumentParser(description="Falabella MK6 — busca SKUs en Falabella.")
    ap.add_argument("input", help="Excel/CSV con columna SKU (y opcionales SKU Easy, Desc.)")
    ap.add_argument("--zones", nargs="*", default=None,
                    help="IDs de zona Falabella (ej: E502 E510). Default: solo E502 Kennedy.")
    ap.add_argument("--no-screenshot", action="store_true",
                    help="No capturar screenshot de cards.")
    ap.add_argument("--output", default=None, help="Ruta del Excel de salida.")
    ap.add_argument("--no-headless", action="store_true", help="Mostrar el browser.")
    args = ap.parse_args(argv)
    asyncio.run(_amain(
        args.input, zone_ids=args.zones,
        screenshot=not args.no_screenshot,
        output_path=args.output, headless=not args.no_headless,
    ))

# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
"""Stock por tienda de Sodimac — módulo ÚNICO compartido (MK7 + Maestra Sección).

Sodimac corre sobre la infraestructura de Falabella y expone una **API pública
JSON sin autenticación** que devuelve, para un SKU, las tiendas con disponibilidad
y las **unidades exactas** en cada una:

    GET https://www.falabella.com/s/geo/v1/stores/cl
        ?offeringId={skuId}&sellerId={sellerId}&latitude={lat}&longitude={lon}

Respuesta (verificada en vivo 2026-07-20):
    {"offeringId":"110100582","sellerId":"SODIMAC_CHILE","stores":[
       {"id":"96","storeName":"SODIMAC HOMECENTER CERRILLOS","distance":2.11,
        "address":{"addressLine1":"...","district":"CERRILLOS","latitude":-33.519,...},
        "stockQuantity":{"number":25,"units":"C/U"}}, ...]}

CLAVES / LIMITACIONES (medidas en vivo):
- Es **geo-acotada por radio** (~19 km): una llamada devuelve sólo las tiendas
  cercanas al lat/lon. Para cobertura nacional se consultan varios ANCHORS.
- Sólo lista tiendas **con stock** del SKU (las que devuelve traen unidades > 0).
  "0 tiendas" = sin stock cerca, NO es error.
- Requiere el `skuId` correcto y el `sellerId` que corresponda: los productos de
  marketplace NO son `SODIMAC_CHILE` (con el seller equivocado devuelve vacío).
- Coste: 1 request por (SKU × ancla). ~9 req/s con 8 workers → por eso el toggle
  en las UIs viene APAGADO por defecto y advierte que ralentiza el proceso.

Los ANCHORS cubren las 56 tiendas Sodimac del país (cosechadas desde esta misma
API el 2026-07-20: Arica → Punta Arenas, con 5 puntos en el Gran Santiago porque
el radio no alcanza a cubrir toda la RM desde un solo punto).
"""
from __future__ import annotations

import json
import ssl
import urllib.request
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

API_URL = "https://www.falabella.com/s/geo/v1/stores/cl"
DEFAULT_SELLER = "SODIMAC_CHILE"

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
_CTX = ssl.create_default_context()
_CTX.check_hostname = False
_CTX.verify_mode = ssl.CERT_NONE

# ── Anclas = LAS 42 TIENDAS DE EASY ─────────────────────────────────────────
# El proyecto entero habla en tiendas Easy (las columnas "Tienda"/"Nombre Tienda"
# de TODOS los outputs son el id/nombre Easy, ver ALL_STORES en maestra_sodimac).
# Por eso el stock se consulta y se reporta POR TIENDA EASY: para cada una se
# pregunta a la API con sus coordenadas y se informa el Sodimac más cercano y sus
# unidades. Así el dato queda accionable ("frente a mi tienda X, el competidor
# tiene N unidades") en vez de con la nomenclatura de tiendas de Sodimac.
# Coordenadas: las de la tienda Sodimac de la misma comuna cuando existe
# (cosechadas de esta API el 2026-07-20); centroide de comuna en el resto.
EASY_STORES = [
    ("E619", "Arica", "Arica", -18.4606, -70.2961),
    ("E534", "Antofagasta", "Antofagasta", -23.6482, -70.4031),
    ("E614", "Calama", "Calama", -22.4480, -68.9220),
    ("E760", "Copiapó", "Copiapó", -27.3734, -70.3356),
    ("E521", "La Serena", "La Serena", -29.9035, -71.2582),
    ("E646", "Quillota", "Quillota", -32.8800, -71.2500),
    ("E900", "Los Andes", "Los Andes", -32.8330, -70.6000),
    ("E508", "Viña del Mar", "Viña del Mar", -33.0070, -71.5453),
    ("E520", "Valparaíso", "Valparaíso", -33.0458, -71.6097),
    ("E781", "El Belloto", "Quilpué", -33.0453, -71.4169),
    ("E988", "Chicureo", "Colina", -33.2800, -70.6700),
    ("E655", "Quilicura", "Quilicura", -33.3666, -70.7280),
    ("E514", "La Dehesa", "Lo Barnechea", -33.3718, -70.5170),
    ("E502", "Kennedy", "Las Condes", -33.4015, -70.5750),
    ("E511", "Costanera", "Providencia", -33.4250, -70.6170),
    ("E512", "La Reina", "La Reina", -33.4626, -70.5448),
    ("E775", "Portal Ñuñoa", "Ñuñoa", -33.4626, -70.5757),
    ("E503", "Maipú", "Maipú", -33.4823, -70.7489),
    ("E643", "Ochagavía", "Pedro Aguirre Cerda", -33.4900, -70.6700),
    ("E518", "Quilín", "Peñalolén", -33.4991, -70.5544),
    ("E513", "El Llano", "San Miguel", -33.5040, -70.6550),
    ("E522", "Cerrillos", "Cerrillos", -33.5190, -70.7150),
    ("E510", "Florida", "La Florida", -33.5197, -70.5945),
    ("E874", "Santa Amalia", "La Florida", -33.5300, -70.5800),
    ("E659", "Puente Alto", "Puente Alto", -33.6033, -70.5766),
    ("E843", "San Bernardo", "San Bernardo", -33.6323, -70.7075),
    ("E504", "Rancagua", "Rancagua", -34.1730, -70.7180),
    ("E592", "Curicó", "Curicó", -34.9800, -71.2400),
    ("E591", "Talca", "Talca", -35.4322, -71.6321),
    ("E524", "Linares", "Linares", -35.8500, -71.6000),
    ("E525", "Chillán", "Chillán", -36.5994, -72.0993),
    ("E633", "Bio Bio", "Hualpén", -36.7979, -73.0702),
    ("E990", "Chiguayante", "Chiguayante", -36.9200, -73.0300),
    ("E983", "Coronel", "Coronel", -37.0300, -73.1400),
    ("E529", "Los Ángeles", "Los Ángeles", -37.4674, -72.3355),
    ("E830", "Villarrica", "Villarrica", -39.2800, -72.2300),
    ("E506", "Portal Temuco", "Temuco", -38.7440, -72.6092),
    ("E517", "Temuco", "Temuco", -38.7429, -72.6412),
    ("E744", "La Unión", "La Unión", -40.2900, -73.0800),
    ("E585", "Osorno", "Osorno", -40.5874, -73.1037),
    ("E748", "Portal Osorno", "Osorno", -40.5750, -73.1300),
    ("E507", "Puerto Montt", "Puerto Montt", -41.4619, -72.9465),
]

# Sólo Región Metropolitana (para runs rápidos): 15 tiendas Easy.
_RM_COMUNAS = {"Colina", "Quilicura", "Lo Barnechea", "Las Condes", "Providencia",
               "La Reina", "Ñuñoa", "Maipú", "Pedro Aguirre Cerda", "Peñalolén",
               "San Miguel", "Cerrillos", "La Florida", "Puente Alto", "San Bernardo"}
EASY_STORES_RM = [s for s in EASY_STORES if s[2] in _RM_COMUNAS]

STOCK_COLS = ["SKU", "Descripción Producto", "Tienda", "Nombre Tienda", "Comuna",
              "Sodimac más cercano", "Distancia (km)", "Unidades", "Unidades en la zona"]
RESUMEN_COLS = ["SKU", "Descripción Producto", "Stock Total (país)",
                "Tiendas Easy con competidor con stock"]


def _fetch(sku, lat, lon, seller_id, *, retries=3, timeout=25):
    url = (f"{API_URL}?offeringId={sku}&sellerId={seller_id}"
           f"&latitude={lat}&longitude={lon}")
    for _ in range(retries):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": _UA, "Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
                return (json.loads(r.read()) or {}).get("stores") or []
        except Exception:  # noqa: BLE001
            import time as _t
            _t.sleep(0.4)
    return []


def stock_for_skus(skus, *, seller_id=DEFAULT_SELLER, stores=None,
                   workers=8, progress_cb=None):
    """Stock por SKU referido a las tiendas EASY.

    Devuelve {sku: {"by_easy": {easy_id: {...}}, "unique": {sodimac_id: unidades}}}.
    Para cada (SKU × tienda Easy) consulta la API con las coordenadas de esa
    tienda: la respuesta trae las tiendas Sodimac con stock alrededor, ordenadas
    por distancia. Se registra el Sodimac **más cercano** (nombre, distancia,
    unidades) y el total de unidades en la zona. `unique` acumula las tiendas
    Sodimac vistas (dedup por id) para poder totalizar a nivel país sin
    duplicar por radios solapados.
    """
    stores = stores or EASY_STORES
    skus = [str(s).strip() for s in skus if str(s).strip()]
    tasks = [(s, st) for s in skus for st in stores]
    out = {s: {"by_easy": {}, "unique": {}} for s in skus}
    done = 0

    def _work(t):
        sku, st = t
        _eid, _nom, _com, lat, lon = st
        return sku, st, _fetch(sku, lat, lon, seller_id)

    with ThreadPoolExecutor(max_workers=workers) as ex:
        for fut in as_completed([ex.submit(_work, t) for t in tasks]):
            sku, st, found = fut.result()
            eid, nom, com, _lat, _lon = st
            found = sorted(found, key=lambda s: s.get("distance") or 9e9)
            zona = sum((s.get("stockQuantity") or {}).get("number") or 0 for s in found)
            for s in found:
                sid = str(s.get("id") or "")
                if sid:
                    out[sku]["unique"][sid] = (s.get("stockQuantity") or {}).get("number") or 0
            if found:
                near = found[0]
                out[sku]["by_easy"][eid] = {
                    "nombre": nom, "comuna": com,
                    "sodimac": near.get("storeName") or "",
                    "distancia": round(near.get("distance") or 0, 1),
                    "unidades": (near.get("stockQuantity") or {}).get("number") or 0,
                    "zona": zona,
                }
            done += 1
            if progress_cb:
                try:
                    progress_cb(done, len(tasks))
                except Exception:
                    pass
    return out


def build_rows(stock_map, names=None):
    """(filas_largo, filas_resumen) listas para escribir a Excel."""
    names = names or {}
    largo, resumen = [], []
    for sku, data in stock_map.items():
        desc = names.get(sku, "")
        by_easy = data.get("by_easy", {})
        for eid, d in sorted(by_easy.items(), key=lambda x: -(x[1]["unidades"] or 0)):
            largo.append({
                "SKU": sku, "Descripción Producto": desc,
                "Tienda": eid, "Nombre Tienda": d["nombre"], "Comuna": d["comuna"],
                "Sodimac más cercano": d["sodimac"], "Distancia (km)": d["distancia"],
                "Unidades": d["unidades"], "Unidades en la zona": d["zona"],
            })
        resumen.append({
            "SKU": sku, "Descripción Producto": desc,
            "Stock Total (país)": sum(data.get("unique", {}).values()),
            "Tiendas Easy con competidor con stock": len(by_easy),
        })
    resumen.sort(key=lambda r: -r["Stock Total (país)"])
    return largo, resumen


def add_stock_sheets(xlsx_path, skus, *, names=None, seller_id=DEFAULT_SELLER,
                     stores=None, workers=8, progress_cb=None):
    """Consulta el stock y AGREGA 2 hojas al Excel ya escrito.

    No toca las hojas existentes (post-proceso con openpyxl) → no hay que
    modificar los writers de MK7/Maestra. Hojas: "Stock por tienda" (formato
    largo, 1 fila por SKU×tienda) y "Resumen stock" (1 fila por SKU).
    Devuelve (n_filas_largo, n_skus_con_stock).
    """
    import openpyxl
    from engines._excel_utils import apply_clean_style

    stock_map = stock_for_skus(skus, seller_id=seller_id, stores=stores,
                               workers=workers, progress_cb=progress_cb)
    largo, resumen = build_rows(stock_map, names)

    wb = openpyxl.load_workbook(xlsx_path)
    for title, cols, rows in (("Stock por tienda", STOCK_COLS, largo),
                              ("Resumen stock", RESUMEN_COLS, resumen)):
        if title in wb.sheetnames:
            del wb[title]
        ws = wb.create_sheet(title)
        ws.append(cols)
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        try:
            apply_clean_style(ws)
        except Exception:
            pass
    wb.save(xlsx_path)
    return len(largo), sum(1 for v in stock_map.values() if v)

# Copyright (c) 2026 Carlos Cruz Errazuriz. All rights reserved.
# Proprietary - see LICENSE file. No unauthorized use, redistribution, or reverse engineering.
"""Base compartida de los engines de PCFactory.

pcfactory.cl corre sobre Modyo (shell Rails + widgets Vue) pero TODO el catálogo
se sirve desde una API REST pública en `api.pcfactory.cl` (JSON, sin auth, sin
challenge de Cloudflare). No hay DOM que scrapear ni Playwright: extracción JSON
pura, igual de limpia que el `variants[]` de Ferni o los engines Tier-A de
Competidores.

Este módulo NO depende de `requests` (usa urllib stdlib) para que los engines
sean testeables fuera de Colab. La única dependencia "pesada" es openpyxl, sólo
al escribir el Excel.

Contrato uniforme que hablan los engines pcf_*.py (para un launcher común):

  RETAILER_NAME : str
  discover_sections() -> [(seccion_nombre, [(subcat_nombre, subcat_id), ...]), ...]
  scrape_section(subcats, *, on_row=None, progress_cb=None, limit=None) -> [row, ...]
  write_excel(rows, path, *, with_images=False) -> None

`subcat_id` es el id de categoría de PCFactory (opaco para el launcher). El
backend hace roll-up: pasar el id de un nodo padre devuelve todos los productos
de sus descendientes, así que basta el id del nodo elegido.
"""
from __future__ import annotations

import json
import ssl
import time
import urllib.request
import urllib.error

# --- Hosts de la API (verificados en vivo 2026-06-04) ------------------------
API_BASE = "https://api.pcfactory.cl"
WEB_BASE = "https://www.pcfactory.cl"
ASSETS_BASE = "https://assets.pcfactory.cl"

# Árbol de categorías (familias -> subcategorías -> hojas).
MENU_URL = f"{API_BASE}/api-dex-catalog/v1/catalog/category/PCF/menu"
# Productos por categoría (paginado). pageSize tope del backend = 48.
QUERY_URL = f"{API_BASE}/pcfactory-services-catalogo/v1/catalogo/productos/query"
PAGE_SIZE = 48

# Columnas unificadas de salida (estilo Maestra Sección del proyecto).
OUTPUT_COLS = [
    "Tienda", "Sección", "Subcategoría", "Marca", "SKU", "Part Number",
    "Descripción Producto", "Precio Efectivo", "Precio Normal",
    "Precio Referencia", "% Descuento", "Stock", "Promoción", "URL",
]

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
_CTX = ssl.create_default_context()
_CTX.check_hostname = False
_CTX.verify_mode = ssl.CERT_NONE


def http_json(url, *, headers=None, timeout=30, retries=3, backoff=1.5):
    """GET `url` y parsea JSON. Reintenta ante errores transitorios.

    Devuelve el objeto parseado o lanza la última excepción tras agotar retries.
    """
    hdr = {"User-Agent": _UA, "Accept": "application/json"}
    if headers:
        hdr.update(headers)
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=hdr)
            with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
                return json.loads(r.read())
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError,
                json.JSONDecodeError) as e:
            last = e
            if attempt < retries - 1:
                time.sleep(backoff * (attempt + 1))
    raise last


def pct_discount(antes, ahora):
    """% de descuento entero (0–100) entre un precio `antes` y un `ahora`.

    Devuelve '' si no aplica (precios inválidos o sin descuento real).
    """
    try:
        a = float(antes)
        b = float(ahora)
        if a > 0 and b > 0 and b < a:
            return round(100 * (a - b) / a)
    except (TypeError, ValueError):
        pass
    return ""


def product_url(slug):
    """URL canónica de la ficha de producto a partir del slug del item."""
    return f"{WEB_BASE}/producto/{slug}" if slug else ""


def image_url(thumbnail):
    """URL absoluta de la imagen a partir del campo `thumbnail` (path relativo)."""
    if not thumbnail:
        return ""
    if thumbnail.startswith("http"):
        return thumbnail
    return f"{ASSETS_BASE}{thumbnail}"


def make_row(item, *, seccion="", subcat=""):
    """Construye una row de OUTPUT_COLS desde un item del endpoint de productos.

    Estructura del item (verificada 2026-06-04):
      id, nombre, marca, categoria{id,nombre}, stock, thumbnail, slug,
      precio{efectivo, normal, referencia, bancoEstado, promocion}
    `partNumber` NO viene en el listado (sólo en el detalle) → queda vacío aquí.
    El % Descuento se calcula referencia→efectivo (el descuento que ve el
    cliente: precio tachado vs precio transferencia); si no hay referencia, cae
    a normal→efectivo.
    """
    precio = item.get("precio") or {}
    efectivo = precio.get("efectivo")
    normal = precio.get("normal")
    referencia = precio.get("referencia")
    base = referencia if (referencia and referencia not in (0, "0")) else normal
    return {
        "Tienda": "PCFactory",
        "Sección": seccion,
        "Subcategoría": subcat or (item.get("categoria") or {}).get("nombre", ""),
        "Marca": item.get("marca") or "",
        "SKU": item.get("id") or "",
        "Part Number": "",  # sólo disponible en el detalle (pcf_detalle, futuro)
        "Descripción Producto": item.get("nombre") or "",
        "Precio Efectivo": efectivo if efectivo not in (None, "") else "",
        "Precio Normal": normal if normal not in (None, "") else "",
        "Precio Referencia": referencia if referencia not in (None, "") else "",
        "% Descuento": pct_discount(base, efectivo),
        "Stock": item.get("stock") or "",
        "Promoción": "Sí" if precio.get("promocion") else "",
        "URL": product_url(item.get("slug")),
        "_img": image_url(item.get("thumbnail")),
    }


# --- Excel -------------------------------------------------------------------
# La estética la pone el helper compartido del proyecto
# (engines/_excel_utils.apply_clean_style): cabecera celeste #2E86C1, auto-filtro,
# anchos auto y URL truncada. Así el Excel queda idéntico al de todas las demás
# herramientas del repo.


def _download_bytes(url, *, timeout=20):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": _UA})
        with urllib.request.urlopen(req, timeout=timeout, context=_CTX) as r:
            return r.read()
    except Exception:
        return None


def _embed_images(ws, rows, img_col_idx, *, cell_px=90, max_images=2000):
    """Descarga (en paralelo) y embebe `_img` en la columna `img_col_idx`.

    Defensivo: si falta PIL/openpyxl.drawing o una imagen falla, salta esa fila
    sin romper el guardado. Los CDN de PCFactory (assets.pcfactory.cl) se
    descargan directo (a diferencia de Sodimac, que bloquea IPs de Google).
    """
    try:
        import io
        from concurrent.futures import ThreadPoolExecutor
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        try:
            from PIL import Image as PILImage
        except Exception:
            PILImage = None
    except Exception:
        return

    urls = [(i, r.get("_img")) for i, r in enumerate(rows) if r.get("_img")]
    urls = urls[:max_images]
    if not urls:
        return

    blobs = {}
    with ThreadPoolExecutor(max_workers=8) as ex:
        for i, data in ex.map(lambda it: (it[0], _download_bytes(it[1])), urls):
            if data:
                blobs[i] = data

    col_letter = get_column_letter(img_col_idx)
    ws.column_dimensions[col_letter].width = cell_px / 7.0
    for i, data in blobs.items():
        try:
            bio = io.BytesIO(data)
            if PILImage is not None:
                im = PILImage.open(bio).convert("RGB")
                im.thumbnail((cell_px, cell_px))
                bio = io.BytesIO()
                im.save(bio, format="PNG")
                bio.seek(0)
            xl = XLImage(bio)
            row_excel = i + 2  # +1 header, +1 a 1-based
            ws.row_dimensions[row_excel].height = cell_px * 0.78
            ws.add_image(xl, f"{col_letter}{row_excel}")
        except Exception:
            continue


def write_excel(rows, path, *, sheet_name="Datos", columns=None, with_images=False):
    """Escribe las rows a un .xlsx con la estética unificada del proyecto.

    Usa openpyxl directo (sin pandas). `Imagen` siempre es la última columna
    (convención del proyecto) y sólo se agrega si `with_images=True`.
    """
    import openpyxl
    from engines._excel_utils import apply_clean_style

    cols = list(columns or OUTPUT_COLS)
    if with_images:
        cols = cols + ["Imagen"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])  # "Imagen" queda vacía; se embebe luego
    apply_clean_style(ws)
    if with_images:
        _embed_images(ws, rows, len(cols))  # última columna = Imagen
    wb.save(path)
